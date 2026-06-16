#!/usr/bin/env python3
"""
LAC Weekly KPI Digest
Sends a Monday morning email with key practice metrics and week-over-week deltas.
Cron: 0 7 * * 1 /home/jlobel/lac_automation/phase5/run_digest.sh
"""

import io
import json
import os
import smtplib
import sys
from datetime import date, datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

load_dotenv('/home/jlobel/lac_automation/.env')

ZOHO_EMAIL        = os.environ['ZOHO_EMAIL']
ZOHO_APP_PASSWORD = os.environ['ZOHO_APP_PASSWORD']
CREDENTIALS_PATH  = os.environ['GOOGLE_CREDENTIALS_PATH']
DIGEST_RECIPIENT  = os.getenv('DIGEST_RECIPIENT', 'jlobel@lobelaccountancy.com')
FRONTEND_URL      = os.getenv('FRONTEND_URL', 'https://lobelaccountancy.github.io')
DASHBOARD_URL     = f"{FRONTEND_URL}/lac-dashboard/"
SNAPSHOT_PATH     = '/home/jlobel/lac_automation/phase5/digest_snapshot.json'
WORKBOOK_NAME     = 'LAC Workbook.xlsx'
DRIVE_SCOPES      = ['https://www.googleapis.com/auth/drive.readonly']

TODAY = date.today()


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def fetch_workbook():
    creds   = service_account.Credentials.from_service_account_file(
                CREDENTIALS_PATH, scopes=DRIVE_SCOPES)
    svc     = build('drive', 'v3', credentials=creds)
    results = svc.files().list(
        q=f"name='{WORKBOOK_NAME}' and "
          f"mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        fields='files(id)',
    ).execute()
    files = results.get('files', [])
    if not files:
        raise RuntimeError(f"'{WORKBOOK_NAME}' not found in Drive")
    req = svc.files().get_media(fileId=files[0]['id'])
    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return load_workbook(buf, read_only=True, data_only=True)


def to_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ('%m/%d/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
    return None


# ---------------------------------------------------------------------------
# Data parsers
# ---------------------------------------------------------------------------

def parse_ar(wb):
    if 'AR Aging' not in wb.sheetnames:
        return {'total_outstanding': 0, 'overdue_amount': 0, 'buckets': {},
                'overdue_clients': [], 'open_count': 0}

    ws      = wb['AR Aging']
    buckets = {k: {'count': 0, 'amount': 0.0}
               for k in ('current', 'd0_30', 'd31_60', 'd61_90', 'd90plus')}
    overdue_clients = []
    total_out = overdue_amt = 0.0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        client      = str(row[0]).strip()
        due_raw     = row[3] if len(row) > 3 else None
        outstanding = row[6] if len(row) > 6 else None
        status      = row[8] if len(row) > 8 else None

        try:
            outstanding = float(outstanding) if outstanding else 0.0
        except (TypeError, ValueError):
            continue
        if outstanding <= 0 or status in ('Paid', 'Written Off'):
            continue

        due   = to_date(due_raw)
        dover = (TODAY - due).days if due else 0
        total_out += outstanding

        if dover <= 0:
            k = 'current'
        elif dover <= 30:
            k = 'd0_30'
        elif dover <= 60:
            k = 'd31_60'
        elif dover <= 90:
            k = 'd61_90'
        else:
            k = 'd90plus'

        buckets[k]['count']  += 1
        buckets[k]['amount'] += outstanding

        if dover > 0:
            overdue_amt += outstanding
            overdue_clients.append({'client': client, 'outstanding': outstanding, 'days': dover})

    for k in buckets:
        buckets[k]['amount'] = round(buckets[k]['amount'], 2)

    overdue_clients.sort(key=lambda x: x['days'], reverse=True)

    return {
        'total_outstanding': round(total_out, 2),
        'overdue_amount':    round(overdue_amt, 2),
        'open_count':        sum(b['count'] for b in buckets.values()),
        'buckets':           buckets,
        'overdue_clients':   overdue_clients[:8],
    }


def parse_pipeline(wb):
    if 'Engagement Pipeline' not in wb.sheetnames:
        return {'active': 0, 'by_stage': {}, 'pending_client': []}

    ws       = wb['Engagement Pipeline']
    by_stage: dict[str, int] = {}
    pending_client = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        client   = str(row[0]).strip()
        eng_type = str(row[1]).strip() if len(row) > 1 and row[1] else 'Unknown'
        stage    = str(row[3]).strip() if len(row) > 3 and row[3] else None
        due_raw  = row[11] if len(row) > 11 else None

        if not client or not stage or stage in ('Complete', 'Lost'):
            continue

        by_stage[stage] = by_stage.get(stage, 0) + 1

        s = stage.lower()
        if any(k in s for k in ('pending client', 'awaiting', 'client', 'info')):
            due = to_date(due_raw)
            pending_client.append({
                'client':   client,
                'type':     eng_type,
                'stage':    stage,
                'due_date': due.strftime('%b %d') if due else None,
            })

    return {
        'active':         sum(by_stage.values()),
        'by_stage':       dict(sorted(by_stage.items(), key=lambda x: x[1], reverse=True)),
        'pending_client': pending_client[:10],
    }


def upcoming_deadlines(wb, days=14):
    if 'Engagement Pipeline' not in wb.sheetnames:
        return []

    ws        = wb['Engagement Pipeline']
    cutoff    = TODAY + __import__('datetime').timedelta(days=days)
    deadlines = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        client   = str(row[0]).strip()
        eng_type = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        stage    = str(row[3]).strip() if len(row) > 3 and row[3] else None
        due_raw  = row[11] if len(row) > 11 else None

        if not client or not stage or stage in ('Complete', 'Lost'):
            continue

        due = to_date(due_raw)
        if not due or due < TODAY or due > cutoff:
            continue

        deadlines.append({
            'client':     client,
            'type':       eng_type,
            'stage':      stage,
            'due_date':   due,
            'days_until': (due - TODAY).days,
        })

    deadlines.sort(key=lambda x: x['due_date'])
    return deadlines[:20]


def client_health(wb):
    if 'AR Aging' not in wb.sheetnames:
        return {'needs_attention': [], 'at_risk': 0, 'healthy': 0, 'total': 0}

    ws      = wb['AR Aging']
    clients: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        client      = str(row[0]).strip()
        outstanding = row[6] if len(row) > 6 else None
        status      = row[8] if len(row) > 8 else None
        due_raw     = row[3] if len(row) > 3 else None

        try:
            outstanding = float(outstanding) if outstanding else 0.0
        except (TypeError, ValueError):
            continue
        if outstanding <= 0 or status in ('Paid', 'Written Off'):
            continue

        due   = to_date(due_raw)
        dover = (TODAY - due).days if due else 0

        if client not in clients:
            clients[client] = {'max_overdue': 0, 'total': 0.0}
        if dover > clients[client]['max_overdue']:
            clients[client]['max_overdue'] = dover
        clients[client]['total'] += outstanding

    needs_attention = []
    at_risk = healthy = 0

    for name, c in sorted(clients.items()):
        if c['max_overdue'] > 30:
            needs_attention.append({'name': name, 'days': c['max_overdue'], 'amount': round(c['total'], 2)})
        elif c['max_overdue'] > 0 or c['total'] > 10_000:
            at_risk += 1
        else:
            healthy += 1

    needs_attention.sort(key=lambda x: x['days'], reverse=True)

    return {
        'needs_attention': needs_attention,
        'at_risk':         at_risk,
        'healthy':         healthy,
        'total':           len(needs_attention) + at_risk + healthy,
    }


# ---------------------------------------------------------------------------
# Snapshot (week-over-week deltas)
# ---------------------------------------------------------------------------

def load_snapshot() -> dict:
    try:
        with open(SNAPSHOT_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_snapshot(ar, pipeline, health):
    data = {
        'generated_at':    TODAY.isoformat(),
        'total_outstanding': ar['total_outstanding'],
        'overdue_amount':    ar['overdue_amount'],
        'active_matters':    pipeline['active'],
        'needs_attention':   len(health['needs_attention']),
    }
    with open(SNAPSHOT_PATH, 'w') as f:
        json.dump(data, f, indent=2)


def delta(current, previous_key, snapshot, fmt='dollar'):
    if not snapshot or previous_key not in snapshot:
        return ''
    diff = current - snapshot[previous_key]
    if diff == 0:
        return ''
    arrow = '▲' if diff > 0 else '▼'
    color = '#C0392B' if diff > 0 else '#1A7A4A'   # up is bad for AR/overdue
    if fmt == 'dollar':
        val = f'${abs(diff):,.0f}'
    else:
        val = str(abs(int(diff)))
    return f'<span style="font-size:11px;color:{color};margin-left:4px;">{arrow}&nbsp;{val}</span>'


def delta_good(current, previous_key, snapshot, fmt='dollar'):
    """Arrow direction is reversed — up is good (e.g., healthy count)."""
    if not snapshot or previous_key not in snapshot:
        return ''
    diff = current - snapshot[previous_key]
    if diff == 0:
        return ''
    arrow = '▲' if diff > 0 else '▼'
    color = '#1A7A4A' if diff > 0 else '#C0392B'
    val = f'${abs(diff):,.0f}' if fmt == 'dollar' else str(abs(int(diff)))
    return f'<span style="font-size:11px;color:{color};margin-left:4px;">{arrow}&nbsp;{val}</span>'


# ---------------------------------------------------------------------------
# Email HTML builder
# ---------------------------------------------------------------------------

def bucket_row(label, bucket, color):
    if bucket['count'] == 0:
        return ''
    return f"""
        <tr>
          <td style="padding:6px 12px;color:{color};font-weight:500;">{label}</td>
          <td style="padding:6px 12px;text-align:right;">{bucket['count']}</td>
          <td style="padding:6px 12px;text-align:right;font-weight:600;">{fmt_dollar(bucket['amount'])}</td>
        </tr>"""


def fmt_dollar(n):
    return f'${n:,.0f}'


def build_email(ar, pipeline, deadlines, health, snapshot):
    week_of = TODAY.strftime('%B %d, %Y')
    today_str = TODAY.strftime('%A, %B %d, %Y')

    # --- KPI summary row ---
    kpi_row = f"""
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td width="33%" style="padding:16px 20px;border-right:1px solid #2E4A6A;text-align:center;">
          <div style="font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:#8BA7C4;margin-bottom:6px;">Total AR Outstanding</div>
          <div style="font-size:26px;font-weight:700;color:white;">{fmt_dollar(ar['total_outstanding'])}</div>
          <div style="font-size:12px;color:#8BA7C4;margin-top:2px;">{ar['open_count']} open invoice{'s' if ar['open_count'] != 1 else ''}
            {delta(ar['total_outstanding'], 'total_outstanding', snapshot)}</div>
        </td>
        <td width="33%" style="padding:16px 20px;border-right:1px solid #2E4A6A;text-align:center;">
          <div style="font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:#8BA7C4;margin-bottom:6px;">Overdue</div>
          <div style="font-size:26px;font-weight:700;color:{'#F87171' if ar['overdue_amount'] > 0 else 'white'};">{fmt_dollar(ar['overdue_amount'])}</div>
          <div style="font-size:12px;color:#8BA7C4;margin-top:2px;">{len(ar['overdue_clients'])} client{'s' if len(ar['overdue_clients']) != 1 else ''}
            {delta(ar['overdue_amount'], 'overdue_amount', snapshot)}</div>
        </td>
        <td width="33%" style="padding:16px 20px;text-align:center;">
          <div style="font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:#8BA7C4;margin-bottom:6px;">Active Matters</div>
          <div style="font-size:26px;font-weight:700;color:white;">{pipeline['active']}</div>
          <div style="font-size:12px;color:#8BA7C4;margin-top:2px;">{len(pipeline['by_stage'])} stage{'s' if len(pipeline['by_stage']) != 1 else ''}
            {delta_good(pipeline['active'], 'active_matters', snapshot, fmt='count')}</div>
        </td>
      </tr>
    </table>"""

    # --- AR aging buckets ---
    buckets_html = ''
    labels = [
        ('current', 'Current (not overdue)', '#1A7A4A'),
        ('d0_30',   '1–30 days overdue',     '#C9920A'),
        ('d31_60',  '31–60 days overdue',    '#C05621'),
        ('d61_90',  '61–90 days overdue',    '#B03030'),
        ('d90plus', '90+ days overdue',      '#7B1818'),
    ]
    for key, label, color in labels:
        buckets_html += bucket_row(label, ar['buckets'][key], color)

    # --- Overdue client list ---
    overdue_rows = ''
    for c in ar['overdue_clients']:
        dot_color = '#C0392B' if c['days'] > 30 else '#D97706'
        overdue_rows += f"""
        <tr>
          <td style="padding:7px 12px;">
            <span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{dot_color};margin-right:8px;"></span>
            {c['client']}
          </td>
          <td style="padding:7px 12px;text-align:right;font-weight:600;">{fmt_dollar(c['outstanding'])}</td>
          <td style="padding:7px 12px;text-align:right;color:#C0392B;">{c['days']}d overdue</td>
        </tr>"""

    # --- Pipeline stages ---
    stage_rows = ''
    for stage, count in pipeline['by_stage'].items():
        stage_rows += f"""
        <tr>
          <td style="padding:6px 12px;">{stage}</td>
          <td style="padding:6px 12px;text-align:right;font-weight:600;">{count}</td>
        </tr>"""

    # --- Pending client follow-ups ---
    pending_html = ''
    if pipeline['pending_client']:
        items = ''.join(
            f'<li style="padding:4px 0;color:#92400E;">'
            f'<strong>{p["client"]}</strong> — {p["type"]}'
            f'{" (" + p["due_date"] + ")" if p["due_date"] else ""}'
            f'</li>'
            for p in pipeline['pending_client']
        )
        pending_html = f"""
        <tr>
          <td colspan="2" style="padding:12px;background:#FEF3C7;border-top:1px solid #E0E6ED;">
            <div style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:#92400E;margin-bottom:8px;">
              Needs Your Follow-up ({len(pipeline['pending_client'])})
            </div>
            <ul style="margin:0;padding-left:18px;font-size:13px;">{items}</ul>
          </td>
        </tr>"""

    # --- Deadlines ---
    deadline_rows = ''
    for d in deadlines:
        urgency_color = '#C0392B' if d['days_until'] <= 7 else '#1B2A3F'
        due_str = d['due_date'].strftime('%b %d')
        days_str = 'Today' if d['days_until'] == 0 else f"{d['days_until']}d"
        deadline_rows += f"""
        <tr>
          <td style="padding:8px 12px;">
            <strong style="color:{urgency_color};">{d['client']}</strong>
            <span style="color:#5A6B7C;"> — {d['type']}</span>
          </td>
          <td style="padding:8px 12px;color:#5A6B7C;">{d['stage']}</td>
          <td style="padding:8px 12px;text-align:right;">
            <strong style="color:{urgency_color};">{due_str}</strong>
            <span style="font-size:11px;color:{urgency_color};margin-left:4px;">({days_str})</span>
          </td>
        </tr>"""

    if not deadline_rows:
        deadline_rows = '<tr><td colspan="3" style="padding:12px;color:#8BA7C4;">No deadlines in the next 14 days.</td></tr>'

    # --- Client health ---
    attention_count = len(health['needs_attention'])
    attention_list = ''
    if health['needs_attention']:
        items = ''.join(
            f'<li style="padding:3px 0;">'
            f'<strong>{c["name"]}</strong> — {fmt_dollar(c["amount"])} · {c["days"]}d overdue'
            f'</li>'
            for c in health['needs_attention']
        )
        attention_list = f'<ul style="margin:10px 0 0;padding-left:18px;font-size:13px;color:#C0392B;">{items}</ul>'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#F0F2F5;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" bgcolor="#F0F2F5">
<tr><td align="center" style="padding:24px 16px;">

  <table width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%;">

    <!-- Header -->
    <tr>
      <td bgcolor="#1B2A3F" style="padding:24px 28px;border-radius:10px 10px 0 0;">
        <table width="100%" cellpadding="0" cellspacing="0">
          <tr>
            <td>
              <div style="font-size:18px;font-weight:700;color:white;">Lobel Accountancy Corporation</div>
              <div style="font-size:13px;color:#8BA7C4;margin-top:3px;">Weekly KPI Digest</div>
            </td>
            <td align="right" style="vertical-align:top;">
              <div style="font-size:13px;color:#8BA7C4;">{today_str}</div>
            </td>
          </tr>
        </table>
      </td>
    </tr>

    <!-- KPI summary row -->
    <tr>
      <td bgcolor="#243651" style="border-bottom:1px solid #2E4A6A;">
        {kpi_row}
      </td>
    </tr>

    <!-- ── AR AGING ── -->
    <tr>
      <td bgcolor="#FFFFFF" style="padding:0;border-left:1px solid #E0E6ED;border-right:1px solid #E0E6ED;">
        <table width="100%" cellpadding="0" cellspacing="0">
          <tr>
            <td style="padding:14px 16px;background:#FAFBFC;border-bottom:1px solid #E0E6ED;">
              <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#5A6B7C;">AR Aging</span>
            </td>
          </tr>
          <tr>
            <td>
              <table width="100%" cellpadding="0" cellspacing="0" style="font-size:13px;">
                <tr style="background:#FAFBFC;">
                  <th style="padding:6px 12px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.04em;color:#8BA7C4;border-bottom:1px solid #E0E6ED;">Bucket</th>
                  <th style="padding:6px 12px;text-align:right;font-size:11px;text-transform:uppercase;letter-spacing:.04em;color:#8BA7C4;border-bottom:1px solid #E0E6ED;">Invoices</th>
                  <th style="padding:6px 12px;text-align:right;font-size:11px;text-transform:uppercase;letter-spacing:.04em;color:#8BA7C4;border-bottom:1px solid #E0E6ED;">Amount</th>
                </tr>
                {buckets_html}
              </table>
            </td>
          </tr>
          {'<tr><td style="padding:0 0 12px;"><table width=\'100%\' cellpadding=\'0\' cellspacing=\'0\' style=\'font-size:13px;\'>' + overdue_rows + '</table></td></tr>' if overdue_rows else ''}
        </table>
      </td>
    </tr>

    <!-- ── PIPELINE ── -->
    <tr>
      <td bgcolor="#FFFFFF" style="padding:0;border-left:1px solid #E0E6ED;border-right:1px solid #E0E6ED;border-top:8px solid #F0F2F5;">
        <table width="100%" cellpadding="0" cellspacing="0">
          <tr>
            <td style="padding:14px 16px;background:#FAFBFC;border-bottom:1px solid #E0E6ED;">
              <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#5A6B7C;">Pipeline — {pipeline['active']} Active Matters</span>
            </td>
          </tr>
          <tr>
            <td>
              <table width="100%" cellpadding="0" cellspacing="0" style="font-size:13px;">
                <tr style="background:#FAFBFC;">
                  <th style="padding:6px 12px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.04em;color:#8BA7C4;border-bottom:1px solid #E0E6ED;">Stage</th>
                  <th style="padding:6px 12px;text-align:right;font-size:11px;text-transform:uppercase;letter-spacing:.04em;color:#8BA7C4;border-bottom:1px solid #E0E6ED;">Count</th>
                </tr>
                {stage_rows}
                {pending_html}
              </table>
            </td>
          </tr>
        </table>
      </td>
    </tr>

    <!-- ── DEADLINES ── -->
    <tr>
      <td bgcolor="#FFFFFF" style="padding:0;border-left:1px solid #E0E6ED;border-right:1px solid #E0E6ED;border-top:8px solid #F0F2F5;">
        <table width="100%" cellpadding="0" cellspacing="0">
          <tr>
            <td style="padding:14px 16px;background:#FAFBFC;border-bottom:1px solid #E0E6ED;">
              <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#5A6B7C;">Upcoming Deadlines — Next 14 Days</span>
            </td>
          </tr>
          <tr>
            <td>
              <table width="100%" cellpadding="0" cellspacing="0" style="font-size:13px;">
                {deadline_rows}
              </table>
            </td>
          </tr>
        </table>
      </td>
    </tr>

    <!-- ── CLIENT HEALTH ── -->
    <tr>
      <td bgcolor="#FFFFFF" style="padding:0;border-left:1px solid #E0E6ED;border-right:1px solid #E0E6ED;border-top:8px solid #F0F2F5;">
        <table width="100%" cellpadding="0" cellspacing="0">
          <tr>
            <td style="padding:14px 16px;background:#FAFBFC;border-bottom:1px solid #E0E6ED;">
              <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#5A6B7C;">Client Health — {health['total']} Clients</span>
              {delta(attention_count, 'needs_attention', snapshot, fmt='count')}
            </td>
          </tr>
          <tr>
            <td style="padding:16px;">
              <table cellpadding="0" cellspacing="0" style="font-size:13px;width:100%;">
                <tr>
                  <td style="padding:6px 16px 6px 0;">
                    <span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#C0392B;margin-right:6px;"></span>
                    <strong>Needs Attention:</strong> {attention_count}
                    {'&nbsp;&nbsp;<span style="font-size:11px;color:#C0392B;">(AR 30+ days overdue)</span>' if attention_count > 0 else ''}
                  </td>
                  <td style="padding:6px 16px;">
                    <span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#D4880A;margin-right:6px;"></span>
                    <strong>At Risk:</strong> {health['at_risk']}
                  </td>
                  <td style="padding:6px 0;">
                    <span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#1A7A4A;margin-right:6px;"></span>
                    <strong>Healthy:</strong> {health['healthy']}
                  </td>
                </tr>
              </table>
              {attention_list}
            </td>
          </tr>
        </table>
      </td>
    </tr>

    <!-- Footer -->
    <tr>
      <td bgcolor="#1B2A3F" style="padding:20px 28px;border-radius:0 0 10px 10px;border-left:1px solid #2E4A6A;border-right:1px solid #2E4A6A;border-bottom:1px solid #2E4A6A;">
        <table width="100%" cellpadding="0" cellspacing="0">
          <tr>
            <td>
              <a href="{DASHBOARD_URL}" style="display:inline-block;background:rgba(255,255,255,.12);color:white;
                 padding:10px 20px;border-radius:6px;text-decoration:none;font-size:13px;font-weight:600;">
                Open Dashboard →
              </a>
            </td>
            <td align="right" style="font-size:12px;color:#8BA7C4;line-height:1.7;">
              Lobel Accountancy Corporation<br>
              Generated {TODAY.strftime('%B %d, %Y')}
            </td>
          </tr>
        </table>
      </td>
    </tr>

  </table>
</td></tr>
</table>
</body></html>"""

    return html


# ---------------------------------------------------------------------------
# Send
# ---------------------------------------------------------------------------

def send_email(html: str, subject: str):
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From']    = ZOHO_EMAIL
    msg['To']      = DIGEST_RECIPIENT
    msg.attach(MIMEText(html, 'html'))
    with smtplib.SMTP_SSL('smtp.zoho.com', 465) as server:
        server.login(ZOHO_EMAIL, ZOHO_APP_PASSWORD)
        server.sendmail(ZOHO_EMAIL, DIGEST_RECIPIENT, msg.as_string())


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    log(f"LAC Weekly KPI Digest — {TODAY.strftime('%A, %B %d, %Y')}")

    log("Fetching workbook from Drive…")
    wb = fetch_workbook()
    log("Workbook loaded.")

    log("Parsing data…")
    ar        = parse_ar(wb)
    pipeline  = parse_pipeline(wb)
    deadlines = upcoming_deadlines(wb, days=14)
    health    = client_health(wb)
    snapshot  = load_snapshot()

    log(f"  AR outstanding: ${ar['total_outstanding']:,.0f}  |  overdue: ${ar['overdue_amount']:,.0f}")
    log(f"  Active matters: {pipeline['active']}  |  deadlines next 14d: {len(deadlines)}")
    log(f"  Client health — attention: {len(health['needs_attention'])}  at_risk: {health['at_risk']}  healthy: {health['healthy']}")

    log("Building email…")
    html    = build_email(ar, pipeline, deadlines, health, snapshot)
    subject = f"LAC Weekly KPI — {TODAY.strftime('%B %d, %Y')}"

    log(f"Sending to {DIGEST_RECIPIENT}…")
    send_email(html, subject)
    log("Email sent.")

    save_snapshot(ar, pipeline, health)
    log("Snapshot saved.")


if __name__ == '__main__':
    try:
        main()
    except Exception as exc:
        log(f"FATAL: {exc}")
        sys.exit(1)
