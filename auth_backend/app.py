import io
import imaplib
import threading
import ipaddress
import json
import os
import jwt
import re
import smtplib
import time
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
import email as _email_pkg
import email.utils as _email_utils
from email.header import decode_header as _decode_mime_header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps

import requests
import pdfplumber
from flask import Flask, request, jsonify, send_from_directory, redirect, has_request_context, Response
from flask_cors import CORS
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook as load_xlsx

load_dotenv('/home/jlobel/lac_automation/.env')

app = Flask(__name__)

# --- Config ---
SECRET_KEY      = os.environ['AUTH_SECRET_KEY']
JWT_SECRET      = os.environ['JWT_SECRET_KEY']
ZOHO_EMAIL      = os.environ['ZOHO_EMAIL']
ZOHO_PASSWORD   = os.environ['ZOHO_APP_PASSWORD']
FRONTEND_URL    = os.getenv('FRONTEND_URL', 'https://lobelaccountancy.github.io')
AUTH_URL        = os.getenv('AUTH_URL', 'https://auth.lobelaccountancy.com')
PORTAL_URL      = os.getenv('PORTAL_URL', 'https://clients.lobelaccountancy.com')
DASHBOARD_URL   = os.getenv('DASHBOARD_URL', 'https://dashboard.lobelaccountancy.com')
PORTAL_DIR      = os.path.join(os.path.dirname(__file__),
                               '../docs/portal')

ALLOWED_EMAILS = {
    e.strip().lower()
    for e in os.getenv('ALLOWED_EMAILS', 'jlobel@lobelaccountancy.com').split(',')
    if e.strip()
}

MAGIC_LINK_EXPIRY = 900   # 15 minutes
JWT_EXPIRY_DAYS   = 30

# DocuSeal e-signature
DOCUSEAL_URL     = os.getenv('DOCUSEAL_URL', 'https://sign.lobelaccountancy.com')
DOCUSEAL_API_KEY = os.getenv('DOCUSEAL_API_KEY', '')

CORS(app, origins=[
    FRONTEND_URL,
    PORTAL_URL,
    'https://dashboard.lobelaccountancy.com',
    'https://ai.lobelaccountancy.com',
    'http://localhost:3000',
    'http://127.0.0.1:3000',
])

serializer = URLSafeTimedSerializer(SECRET_KEY)

# Single-use token tracking: {token: used_at}
# Pruned on each verify call to stay bounded.
_used_tokens: dict[str, float] = {}

# Rate limiting: email -> list of request timestamps
_rate_limits: dict[str, list] = defaultdict(list)


def _prune_used_tokens():
    cutoff = time.time() - MAGIC_LINK_EXPIRY
    expired = [t for t, ts in _used_tokens.items() if ts < cutoff]
    for t in expired:
        del _used_tokens[t]


def _rate_limit_ok(email: str, max_req: int = 3, window: int = 3600) -> bool:
    now = time.time()
    recent = [t for t in _rate_limits[email] if now - t < window]
    if len(recent) >= max_req:
        return False
    recent.append(now)
    _rate_limits[email] = recent
    return True


def _send_magic_link(email: str, token: str):
    magic_url = f"{FRONTEND_URL}/auth.html?token={token}"
    expires_str = datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p UTC')

    body = f"""
<html><body style="font-family:Calibri,sans-serif;color:#1B2A3F;max-width:600px;margin:0 auto;">
  <div style="background:#1B2A3F;padding:24px;border-radius:8px 8px 0 0;">
    <h1 style="color:white;margin:0;font-size:22px;">Lobel Accountancy Corporation</h1>
    <p style="color:#8BA7C4;margin:4px 0 0 0;font-size:13px;">Practice Management System</p>
  </div>
  <div style="background:#f8f9fa;padding:32px;border-radius:0 0 8px 8px;border:1px solid #e0e0e0;">
    <h2 style="color:#1B2A3F;margin-top:0;">Your sign-in link</h2>
    <p>Click the button below to sign in to the LAC dashboard.
       This link expires in 15 minutes and can only be used once.</p>
    <p>
      <a href="{magic_url}"
         style="display:inline-block;background:#1B2A3F;color:white;
                padding:14px 28px;border-radius:6px;text-decoration:none;
                font-size:15px;font-weight:600;">
        Sign in to LAC Dashboard &rarr;
      </a>
    </p>
    <p style="margin-top:24px;font-size:12px;color:#999;">
      If you didn&rsquo;t request this link, you can safely ignore this email.<br>
      Link expires: {expires_str}
    </p>
  </div>
</body></html>
"""

    msg = MIMEMultipart('alternative')
    msg['Subject'] = 'LAC Dashboard — Sign-in link'
    msg['From']    = ZOHO_EMAIL
    msg['To']      = email
    msg.attach(MIMEText(body, 'html'))

    with smtplib.SMTP_SSL('smtp.zoho.com', 465) as server:
        server.login(ZOHO_EMAIL, ZOHO_PASSWORD)
        server.sendmail(ZOHO_EMAIL, email, msg.as_string())


def require_jwt(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '):
            return jsonify({'error': 'Missing authorization header'}), 401
        token = auth[7:]
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
            request.user_email  = payload['email']
            request.client_name = '__staff__'
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Session expired — please sign in again'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'service': 'lac-auth'})


@app.route('/auth/request', methods=['POST'])
def auth_request():
    data  = request.get_json(silent=True) or {}
    email = (data.get('email') or '').strip().lower()

    if not email:
        return jsonify({'error': 'Email is required'}), 400

    # Always return the same message — don't reveal whether email is allowed
    generic = {'message': 'If this email is authorized, a sign-in link has been sent.'}

    if email not in ALLOWED_EMAILS:
        return jsonify(generic), 200

    if not _rate_limit_ok(email):
        return jsonify({'error': 'Too many requests — try again in an hour.'}), 429

    token = serializer.dumps(email, salt='magic-link')

    try:
        _send_magic_link(email, token)
    except Exception as exc:
        app.logger.error('Magic link email failed for %s: %s', email, exc)
        return jsonify({'error': 'Failed to send email. Check server logs.'}), 500

    return jsonify(generic), 200


@app.route('/auth/verify', methods=['POST'])
def auth_verify():
    data  = request.get_json(silent=True) or {}
    token = (data.get('token') or '').strip()

    if not token:
        return jsonify({'error': 'Token is required'}), 400

    _prune_used_tokens()

    if token in _used_tokens:
        return jsonify({'error': 'This link has already been used. Request a new one.'}), 401

    try:
        email = serializer.loads(token, salt='magic-link', max_age=MAGIC_LINK_EXPIRY)
    except SignatureExpired:
        return jsonify({'error': 'Link expired. Request a new one.'}), 401
    except BadSignature:
        return jsonify({'error': 'Invalid or tampered token'}), 401

    _used_tokens[token] = time.time()

    now = int(time.time())
    jwt_payload = {
        'email': email,
        'iat': now,
        'exp': now + (JWT_EXPIRY_DAYS * 86400),
        'iss': 'lac-auth',
    }
    jwt_token = jwt.encode(jwt_payload, JWT_SECRET, algorithm='HS256')

    return jsonify({
        'jwt':             jwt_token,
        'email':           email,
        'expires_in_days': JWT_EXPIRY_DAYS,
    })


_LOCAL_NETWORKS = [
    ipaddress.ip_network('10.0.0.0/8'),       # RFC 1918
    ipaddress.ip_network('172.16.0.0/12'),     # RFC 1918
    ipaddress.ip_network('192.168.0.0/16'),    # RFC 1918
    ipaddress.ip_network('100.64.0.0/10'),     # CGNAT / Crostini / Tailscale
]


def _is_local_ip(addr):
    """Return True for private LAN addresses — excluding loopback so Cloudflare tunnel IPs are rejected."""
    try:
        ip = ipaddress.ip_address(addr)
        return any(ip in net for net in _LOCAL_NETWORKS)
    except ValueError:
        return False


@app.route('/auth/local-login')
def auth_local_login():
    """Issue a JWT without email verification for requests arriving from the local LAN."""
    if not _is_local_ip(request.remote_addr):
        return 'Local network login is only available from the home LAN.', 403

    owner_email = next(iter(sorted(ALLOWED_EMAILS)), '')
    if not owner_email:
        return 'No authorized email configured.', 503

    now = int(time.time())
    payload = {
        'email': owner_email,
        'iat': now,
        'exp': now + (JWT_EXPIRY_DAYS * 86400),
        'iss': 'lac-auth',
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm='HS256')
    return redirect(f'{DASHBOARD_URL}/auth.html#jwt={token}')


@app.route('/auth/me', methods=['GET'])
@require_jwt
def auth_me():
    return jsonify({'email': request.user_email, 'authenticated': True})


# ---------------------------------------------------------------------------
# Data: Morning Briefing
# ---------------------------------------------------------------------------

CREDENTIALS_PATH  = os.getenv('GOOGLE_CREDENTIALS_PATH')
WORKBOOK_NAME     = 'LAC Workbook.xlsx'
_DRIVE_SCOPES     = ['https://www.googleapis.com/auth/drive.readonly']
_DRIVE_RW_SCOPES  = ['https://www.googleapis.com/auth/drive']

_wb_cache: dict = {'wb': None, 'fetched_at': 0.0}
_WB_TTL = 600  # 10-minute cache
_wb_lock = threading.Lock()
_wb_file_id_cache: dict = {'id': None}  # permanent cache — file ID never changes


def _drive_service():
    creds = service_account.Credentials.from_service_account_file(
        CREDENTIALS_PATH, scopes=_DRIVE_SCOPES)
    return build('drive', 'v3', credentials=creds)


def _fetch_workbook():
    svc = _drive_service()
    file_id = _wb_file_id_cache.get('id')
    if not file_id:
        results = svc.files().list(
            q=f"name='{WORKBOOK_NAME}' and "
              f"mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
            fields='files(id)',
        ).execute()
        files = results.get('files', [])
        if not files:
            raise RuntimeError(f"'{WORKBOOK_NAME}' not found in Drive")
        file_id = files[0]['id']
        _wb_file_id_cache['id'] = file_id

    req = svc.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return load_xlsx(buf, read_only=True, data_only=True)


def _workbook():
    now = time.time()
    if _wb_cache['wb'] is not None and now - _wb_cache['fetched_at'] <= _WB_TTL:
        return _wb_cache['wb']
    with _wb_lock:
        # Re-check inside lock — another thread may have refreshed while we waited
        now = time.time()
        if _wb_cache['wb'] is None or now - _wb_cache['fetched_at'] > _WB_TTL:
            _wb_cache['wb']         = _fetch_workbook()
            _wb_cache['fetched_at'] = now
    return _wb_cache['wb']


# ---------------------------------------------------------------------------
# Workbook write helpers — download writable copy, save, upload to Drive
# ---------------------------------------------------------------------------

_WB_LOCAL_PATH = '/tmp/lac_workbook_rw.xlsx'

def _drive_rw_service():
    creds = service_account.Credentials.from_service_account_file(
        CREDENTIALS_PATH, scopes=_DRIVE_RW_SCOPES)
    return build('drive', 'v3', credentials=creds)

def _wb_file_id(svc):
    results = svc.files().list(
        q=f"name='{WORKBOOK_NAME}' and "
          "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        fields='files(id)',
    ).execute()
    files = results.get('files', [])
    if not files:
        raise RuntimeError(f"'{WORKBOOK_NAME}' not found in Drive")
    return files[0]['id']

def _wb_download_writable():
    """Download workbook from Drive as a writable openpyxl workbook.
    Returns (wb, file_id, svc) so caller can upload after editing."""
    from googleapiclient.http import MediaIoBaseDownload as _DL
    svc     = _drive_rw_service()
    fid     = _wb_file_id(svc)
    req     = svc.files().get_media(fileId=fid)
    buf     = io.BytesIO()
    dl      = _DL(buf, req)
    done    = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    with open(_WB_LOCAL_PATH, 'wb') as f:
        f.write(buf.read())
    wb = load_xlsx(_WB_LOCAL_PATH, data_only=True)
    return wb, fid, svc

def _wb_save_and_upload(wb, fid, svc):
    """Save workbook and upload — delegates to the main _wb_upload helper defined below."""
    _wb_upload(wb, fid, svc)

# ---------------------------------------------------------------------------
# Engagement Pipeline column map (same as invoice_automation.py)
# ---------------------------------------------------------------------------
_EP_COL = {
    'client': 0, 'client_num': 1, 'stage': 2, 'eng_type': 3,
    'billing_type': 4, 'entity_type': 5, 'fee': 6, 'status': 7,
    'start_date': 8, 'issue_date': 9, 'fye': 10,
    'client_title': 11, 'email': 12, 'address': 13,
}
# AR Aging column map
_AR_COL = {
    'client': 0, 'client_num': 1, 'invoice': 2, 'inv_date': 3,
    'due_date': 4, 'service': 5, 'amount': 6, 'paid': 7,
    'outstanding': 8, 'days': 9, 'status': 10,
    'email': 11, 'address': 12, 'reminder': 13,
}
_EP_DATA_START = 3   # Engagement Pipeline: row 2 = headers, row 3+ = data
_AR_DATA_START = 13  # AR Aging: row 12 = headers, row 13+ = data


def _to_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ('%m/%d/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
    return None


def _parse_ar(wb):
    if 'AR Aging' not in wb.sheetnames:
        return {'total_outstanding': 0, 'overdue_count': 0, 'overdue_amount': 0, 'items': []}

    ws  = wb['AR Aging']
    today = date.today()
    items = []
    total_out = overdue_amt = 0

    for row in ws.iter_rows(min_row=AR_DATA_START, values_only=True):
        if len(row) <= AR_C_OUTSTANDING:
            continue
        client      = row[AR_C_CLIENT]
        invoice     = row[AR_C_INVOICE]
        due_raw     = row[AR_C_DUE_DATE]
        amount      = row[AR_C_AMOUNT]
        outstanding = row[AR_C_OUTSTANDING]
        status      = row[AR_C_STATUS]      if len(row) > AR_C_STATUS  else None

        try:
            amount      = float(amount)      if amount      is not None else 0.0
            outstanding = float(outstanding) if outstanding is not None else 0.0
        except (TypeError, ValueError):
            continue
        if not client or outstanding <= 0:
            continue
        if status in ('Paid', 'Written Off', 'Void'):
            continue

        due = _to_date(due_raw)
        days_over = (today - due).days if due else 0
        total_out += outstanding
        if days_over > 0:
            overdue_amt += outstanding

        items.append({
            'client':      str(client),
            'invoice':     str(invoice) if invoice else '',
            'due_date':    due.isoformat() if due else None,
            'amount':      round(amount, 2),
            'outstanding': round(outstanding, 2),
            'days_overdue': days_over,
            'status':      str(status) if status else 'Open',
        })

    items.sort(key=lambda x: x['days_overdue'], reverse=True)
    return {
        'total_outstanding': round(total_out, 2),
        'overdue_count':     sum(1 for i in items if i['days_overdue'] > 0),
        'overdue_amount':    round(overdue_amt, 2),
        'items':             items,
    }


def _parse_pipeline(wb):
    if 'Engagement Pipeline' not in wb.sheetnames:
        return {'total_active': 0, 'by_stage': {}, 'upcoming': []}

    ws    = wb['Engagement Pipeline']
    today = date.today()
    by_stage: dict[str, int] = {}
    upcoming = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 3:
            continue
        client   = row[0]
        stage    = row[2]
        eng_type = row[3] if len(row) > 3 else None
        due_raw  = row[9] if len(row) > 9 else None

        if not client or stage in ('Complete', 'Lost', None):
            continue

        by_stage[str(stage)] = by_stage.get(str(stage), 0) + 1
        due = _to_date(due_raw)
        days_until = (due - today).days if due else None

        if days_until is not None and days_until <= 30:
            upcoming.append({
                'client':     str(client),
                'type':       str(eng_type) if eng_type else '',
                'stage':      str(stage),
                'due_date':   due.isoformat() if due else None,
                'days_until': days_until,
            })

    upcoming.sort(key=lambda x: x['days_until'])
    return {
        'total_active': sum(by_stage.values()),
        'by_stage':     by_stage,
        'upcoming':     upcoming[:20],
    }


@app.route('/data/morning-briefing', methods=['GET'])
@require_jwt
def morning_briefing():
    try:
        wb = _workbook()
    except Exception as exc:
        app.logger.error('Workbook fetch failed: %s', exc)
        return jsonify({'error': 'Could not load workbook from Drive'}), 503

    return jsonify({
        'date':          date.today().isoformat(),
        'ar':            _parse_ar(wb),
        'pipeline':      _parse_pipeline(wb),
        'revenue_mtd':   _parse_is_revenue(wb),
        'revenue_trend': _parse_revenue_trend(wb),
        'cached_age':    round(time.time() - _wb_cache['fetched_at']),
    })


# ---------------------------------------------------------------------------
# Data: Client Health
# ---------------------------------------------------------------------------

def _build_client_map(wb):
    """Aggregate AR Aging + Engagement Pipeline records keyed by client name."""
    today = date.today()
    clients: dict[str, dict] = {}

    def ensure(name):
        if name not in clients:
            clients[name] = {'ar': [], 'pipeline': [], 'email': None}
        return clients[name]

    if 'AR Aging' in wb.sheetnames:
        for row in wb['AR Aging'].iter_rows(min_row=AR_DATA_START, values_only=True):
            if not row or not row[AR_C_CLIENT]:
                continue
            name = str(row[AR_C_CLIENT]).strip()
            invoice_val = str(row[AR_C_INVOICE]).strip() if len(row) > AR_C_INVOICE and row[AR_C_INVOICE] else ''
            if not name or not invoice_val:
                continue
            c = ensure(name)

            invoice     = row[AR_C_INVOICE]
            inv_date_raw= row[AR_C_INV_DATE]     if len(row) > AR_C_INV_DATE     else None
            due_raw     = row[AR_C_DUE_DATE]     if len(row) > AR_C_DUE_DATE     else None
            service     = row[AR_C_SERVICE]      if len(row) > AR_C_SERVICE      else None
            amount      = row[AR_C_AMOUNT]       if len(row) > AR_C_AMOUNT       else None
            outstanding = row[AR_C_OUTSTANDING]  if len(row) > AR_C_OUTSTANDING  else None
            status      = row[AR_C_STATUS]       if len(row) > AR_C_STATUS       else None
            email       = row[AR_C_EMAIL]        if len(row) > AR_C_EMAIL        else None

            if email and not c['email']:
                c['email'] = str(email)

            try:
                amount      = float(amount)      if amount      else 0.0
                outstanding = float(outstanding) if outstanding else 0.0
            except (TypeError, ValueError):
                amount = outstanding = 0.0

            inv_date  = _to_date(inv_date_raw)
            due       = _to_date(due_raw)
            days_over = (today - due).days if due else 0

            c['ar'].append({
                'invoice':      str(invoice) if invoice else '',
                'amount':       round(amount, 2),
                'outstanding':  round(outstanding, 2),
                'inv_date':     inv_date.isoformat() if inv_date else None,
                'due_date':     due.isoformat() if due else None,
                'days_overdue': days_over,
                'service':      str(service) if service else '',
                'status':       str(status) if status else 'Open',
            })

    if 'Engagement Pipeline' in wb.sheetnames:
        for row in wb['Engagement Pipeline'].iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            name  = str(row[0]).strip()
            stage = str(row[2]).strip() if len(row) > 2 and row[2] else None
            if not name or stage in ('Complete', 'Lost', None, 'None'):
                continue

            c          = ensure(name)
            eng_type   = str(row[3]) if len(row) > 3 and row[3] else 'Unknown'
            due_raw    = row[9] if len(row) > 9 else None
            due        = _to_date(due_raw)
            days_until = (due - today).days if due else None

            c['pipeline'].append({
                'type':       eng_type,
                'stage':      stage,
                'due_date':   due.isoformat() if due else None,
                'days_until': days_until,
            })

    return clients


def _health_score(open_ar: list) -> str:
    if not open_ar:
        return 'healthy'
    max_over = max((i['days_overdue'] for i in open_ar if i['days_overdue'] > 0), default=0)
    total    = sum(i['outstanding'] for i in open_ar)
    if max_over > 30:
        return 'needs_attention'
    if max_over > 0 or total > 10_000:
        return 'at_risk'
    return 'healthy'


@app.route('/data/clients', methods=['GET'])
@require_jwt
def client_health():
    try:
        wb = _workbook()
    except Exception as exc:
        app.logger.error('Workbook fetch failed: %s', exc)
        return jsonify({'error': 'Could not load workbook from Drive'}), 503

    today      = date.today()
    ytd_start  = date(today.year, 1, 1).isoformat()
    client_map = _build_client_map(wb)
    result     = []

    for name, data in sorted(client_map.items()):
        ar       = data['ar']
        pipeline = data['pipeline']

        open_ar  = [i for i in ar if i['outstanding'] > 0
                    and i['status'] not in ('Paid', 'Written Off', 'Void')]
        total_out   = round(sum(i['outstanding'] for i in open_ar), 2)
        overdue_amt = round(sum(i['outstanding'] for i in open_ar if i['days_overdue'] > 0), 2)
        max_over    = max((i['days_overdue'] for i in open_ar if i['days_overdue'] > 0), default=0)
        ytd_billed  = round(sum(i['amount'] for i in ar
                                if i['due_date'] and i['due_date'] >= ytd_start), 2)

        upcoming = sorted(
            [p for p in pipeline if p['days_until'] is not None],
            key=lambda x: x['days_until'],
        )

        result.append({
            'name':   name,
            'email':  data['email'],
            'health': _health_score(open_ar),
            'ar': {
                'total_outstanding':  total_out,
                'overdue_amount':     overdue_amt,
                'max_overdue_days':   max_over,
                'open_invoice_count': len(open_ar),
                'invoices': sorted(ar, key=lambda x: x['days_overdue'], reverse=True),
            },
            'pipeline': {
                'active_count':  len(pipeline),
                'next_deadline': upcoming[0] if upcoming else None,
                'matters':       pipeline,
            },
            'ytd_billed': ytd_billed,
        })

    order = {'needs_attention': 0, 'at_risk': 1, 'healthy': 2}
    result.sort(key=lambda x: (order[x['health']], -x['ar']['max_overdue_days'], x['name']))

    return jsonify({
        'date':    today.isoformat(),
        'summary': {
            'total':             len(result),
            'needs_attention':   sum(1 for c in result if c['health'] == 'needs_attention'),
            'at_risk':           sum(1 for c in result if c['health'] == 'at_risk'),
            'healthy':           sum(1 for c in result if c['health'] == 'healthy'),
        },
        'clients':    result,
        'cached_age': round(time.time() - _wb_cache['fetched_at']),
    })


# ---------------------------------------------------------------------------
# Client Portal
# ---------------------------------------------------------------------------

def _find_client_by_email(wb, email: str) -> str | None:
    """Return client name from AR Aging whose email matches, or None."""
    if 'AR Aging' not in wb.sheetnames:
        return None
    el = email.lower()
    for row in wb['AR Aging'].iter_rows(min_row=AR_DATA_START, values_only=True):
        if len(row) > AR_C_EMAIL and row[AR_C_EMAIL] and row[AR_C_CLIENT]:
            if str(row[AR_C_EMAIL]).strip().lower() == el:
                return str(row[AR_C_CLIENT]).strip()
    return None


def _client_stage_label(stage: str) -> str:
    s = stage.lower()
    if any(k in s for k in ('proposal', 'prospect', 'lead')):
        return 'Your engagement is being prepared.'
    if any(k in s for k in ('signed', 'onboard', 'setup', 'engagement')):
        return 'Your engagement is confirmed and getting started.'
    if any(k in s for k in ('progress', 'active', 'working', 'prep', 'wip')):
        return 'Your work is currently in progress.'
    if any(k in s for k in ('review', 'qc', 'partner', 'manager')):
        return 'Your work is in final review — almost ready.'
    if any(k in s for k in ('pending client', 'awaiting', 'info', 'client')):
        return 'Action needed — we need information from you. Please contact us.'
    if any(k in s for k in ('pending filing', 'efile', 'submitted', 'filed', 'delivered')):
        return 'Your work is complete and ready for delivery.'
    return f'Status: {stage}'


def _stage_needs_action(stage: str) -> bool:
    s = stage.lower()
    return any(k in s for k in ('pending client', 'awaiting', 'info', 'action'))


def _parse_client_data(wb, client_name: str) -> dict:
    today = date.today()
    invoices: list[dict] = []

    if 'AR Aging' in wb.sheetnames:
        for row in wb['AR Aging'].iter_rows(min_row=AR_DATA_START, values_only=True):
            if not row or not row[AR_C_CLIENT] or str(row[AR_C_CLIENT]).strip() != client_name:
                continue
            invoice     = row[AR_C_INVOICE]                                        if len(row) > AR_C_INVOICE     else None
            due_raw     = row[AR_C_DUE_DATE]                                       if len(row) > AR_C_DUE_DATE    else None
            amount      = row[AR_C_AMOUNT]                                         if len(row) > AR_C_AMOUNT      else None
            outstanding = row[AR_C_OUTSTANDING]                                    if len(row) > AR_C_OUTSTANDING else None
            status      = row[AR_C_STATUS]                                         if len(row) > AR_C_STATUS      else None
            try:
                amount      = float(amount)      if amount      else 0.0
                outstanding = float(outstanding) if outstanding else 0.0
            except (TypeError, ValueError):
                continue
            due       = _to_date(due_raw)
            days_over = (today - due).days if due else 0
            invoices.append({
                'invoice':      str(invoice) if invoice else '',
                'amount':       round(amount, 2),
                'outstanding':  round(outstanding, 2),
                'due_date':     due.isoformat() if due else None,
                'days_overdue': days_over,
                'status':       str(status) if status else 'Open',
                'paid':         str(status) == 'Paid' or outstanding == 0,
            })

    matters: list[dict] = []
    if 'Engagement Pipeline' in wb.sheetnames:
        for row in wb['Engagement Pipeline'].iter_rows(min_row=3, values_only=True):
            if not row or not row[0] or str(row[0]).strip() != client_name:
                continue
            stage = str(row[2]).strip() if len(row) > 2 and row[2] else None
            if not stage or stage in ('Complete', 'Lost'):
                continue
            eng_type   = str(row[3]) if len(row) > 3 and row[3] else 'Engagement'
            due_raw    = row[9] if len(row) > 9 else None
            due        = _to_date(due_raw)
            days_until = (due - today).days if due else None
            matters.append({
                'type':         eng_type,
                'stage':        stage,
                'stage_label':  _client_stage_label(stage),
                'needs_action': _stage_needs_action(stage),
                'due_date':     due.isoformat() if due else None,
                'days_until':   days_until,
            })

    open_inv       = [i for i in invoices if not i['paid']]
    total_out      = round(sum(i['outstanding'] for i in open_inv), 2)
    overdue_amount = round(sum(i['outstanding'] for i in open_inv if i['days_overdue'] > 0), 2)

    return {
        'client_name':       client_name,
        'total_outstanding': total_out,
        'overdue_amount':    overdue_amount,
        'action_required':   any(m['needs_action'] for m in matters),
        'invoices':          invoices,
        'matters':           matters,
    }


def _send_portal_magic_link(email: str, client_name: str, token: str):
    magic_url   = f"{PORTAL_URL}/auth.html?token={token}"
    expires_str = datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p UTC')
    body = f"""
<html><body style="font-family:Calibri,sans-serif;color:#1B2A3F;max-width:600px;margin:0 auto;">
  <div style="background:#1B2A3F;padding:24px 32px;border-radius:8px 8px 0 0;">
    <h1 style="color:white;margin:0;font-size:20px;">Lobel Accountancy Corporation</h1>
    <p style="color:#8BA7C4;margin:4px 0 0;font-size:13px;">Client Portal</p>
  </div>
  <div style="background:#f8f9fa;padding:32px;border-radius:0 0 8px 8px;border:1px solid #e0e0e0;">
    <h2 style="margin-top:0;font-size:18px;">Hello, {client_name}</h2>
    <p style="color:#4A5568;">Click the button below to sign in to your client portal and view
       your engagement status and invoices. This link expires in 15&nbsp;minutes and can only
       be used once.</p>
    <p style="margin:24px 0;">
      <a href="{magic_url}"
         style="background:#1B2A3F;color:white;padding:14px 28px;border-radius:6px;
                text-decoration:none;font-size:15px;font-weight:600;display:inline-block;">
        Sign in to your portal &rarr;
      </a>
    </p>
    <hr style="border:none;border-top:1px solid #e0e0e0;margin:24px 0;">
    <p style="font-size:12px;color:#999;">
      If you didn&rsquo;t request this link, please ignore this email.<br>
      Questions? Reply to this email or call (949)&nbsp;345-1925.<br>
      Link expires: {expires_str}
    </p>
  </div>
</body></html>"""
    msg = MIMEMultipart('alternative')
    msg['Subject'] = 'Lobel Accountancy — Your client portal sign-in link'
    msg['From']    = ZOHO_EMAIL
    msg['To']      = email
    msg.attach(MIMEText(body, 'html'))
    with smtplib.SMTP_SSL('smtp.zoho.com', 465) as server:
        server.login(ZOHO_EMAIL, ZOHO_PASSWORD)
        server.sendmail(ZOHO_EMAIL, email, msg.as_string())


def require_client_jwt(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '):
            return jsonify({'error': 'Missing authorization header'}), 401
        token = auth[7:]
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
            if payload.get('role') not in ('client', 'staff'):
                return jsonify({'error': 'Not authorized for client portal'}), 403
            request.client_email = payload['email']
            request.client_name  = payload.get('client_name', '')
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Session expired — please sign in again'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401
        return f(*args, **kwargs)
    return decorated


@app.route('/portal/auth/request', methods=['POST'])
def portal_auth_request():
    data  = request.get_json(silent=True) or {}
    email = (data.get('email') or '').strip().lower()
    if not email:
        return jsonify({'error': 'Email is required'}), 400

    generic = {'message': 'If this email is on file, a sign-in link has been sent.'}

    try:
        wb = _workbook()
    except Exception:
        return jsonify({'error': 'Service temporarily unavailable'}), 503

    is_staff = email in ALLOWED_EMAILS
    client_name = _find_client_by_email(wb, email) if not is_staff else '__staff__'
    if not client_name:
        return jsonify(generic), 200

    if not _rate_limit_ok(f'portal:{email}'):
        return jsonify({'error': 'Too many requests — try again in an hour.'}), 429

    token = serializer.dumps(email, salt='portal-magic-link')
    try:
        _send_portal_magic_link(email, client_name if not is_staff else 'Jeffrey', token)
    except Exception as exc:
        app.logger.error('Portal magic link failed for %s: %s', email, exc)
        return jsonify({'error': 'Failed to send email. Please try again.'}), 500

    return jsonify(generic), 200


@app.route('/portal/auth/verify', methods=['POST'])
def portal_auth_verify():
    data  = request.get_json(silent=True) or {}
    token = (data.get('token') or '').strip()
    if not token:
        return jsonify({'error': 'Token is required'}), 400

    _prune_used_tokens()
    if token in _used_tokens:
        return jsonify({'error': 'This link has already been used. Request a new one.'}), 401

    try:
        email = serializer.loads(token, salt='portal-magic-link', max_age=MAGIC_LINK_EXPIRY)
    except SignatureExpired:
        return jsonify({'error': 'Link expired. Request a new one.'}), 401
    except BadSignature:
        return jsonify({'error': 'Invalid or tampered token'}), 401

    _used_tokens[token] = time.time()

    try:
        wb = _workbook()
    except Exception:
        return jsonify({'error': 'Service temporarily unavailable'}), 503

    is_staff    = email in ALLOWED_EMAILS
    client_name = '__staff__' if is_staff else _find_client_by_email(wb, email)
    if not client_name:
        return jsonify({'error': 'Email not found in client records'}), 401

    now = int(time.time())
    jwt_payload = {
        'email':       email,
        'client_name': client_name,
        'role':        'staff' if is_staff else 'client',
        'iat':         now,
        'exp':         now + (7 * 86400),
        'iss':         'lac-portal',
    }
    jwt_token = jwt.encode(jwt_payload, JWT_SECRET, algorithm='HS256')
    return jsonify({'jwt': jwt_token, 'client_name': client_name, 'email': email})


@app.route('/portal/me', methods=['GET'])
@require_client_jwt
def portal_me():
    try:
        wb = _workbook()
    except Exception as exc:
        app.logger.error('Workbook fetch failed: %s', exc)
        return jsonify({'error': 'Service temporarily unavailable'}), 503

    is_staff    = (request.client_name == '__staff__')
    selected    = request.args.get('client') or ''

    if is_staff:
        # Only show clients who are actively engaged (portal access mirrors engagement status)
        PORTAL_STAGES = {'Engaged', 'In Progress', 'Under Review', 'Complete'}
        all_clients = []
        if 'Engagement Pipeline' in wb.sheetnames:
            seen = set()
            for row in wb['Engagement Pipeline'].iter_rows(min_row=3, values_only=True):
                name  = str(row[0]).strip() if row and row[0] else None
                stage = str(row[2]).strip() if row and len(row) > 2 and row[2] else ''
                if name and name not in seen and stage in PORTAL_STAGES:
                    seen.add(name)
                    all_clients.append(name)
        if selected and selected in all_clients:
            data = _parse_client_data(wb, selected)
        elif all_clients:
            data = _parse_client_data(wb, all_clients[0])
        else:
            data = {}
        return jsonify({
            'date':        date.today().isoformat(),
            'role':        'staff',
            'all_clients': all_clients,
            **data,
        })

    client_name = request.client_name or _find_client_by_email(wb, request.client_email)
    if not client_name:
        return jsonify({'error': 'Client record not found'}), 404

    return jsonify({'date': date.today().isoformat(), 'role': 'client',
                    **_parse_client_data(wb, client_name)})


@app.route('/portal/documents', methods=['GET'])
@require_client_jwt
def portal_documents():
    client_name = request.client_name
    if client_name == '__staff__':
        client_name = request.args.get('client', '').strip()
    if not client_name:
        return jsonify({'documents': []})
    docs = _load_client_docs()
    client_docs = docs.get(client_name, [])
    PORTAL_DOC_TYPES = {'engagement_letter', 'rep_letter', 'invoice_1', 'invoice_2'}
    visible = [d for d in client_docs if d.get('doc_type') in PORTAL_DOC_TYPES]
    return jsonify({
        'documents': sorted(visible, key=lambda x: x.get('date', ''), reverse=True)
    })


@app.route('/portal/doc/<doc_id>', methods=['GET'])
def portal_doc_serve(doc_id):
    """Serve a client document PDF. Accepts ?token= or Authorization header."""
    from flask import send_file as _send_file2
    token = request.args.get('token') or (
        request.headers.get('Authorization', '').replace('Bearer ', '').strip() or None
    )
    if not token:
        return jsonify({'error': 'Missing authorization'}), 401
    try:
        jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except Exception:
        return jsonify({'error': 'Invalid token'}), 401

    docs = _load_client_docs()
    for client_list in docs.values():
        for d in client_list:
            if d.get('id') == doc_id:
                pdf_path = d.get('pdf_path', '')
                if not pdf_path or not os.path.exists(pdf_path):
                    return jsonify({'error': 'File not found'}), 404
                return _send_file2(pdf_path, mimetype='application/pdf',
                                   as_attachment=True,
                                   download_name=d.get('filename', f'{doc_id}.pdf'))
    return jsonify({'error': 'Document not found'}), 404


# ---------------------------------------------------------------------------
# Data: Revenue Forecast & Capacity Planning
# ---------------------------------------------------------------------------

def _stage_probability(stage: str) -> float:
    s = stage.lower()
    if any(k in s for k in ('proposal', 'prospect', 'quote', 'lead')):        return 0.25
    if any(k in s for k in ('engagement', 'signed', 'onboard', 'setup')):     return 0.50
    if any(k in s for k in ('progress', 'active', 'working', 'prep', 'wip')): return 0.75
    if any(k in s for k in ('review', 'partner', 'manager', 'qc')):           return 0.90
    if any(k in s for k in ('client', 'awaiting', 'info', 'pending')):        return 0.80
    if any(k in s for k in ('filed', 'efile', 'submitted', 'delivered')):     return 0.95
    return 0.65


def _months_ahead(today: date, count: int = 6) -> list[tuple[int, int]]:
    result = []
    for i in range(1, count + 1):
        m = today.month + i
        y = today.year
        while m > 12:
            m -= 12
            y += 1
        result.append((y, m))
    return result


def _parse_forecast(wb):
    today    = date.today()
    capacity = int(os.getenv('MONTHLY_CAPACITY', '20'))

    # --- AR Aging: historical actuals + avg fee baseline ---
    hist_monthly: dict[str, float] = {}
    all_amounts:  list[float]      = []

    if 'AR Aging' in wb.sheetnames:
        for row in wb['AR Aging'].iter_rows(min_row=AR_DATA_START, values_only=True):
            amount  = row[AR_C_AMOUNT]   if len(row) > AR_C_AMOUNT   else None
            due_raw = row[AR_C_INV_DATE] if len(row) > AR_C_INV_DATE else None
            if not amount:
                continue
            try:
                amt = float(amount)
            except (TypeError, ValueError):
                continue
            if amt <= 0:
                continue
            all_amounts.append(amt)
            due = _to_date(due_raw)
            if due:
                hist_monthly[due.strftime('%Y-%m')] = hist_monthly.get(due.strftime('%Y-%m'), 0) + amt

    avg_fee = round(sum(all_amounts) / len(all_amounts), 2) if all_amounts else 0.0

    budget_monthly = _parse_budget_projections(wb)

    actuals = []
    for y, m in _months_seq(today, 6):
        mk = f'{y}-{m:02d}'
        actuals.append({'month': mk, 'label': date(y, m, 1).strftime('%b %y'),
                        'billed':    round(hist_monthly.get(mk, 0), 2),
                        'budgeted':  budget_monthly.get(mk, 0)})

    # --- Engagement Pipeline: forecast + workload + stage summary ---
    forecast_monthly: dict[str, float]       = {}
    workload_monthly: dict[str, dict]        = {}
    by_stage:         dict[str, dict]        = {}
    total_weighted  = 0.0
    total_est_value = 0.0

    if 'Engagement Pipeline' in wb.sheetnames:
        for row in wb['Engagement Pipeline'].iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            client  = str(row[0]).strip()
            stage   = str(row[2]).strip() if len(row) > 2 and row[2] else None
            due_raw = row[9] if len(row) > 9 else None

            if not client or not stage or stage in ('Complete', 'Lost'):
                continue

            try:
                deal_fee = float(row[6]) if len(row) > 6 and row[6] else avg_fee
            except (TypeError, ValueError):
                deal_fee = avg_fee

            prob     = _stage_probability(stage)
            weighted = deal_fee * prob
            total_est_value += deal_fee
            total_weighted  += weighted

            if stage not in by_stage:
                by_stage[stage] = {'count': 0, 'est_value': 0.0, 'weighted': 0.0, 'probability': prob}
            by_stage[stage]['count']     += 1
            by_stage[stage]['est_value'] += deal_fee
            by_stage[stage]['weighted']  += weighted

            due = _to_date(due_raw)
            if not due:
                continue
            days_until = (due - today).days
            if 0 <= days_until <= 180:
                mk = due.strftime('%Y-%m')
                forecast_monthly[mk] = forecast_monthly.get(mk, 0) + weighted
                if mk not in workload_monthly:
                    workload_monthly[mk] = {'count': 0, 'clients': set()}
                workload_monthly[mk]['count']   += 1
                workload_monthly[mk]['clients'].add(client)

    forecast = []
    for y, m in _months_ahead(today, 6):
        mk = f'{y}-{m:02d}'
        forecast.append({'month': mk, 'label': date(y, m, 1).strftime('%b %y'),
                         'projected': round(forecast_monthly.get(mk, 0), 2),
                         'budgeted':  budget_monthly.get(mk, 0)})

    workload = []
    for y, m in _months_ahead(today, 6):
        mk = f'{y}-{m:02d}'
        w  = workload_monthly.get(mk, {'count': 0, 'clients': set()})
        workload.append({
            'month':          mk,
            'label':          date(y, m, 1).strftime('%b %y'),
            'count':          w['count'],
            'unique_clients': len(w['clients']),
            'capacity':       capacity,
            'over_capacity':  w['count'] > capacity,
        })

    pipeline_stages = sorted(
        [{'stage': s, 'count': d['count'],
          'est_value': round(d['est_value'], 2),
          'weighted':  round(d['weighted'],  2),
          'probability': d['probability']}
         for s, d in by_stage.items()],
        key=lambda x: x['weighted'], reverse=True,
    )

    return {
        'avg_fee': avg_fee,
        'actuals': actuals,
        'forecast': forecast,
        'workload': workload,
        'pipeline': {
            'total_est_value': round(total_est_value, 2),
            'total_weighted':  round(total_weighted,  2),
            'by_stage':        pipeline_stages,
            'capacity':        capacity,
        },
    }


def _parse_budget_projections(wb):
    """Read Total Revenue monthly budget from Budget & Projections tab. Returns {'YYYY-MM': amount}."""
    if 'Budget & Projections' not in wb.sheetnames:
        return {}
    ws = wb['Budget & Projections']
    import re
    fy = date.today().year
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        if row and row[0]:
            m = re.search(r'FY(\d{4})', str(row[0]))
            if m:
                fy = int(m.group(1))
        break

    monthly = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        if str(row[0]).strip().lower() == 'total revenue':
            for idx in range(12):
                val = row[2 + idx] if len(row) > 2 + idx else None
                try:
                    amount = float(val) if val else 0.0
                except (TypeError, ValueError):
                    amount = 0.0
                monthly[f'{fy}-{idx + 1:02d}'] = amount
            break
    return monthly


def _parse_compliance_dates(wb):
    """Compute compliance dates from input assumptions — does not rely on formula cache."""
    if 'Key Compliance Dates' not in wb.sheetnames:
        return []

    ws    = wb['Key Compliance Dates']
    today = date.today()

    # Read plain-value input assumptions (rows 4–10, col B)
    assumptions: dict[str, date | None] = {}
    for row in ws.iter_rows(min_row=4, max_row=11, values_only=True):
        if row and row[0] and row[1]:
            assumptions[str(row[0]).strip()] = _to_date(row[1])

    incorp  = assumptions.get('Incorporation Date')
    fy_start = assumptions.get('FY Start Date')
    f2553   = assumptions.get('Form 2553 Filed Date')
    fy      = fy_start.year if fy_start else today.year

    # Read manually-entered DONE status from data rows (col F = index 5)
    done_set: set[str] = set()
    for row in ws.iter_rows(min_row=14, max_row=60, values_only=True):
        if not row or not row[1]:
            continue
        ob  = str(row[1]).strip()
        st  = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        if st.upper() == 'DONE':
            done_set.add(ob)

    items: list[dict] = []

    def add(category, obligation, due, frequency='', notes=''):
        if not due or obligation in done_set:
            return
        days = (due - today).days
        items.append({
            'category':   category,
            'obligation': obligation,
            'due_date':   due.isoformat(),
            'days_until': days,
            'frequency':  frequency,
            'notes':      notes,
            'status':     '',
            'overdue':    days < 0,
        })

    # ── Federal Tax ────────────────────────────────────────────────────────────
    if f2553:
        add('Federal Tax', 'Form 2553 — S-Corp Election', f2553, 'One-time')
    add('Federal Tax', f'Form 1120-S — S-Corp Return FY{fy}',   date(fy+1, 3, 15),  'Annual')
    add('Federal Tax', f'Form 7004 — Extension (if needed)',      date(fy+1, 3, 15),  'Annual')
    add('Federal Tax', f'Fed Est. Tax Q1 {fy+1}',                date(fy+1, 4, 15),  'Quarterly')
    add('Federal Tax', f'Fed Est. Tax Q2 {fy+1}',                date(fy+1, 6, 16),  'Quarterly')
    add('Federal Tax', f'Fed Est. Tax Q3 {fy+1}',                date(fy+1, 9, 15),  'Quarterly')
    add('Federal Tax', f'Fed Est. Tax Q4 {fy+1}',                date(fy+2, 1, 15),  'Quarterly')
    add('Federal Tax', f'Form 940 — FUTA Annual',                 date(fy+1, 1, 31),  'Annual', 'Only if payroll is run')
    add('Federal Tax', f'W-2 / W-3 to employees & SSA',          date(fy+1, 1, 31),  'Annual')

    # ── CA Tax ─────────────────────────────────────────────────────────────────
    add('CA Tax', f'CA Form 100-S — S-Corp Return FY{fy}', date(fy+1, 3, 15),  'Annual')
    add('CA Tax', f'CA Est. Tax Q1 {fy}',                   date(fy,   4, 15),  'Quarterly')
    add('CA Tax', f'CA Est. Tax Q2 {fy}',                   date(fy,   6, 15),  'Quarterly')
    add('CA Tax', f'CA Est. Tax Q3 {fy}',                   date(fy,   9, 15),  'Quarterly')
    add('CA Tax', f'CA Est. Tax Q4 {fy}',                   date(fy+1, 1, 15),  'Quarterly')
    add('CA Tax', 'CA Minimum Franchise Tax $800',           date(fy+1, 3, 15),  'Annual', 'Waived FY2026; $800 due FY2027+')
    add('CA Tax', 'CA SDI / Payroll Registration',           date(fy,  12, 31),  'One-time', 'Register with EDD if paying salary')

    # ── CA SOS ─────────────────────────────────────────────────────────────────
    if incorp:
        add('CA SOS', 'Statement of Information (SI-550)',
            incorp + timedelta(days=90), 'Annual', 'Due within 90 days of incorp; $25')
        add('CA SOS', 'Biennial Statement of Information',
            date(incorp.year+2, incorp.month, incorp.day), 'Biennial')

    # ── CBA ────────────────────────────────────────────────────────────────────
    if incorp:
        add('CBA', 'Notify CBA — Address Change',
            incorp + timedelta(days=30), 'Within 30 days')
    add('CBA', 'Individual CPA License Renewal', date(2027, 6, 30), 'Biennial', 'CA CPA license 132272')

    # ── Insurance & Registered Agent ───────────────────────────────────────────
    if incorp:
        renewal = date(incorp.year+1, incorp.month, incorp.day)
        add('Insurance',         'E&O Insurance Renewal',                    renewal, 'Annual', 'Min $100K per claim')
        add('Insurance',         'General Liability Renewal',                renewal, 'Annual')
        add('Registered Agent',  'Northwest Registered Agent Renewal',       renewal, 'Annual', '$145/yr')

    # ── Business License & EFTPS ───────────────────────────────────────────────
    add('Business License', 'Newport Beach Business License', date(fy+1, 1, 1), 'Annual')
    if incorp:
        add('Federal', 'EFTPS Registration Complete',
            incorp + timedelta(days=66), 'One-time', 'Must be active before first payroll')

    items.sort(key=lambda x: x['days_until'])
    return items


@app.route('/data/compliance-dates', methods=['GET'])
@require_jwt
def compliance_dates():
    try:
        wb = _workbook()
    except Exception as exc:
        app.logger.error('Workbook fetch failed: %s', exc)
        return jsonify({'error': 'Could not load workbook'}), 503

    items = _parse_compliance_dates(wb)
    overdue_count = sum(1 for i in items if i['overdue'])
    upcoming_30   = sum(1 for i in items if 0 <= i['days_until'] <= 30)
    # Show overdue + next 90 days on briefing page
    visible = [i for i in items if i['days_until'] <= 90]
    return jsonify({
        'date':           date.today().isoformat(),
        'items':          visible[:15],
        'overdue_count':  overdue_count,
        'upcoming_30':    upcoming_30,
    })


@app.route('/data/compliance-complete', methods=['POST'])
@require_jwt
def compliance_complete():
    """Mark a compliance item complete and write the date back to the workbook."""
    body       = request.get_json(force=True) or {}
    obligation = (body.get('obligation') or '').strip()
    completed  = (body.get('completed_date') or date.today().isoformat())

    if not obligation:
        return jsonify({'error': 'obligation required'}), 400

    try:
        wb, file_id, svc = _wb_download_fresh()

        if 'Key Compliance Dates' not in wb.sheetnames:
            return jsonify({'error': 'Key Compliance Dates tab not found'}), 404

        ws    = wb['Key Compliance Dates']
        found = False
        for row in ws.iter_rows(min_row=14):
            cell_ob = row[1] if len(row) > 1 else None
            if cell_ob and str(cell_ob.value or '').strip() == obligation:
                # col F (index 5) = Status, col G (index 6) = Date Completed
                if len(row) > 5:
                    row[5].value = 'DONE'
                if len(row) > 6:
                    row[6].value = completed
                found = True
                break

        if not found:
            return jsonify({'error': f"Obligation '{obligation}' not found"}), 404

        wb.calculation.calcMode = 'auto'
        wb.calculation.calcOnSave = True
        _wb_upload_async(wb, file_id, svc)

    except Exception as exc:
        app.logger.error('compliance_complete error: %s', exc)
        return jsonify({'error': str(exc)}), 500

    return jsonify({'success': True, 'obligation': obligation, 'completed_date': completed})


def _calendar_service():
    creds = service_account.Credentials.from_service_account_file(
        CREDENTIALS_PATH,
        scopes=['https://www.googleapis.com/auth/calendar.readonly'])
    return build('calendar', 'v3', credentials=creds)


@app.route('/data/calendar', methods=['GET'])
@require_jwt
def calendar_events():
    cal_id = os.getenv('GOOGLE_CALENDAR_ID', 'jlobel@lobelaccountancy.com')
    try:
        svc = _calendar_service()
        now    = datetime.now(timezone.utc).isoformat()
        end_dt = (datetime.now(timezone.utc) + timedelta(days=14)).isoformat()
        result = svc.events().list(
            calendarId=cal_id,
            timeMin=now,
            timeMax=end_dt,
            maxResults=20,
            singleEvents=True,
            orderBy='startTime',
        ).execute()

        events = []
        for ev in result.get('items', []):
            start   = ev['start'].get('dateTime', ev['start'].get('date', ''))
            end_t   = ev['end'].get('dateTime',   ev['end'].get('date',   ''))
            all_day = 'dateTime' not in ev['start']
            events.append({
                'id':          ev['id'],
                'title':       ev.get('summary', '(No title)'),
                'start':       start,
                'end':         end_t,
                'all_day':     all_day,
                'location':    ev.get('location', ''),
            })
        return jsonify({'events': events})
    except Exception as exc:
        app.logger.error('Calendar fetch failed: %s', exc)
        return jsonify({'events': [], 'error': str(exc)})


@app.route('/data/forecast', methods=['GET'])
@require_jwt
def revenue_forecast():
    try:
        wb = _workbook()
    except Exception as exc:
        app.logger.error('Workbook fetch failed: %s', exc)
        return jsonify({'error': 'Could not load workbook from Drive'}), 503

    return jsonify({
        'date':       date.today().isoformat(),
        **_parse_forecast(wb),
        'cached_age': round(time.time() - _wb_cache['fetched_at']),
    })


# ---------------------------------------------------------------------------
# Data: BI Dashboard
# ---------------------------------------------------------------------------

def _months_seq(today: date, count: int = 12) -> list[tuple[int, int]]:
    """Return list of (year, month) tuples for the last `count` months, oldest first."""
    result = []
    for i in range(count - 1, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        result.append((y, m))
    return result


def _parse_is_revenue(wb):
    """Read Total Revenue for the current month from the Income Statement tab.
    Scans dynamically for the header row and Total Revenue row so it stays
    correct if rows/columns are ever inserted.
    """
    if 'Income Statement' not in wb.sheetnames:
        return None

    ws         = wb['Income Statement']
    today      = date.today()
    month_abbr = today.strftime('%b')  # e.g. 'Jun'

    # Find the header row: col A == 'Account' with month names alongside
    header_row_vals = None
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        if row and str(row[0] or '').strip() == 'Account':
            header_row_vals = row
            break
    if header_row_vals is None:
        return None

    # Locate current month's column (match 3-char abbreviation, case-insensitive)
    month_col = None
    for col_idx, val in enumerate(header_row_vals):
        if val and str(val).strip()[:3].lower() == month_abbr.lower():
            month_col = col_idx
            break
    if month_col is None:
        return None

    # Find the Total Revenue row by scanning col A
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and str(row[0] or '').strip() == 'Total Revenue':
            val = row[month_col] if len(row) > month_col else None
            try:
                return float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                return 0.0

    return None


def _parse_revenue_trend(wb):
    """Return [{month, revenue, expenses, net}] for all available months."""
    if 'Income Statement' not in wb.sheetnames:
        return []
    ws = wb['Income Statement']
    header_vals = None
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        if row and str(row[0] or '').strip() == 'Account':
            header_vals = list(row)
            break
    if not header_vals:
        return []
    MONTHS = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
    month_cols = {}
    for ci, val in enumerate(header_vals):
        s = str(val or '').strip()
        if s[:3].lower() in MONTHS:
            month_cols[ci] = s[:3].capitalize()
    rev_row = exp_row = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        label = str(row[0] or '').strip()
        if label == 'Total Revenue':
            rev_row = list(row)
        elif 'Total' in label and 'Expense' in label and exp_row is None:
            exp_row = list(row)
        if rev_row and exp_row:
            break
    result = []
    for ci in sorted(month_cols):
        month = month_cols[ci]
        rev = 0.0
        exp = 0.0
        try:
            rev = float(rev_row[ci] or 0) if rev_row and ci < len(rev_row) else 0.0
        except (TypeError, ValueError):
            pass
        try:
            exp = float(exp_row[ci] or 0) if exp_row and ci < len(exp_row) else 0.0
        except (TypeError, ValueError):
            pass
        if rev != 0 or exp != 0:
            result.append({'month': month, 'revenue': round(rev, 2), 'expenses': round(exp, 2), 'net': round(rev - exp, 2)})
    return result


def _parse_bi_revenue(wb):
    if 'AR Aging' not in wb.sheetnames:
        return {'by_month': [], 'ytd_billed': 0, 'by_type': [], 'top_clients': []}

    # Build client → engagement type lookup from Engagement Pipeline
    ep_type: dict[str, str] = {}
    if 'Engagement Pipeline' in wb.sheetnames:
        for row in wb['Engagement Pipeline'].iter_rows(min_row=3, values_only=True):
            if row and row[0] and row[3]:
                ep_type[str(row[0]).strip()] = str(row[3]).strip()

    ws    = wb['AR Aging']
    today = date.today()
    ytd_start = date(today.year, 1, 1)

    monthly:   dict[str, float] = {}
    by_type:   dict[str, float] = {}
    by_client: dict[str, float] = {}
    ytd_billed = 0.0

    for row in ws.iter_rows(min_row=AR_DATA_START, values_only=True):
        if len(row) <= AR_C_AMOUNT:
            continue
        client  = row[AR_C_CLIENT]
        due_raw = row[AR_C_INV_DATE]
        amount  = row[AR_C_AMOUNT]

        if not client or not amount:
            continue
        try:
            amount = float(amount)
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue

        due = _to_date(due_raw)
        if not due:
            continue

        eng_type = ep_type.get(str(client).strip(), 'Other')

        month_key = due.strftime('%Y-%m')
        monthly[month_key]         = monthly.get(month_key, 0)     + amount
        by_client[str(client)]     = by_client.get(str(client), 0) + amount
        by_type[eng_type]          = by_type.get(eng_type, 0)       + amount
        if due >= ytd_start:
            ytd_billed += amount

    months = []
    for y, m in _months_seq(today):
        mk  = f'{y}-{m:02d}'
        lbl = date(y, m, 1).strftime('%b %y')
        months.append({'month': mk, 'label': lbl, 'billed': round(monthly.get(mk, 0), 2)})

    top_clients = sorted(by_client.items(), key=lambda x: x[1], reverse=True)[:8]
    top_types   = sorted(by_type.items(),   key=lambda x: x[1], reverse=True)

    return {
        'by_month':    months,
        'ytd_billed':  round(ytd_billed, 2),
        'by_type':     [{'type': t, 'billed': round(a, 2)} for t, a in top_types],
        'top_clients': [{'client': c, 'billed': round(a, 2)} for c, a in top_clients],
    }


def _parse_bi_ar_buckets(wb):
    empty = {'count': 0, 'amount': 0.0}
    buckets = {k: dict(empty) for k in ('current', 'd0_30', 'd31_60', 'd61_90', 'd90plus')}

    if 'AR Aging' not in wb.sheetnames:
        return buckets

    ws    = wb['AR Aging']
    today = date.today()

    for row in ws.iter_rows(min_row=AR_DATA_START, values_only=True):
        if len(row) <= AR_C_OUTSTANDING:
            continue
        client      = row[AR_C_CLIENT]
        due_raw     = row[AR_C_DUE_DATE]
        outstanding = row[AR_C_OUTSTANDING]
        status      = row[AR_C_STATUS] if len(row) > AR_C_STATUS else None

        if not client or not outstanding:
            continue
        try:
            outstanding = float(outstanding)
        except (TypeError, ValueError):
            continue
        if outstanding <= 0 or status in ('Paid', 'Written Off', 'Void'):
            continue

        due   = _to_date(due_raw)
        dover = (today - due).days if due else 0

        if dover <= 0:
            k = 'current'
        elif dover <= 30:
            k = 'd0_30'
        elif dover <= 60:
            k = 'd31_60'
        elif dover <= 90:
            k = 'd61_90'
        else:
            k = 'd90plus'

        buckets[k]['count']  += 1
        buckets[k]['amount'] += outstanding

    for k in buckets:
        buckets[k]['amount'] = round(buckets[k]['amount'], 2)
    return buckets


def _parse_bi_pipeline(wb):
    if 'Engagement Pipeline' not in wb.sheetnames:
        return {'by_stage': {}, 'by_type': {}, 'due_by_month': []}

    ws    = wb['Engagement Pipeline']
    today = date.today()
    by_stage: dict[str, int] = {}
    by_type:  dict[str, int] = {}
    due_monthly: dict[str, int] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 3:
            continue
        client   = row[0]
        stage    = row[2] if len(row) > 2 else None
        eng_type = row[3] if len(row) > 3 else None
        due_raw  = row[9] if len(row) > 9 else None

        if not client or stage in ('Complete', 'Lost', None):
            continue

        by_stage[str(stage)] = by_stage.get(str(stage), 0) + 1
        if eng_type:
            by_type[str(eng_type)] = by_type.get(str(eng_type), 0) + 1

        due = _to_date(due_raw)
        if due and due >= today:
            mk = due.strftime('%Y-%m')
            due_monthly[mk] = due_monthly.get(mk, 0) + 1

    months = []
    for y, m in _months_seq(today, 6):
        mk  = f'{y}-{m:02d}'
        lbl = date(y, m, 1).strftime('%b %y')
        months.append({'month': mk, 'label': lbl, 'count': due_monthly.get(mk, 0)})

    return {
        'by_stage': dict(sorted(by_stage.items(), key=lambda x: x[1], reverse=True)),
        'by_type':  dict(sorted(by_type.items(),  key=lambda x: x[1], reverse=True)),
        'due_by_month': months,
    }


def _parse_is_net_income(wb):
    """Return current month net income/loss from Income Statement: {value, label}."""
    import re as _re
    if 'Income Statement' not in wb.sheetnames:
        return None
    ws         = wb['Income Statement']
    month_abbr = date.today().strftime('%b')

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        if row and str(row[0] or '').strip() == 'Account':
            header_row = row
            break
    if header_row is None:
        return None

    month_col = None
    for ci, val in enumerate(header_row):
        if val and str(val).strip()[:3].lower() == month_abbr.lower():
            month_col = ci
            break
    if month_col is None:
        return None

    for row in ws.iter_rows(min_row=1, values_only=True):
        label = str(row[0] or '').strip() if row else ''
        if _re.search(r'NET\s+(INCOME|LOSS)', label, _re.I):
            val = row[month_col] if len(row) > month_col else None
            try:
                value = float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                value = 0.0
            return {'value': value, 'label': label}

    return None


@app.route('/data/bi', methods=['GET'])
@require_jwt
def bi_dashboard():
    try:
        wb = _workbook()
    except Exception as exc:
        app.logger.error('Workbook fetch failed: %s', exc)
        return jsonify({'error': 'Could not load workbook from Drive'}), 503

    return jsonify({
        'date':           date.today().isoformat(),
        'revenue':        _parse_bi_revenue(wb),
        'revenue_mtd':    _parse_is_revenue(wb),
        'net_income_mtd': _parse_is_net_income(wb),
        'ar_buckets':     _parse_bi_ar_buckets(wb),
        'pipeline':       _parse_bi_pipeline(wb),
        'cached_age':     round(time.time() - _wb_cache['fetched_at']),
    })


REG_STATE_PATH   = '/home/jlobel/lac_automation/phase6/regulatory_state.json'
WB_LOCAL_PATH    = '/home/jlobel/lac_automation/LAC_Workbook.xlsx'
DRIVE_RW_SCOPES  = ['https://www.googleapis.com/auth/drive']
AR_HEADER_ROW    = 12   # row where Client/Invoice/... headers live in AR Aging
AR_DATA_START    = 13   # first invoice data row
# AR Aging column indices (0-based), from row 12 headers
AR_C_CLIENT      = 0   # Client
AR_C_CLIENT_NUM  = 1   # Client Number
AR_C_INVOICE     = 2   # Invoice #
AR_C_INV_DATE    = 3   # Invoice Date
AR_C_DUE_DATE    = 4   # Due Date
AR_C_SERVICE     = 5   # Service
AR_C_AMOUNT      = 6   # Amount ($)
AR_C_PAID        = 7   # Paid ($)
AR_C_OUTSTANDING = 8   # Outstanding ($)
AR_C_DAYS        = 9   # Days Outstanding
AR_C_STATUS      = 10  # Status
AR_C_EMAIL       = 11  # Client Email
AR_C_ADDRESS     = 12  # Client Address
AR_C_REMINDER    = 13  # Last Reminder Sent


def _wb_download_fresh():
    """Download workbook from Drive, return (wb, file_id, svc). Raises on error."""
    import io as _io
    from google.oauth2 import service_account as _sa
    from googleapiclient.discovery import build as _build
    from googleapiclient.http import MediaIoBaseDownload as _DL
    creds   = _sa.Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=DRIVE_RW_SCOPES)
    svc     = _build('drive', 'v3', credentials=creds)
    file_id = _wb_file_id_cache.get('id')
    if not file_id:
        files = svc.files().list(
            q=f"name='{WORKBOOK_NAME}' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
            fields='files(id)'
        ).execute().get('files', [])
        if not files:
            raise RuntimeError('Workbook not found on Drive')
        file_id = files[0]['id']
        _wb_file_id_cache['id'] = file_id
    buf = _io.BytesIO()
    dl  = _DL(buf, svc.files().get_media(fileId=file_id))
    done = False
    while not done: _, done = dl.next_chunk()
    buf.seek(0)
    from openpyxl import load_workbook as _lw
    wb = _lw(buf)
    return wb, file_id, svc


def _wb_upload(wb, file_id, svc):
    """Save workbook locally and push to Drive, then bust cache."""
    import io as _io
    from googleapiclient.http import MediaFileUpload as _UL
    out = _io.BytesIO()
    wb.save(out)
    out.seek(0)
    with open(WB_LOCAL_PATH, 'wb') as f:
        f.write(out.read())
    svc.files().update(
        fileId=file_id,
        media_body=_UL(WB_LOCAL_PATH,
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    ).execute()
    _wb_cache['wb'] = None
    _wb_cache['fetched_at'] = 0


def _wb_upload_async(wb, file_id, svc):
    """Update the read cache immediately with modified wb, then upload to Drive in background.

    Reads that arrive during the upload (typically 3-5s) see the correct post-change
    data from cache rather than the stale pre-change version.
    """
    import io as _io
    import threading as _threading
    from googleapiclient.http import MediaFileUpload as _UL

    # Immediately make modified data visible to reads
    _wb_cache['wb'] = wb
    _wb_cache['fetched_at'] = time.time()

    def _do_upload():
        try:
            out = _io.BytesIO()
            wb.save(out)
            out.seek(0)
            with open(WB_LOCAL_PATH, 'wb') as f:
                f.write(out.read())
            svc.files().update(
                fileId=file_id,
                media_body=_UL(WB_LOCAL_PATH,
                               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            ).execute()
            # Reset TTL from upload-complete time
            _wb_cache['fetched_at'] = time.time()
        except Exception as e:
            app.logger.error('Background wb upload failed: %s', e)
            # Bust cache so next read re-fetches from Drive
            _wb_cache['wb'] = None
            _wb_cache['fetched_at'] = 0

    _threading.Thread(target=_do_upload, daemon=True).start()


@app.route('/ar/payment', methods=['POST'])
@require_jwt
def ar_payment():
    """Record a full or partial payment against an invoice in AR Aging.
    Body: { invoice, paid_amount, note (optional) }
    Staff only.
    """
    if request.client_name != '__staff__':
        return jsonify({'error': 'Staff only'}), 403

    body       = request.get_json(silent=True) or {}
    invoice_id = (body.get('invoice') or '').strip()
    try:
        paid_amount = float(body.get('paid_amount', 0))
    except (TypeError, ValueError):
        return jsonify({'error': 'paid_amount must be a number'}), 400
    note = (body.get('note') or '').strip()

    if not invoice_id or paid_amount <= 0:
        return jsonify({'error': 'invoice and paid_amount required'}), 400

    try:
        wb, file_id, svc = _wb_download_fresh()
        if 'AR Aging' not in wb.sheetnames:
            return jsonify({'error': 'AR Aging tab not found'}), 404

        ws = wb['AR Aging']
        found = False
        for row in ws.iter_rows(min_row=AR_DATA_START):
            inv_cell = row[AR_C_INVOICE] if len(row) > AR_C_INVOICE else None
            if not inv_cell or str(inv_cell.value or '').strip() != invoice_id:
                continue

            amount_cell      = row[AR_C_AMOUNT]      if len(row) > AR_C_AMOUNT      else None
            paid_cell        = row[AR_C_PAID]        if len(row) > AR_C_PAID        else None
            outstanding_cell = row[AR_C_OUTSTANDING] if len(row) > AR_C_OUTSTANDING else None
            status_cell      = row[AR_C_STATUS]      if len(row) > AR_C_STATUS      else None

            try:
                total = float(amount_cell.value or 0)
            except (TypeError, ValueError):
                total = 0.0

            # Accumulate (support multiple partial payments)
            try:
                prev_paid = float(paid_cell.value or 0)
            except (TypeError, ValueError):
                prev_paid = 0.0

            new_paid        = round(min(prev_paid + paid_amount, total), 2)
            new_outstanding = round(max(total - new_paid, 0), 2)
            new_status      = 'Paid' if new_outstanding == 0 else 'Partial'

            paid_cell.value        = new_paid
            outstanding_cell.value = new_outstanding
            status_cell.value      = new_status

            days_cell = row[AR_C_DAYS] if len(row) > AR_C_DAYS else None
            if days_cell and new_status == 'Paid':
                days_cell.value = 0

            found = True
            client_name = str(row[0].value or '').strip()
            break

        if not found:
            return jsonify({'error': f'Invoice {invoice_id} not found'}), 404

        _wb_upload_async(wb, file_id, svc)

        actor = 'admin'
        detail = f'{invoice_id}: ${paid_amount:,.2f} payment recorded → {new_status}'
        if note:
            detail += f' — {note}'
        _log_activity(client_name, actor, 'Payment Recorded', detail)

        return jsonify({
            'success':     True,
            'invoice':     invoice_id,
            'paid':        new_paid,
            'outstanding': new_outstanding,
            'status':      new_status,
        })

    except Exception as exc:
        app.logger.error('ar_payment error: %s', exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/ar/delete', methods=['POST'])
@require_jwt
def ar_delete():
    """Delete an invoice row from AR Aging by clearing its cells. Staff only."""
    if request.client_name != '__staff__':
        return jsonify({'error': 'Staff only'}), 403

    body       = request.get_json(silent=True) or {}
    invoice_id = (body.get('invoice') or '').strip()
    if not invoice_id:
        return jsonify({'error': 'invoice required'}), 400

    try:
        wb, file_id, svc = _wb_download_fresh()
        if 'AR Aging' not in wb.sheetnames:
            return jsonify({'error': 'AR Aging tab not found'}), 404

        ws = wb['AR Aging']
        found_row   = None
        client_name = ''
        for row in ws.iter_rows(min_row=AR_DATA_START):
            inv_cell = row[AR_C_INVOICE] if len(row) > AR_C_INVOICE else None
            if not inv_cell or str(inv_cell.value or '').strip() != invoice_id:
                continue
            found_row   = inv_cell.row
            client_name = str(row[AR_C_CLIENT].value or '').strip()
            break

        if found_row is None:
            # Invoice is absent from Drive — was cleared externally or already deleted.
            # Bust the stale in-memory cache so the next read reflects the clean workbook.
            _wb_cache['wb'] = None
            _wb_cache['fetched_at'] = 0
            return jsonify({'ok': True, 'invoice': invoice_id})

        # Mark as Void rather than wiping the row entirely. Keeping the client/
        # invoice/service cells intact lets invoice_automation.py see that an
        # invoice was already issued for this engagement and prevents it from
        # re-creating the invoice the next time it runs.
        ws.cell(row=found_row, column=AR_C_AMOUNT      + 1).value = 0
        ws.cell(row=found_row, column=AR_C_PAID        + 1).value = 0
        ws.cell(row=found_row, column=AR_C_OUTSTANDING + 1).value = 0
        ws.cell(row=found_row, column=AR_C_STATUS      + 1).value = 'Void'

        _wb_upload_async(wb, file_id, svc)
        _log_activity(client_name, 'admin', 'Invoice Voided', f'{invoice_id} voided in AR Aging')

        return jsonify({'ok': True, 'invoice': invoice_id})

    except Exception as exc:
        app.logger.error('ar_delete error: %s', exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/data/regulatory', methods=['GET'])
@require_jwt
def regulatory_updates():
    import json as _json
    from datetime import timedelta

    try:
        with open(REG_STATE_PATH) as f:
            state = _json.load(f)
    except (FileNotFoundError, _json.JSONDecodeError):
        return jsonify({'items': [], 'last_digest': None, 'source': 'state_missing'})

    pending     = state.get('pending', [])
    last_digest = state.get('last_digest')

    # Always show items found in the last 14 days from pending
    cutoff = (date.today() - timedelta(days=14)).isoformat()
    items  = [
        {
            'source':  i.get('source', ''),
            'title':   i.get('title', ''),
            'url':     i.get('url', ''),
            'found':   i.get('found', ''),
            'urgent':  i.get('urgent', False),
        }
        for i in pending
        if i.get('found', '') >= cutoff
    ]

    # Sort: urgent first, then most recent
    items.sort(key=lambda x: (not x['urgent'], x['found']), reverse=False)
    items.sort(key=lambda x: not x['urgent'])

    return jsonify({
        'items':       items[:12],
        'total':       len(items),
        'last_digest': last_digest,
    })


@app.route('/budget/clients', methods=['GET'])
@require_jwt
def budget_clients():
    """Return client list for the budget dropdown."""
    # First try local budget JSON (fast)
    budget_json = _load_budget_json()
    if budget_json:
        clients = sorted({v['client'] for v in budget_json.values() if v.get('client')})
        if clients:
            return jsonify({'clients': clients})

    # Fall back to Engagement Pipeline in workbook
    try:
        wb = _workbook()
    except Exception as exc:
        return jsonify({'error': str(exc)}), 503
    if 'Engagement Pipeline' in wb.sheetnames:
        clients = sorted({
            str(row[0]).strip()
            for row in wb['Engagement Pipeline'].iter_rows(min_row=3, values_only=True)
            if row and row[0] and len(row) > 2 and row[2]
            and str(row[2]).strip() not in ('', 'None')
        })
    else:
        clients = []
    return jsonify({'clients': clients})


@app.route('/budget/data', methods=['GET'])
@require_jwt
def budget_data():
    """Return Budget Backups rows and Clockify actuals for a client."""
    client = request.args.get('client', '').strip()
    if not client:
        return jsonify({'error': 'client required'}), 400
    try:
        wb = _workbook()
    except Exception as exc:
        return jsonify({'error': str(exc)}), 503

    HOURLY_RATE = float(os.getenv('CLOCKIFY_HOURLY_RATE', 225))

    # Budget data from workbook Budget Data tab (two-way sync); fall back to JSON
    budget_map = {}
    if 'Budget Data' in wb.sheetnames:
        for row in wb['Budget Data'].iter_rows(min_row=2, values_only=True):
            if row and str(row[0] or '').strip() == client and row[1] and row[2] is not None:
                budget_map[str(row[1]).strip()] = float(row[2])
    if not budget_map:
        budget_json = _load_budget_json()
        budget_map  = {
            v['work_area']: v['budget_hrs']
            for k, v in budget_json.items()
            if v.get('client') == client
        }

    # Clockify actuals from Time Entries
    actual_map = {}
    if 'Time Entries' in wb.sheetnames:
        for row in wb['Time Entries'].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            entry_client = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            work_area    = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            hours        = float(row[4]) if len(row) > 4 and row[4] else 0.0
            if entry_client == client and work_area:
                actual_map[work_area] = actual_map.get(work_area, 0.0) + hours

    WORK_AREAS = [
        'Engagement Planning & Administration',
        'Cash & Restricted Cash',
        'Accounts Receivable & Revenue Cutoff',
        'Fixed Assets & Depreciation',
        'Prepaid Expenses & Other Current Assets',
        'Other Assets (Incl. Income Taxes & Deferred Taxes)',
        'Accounts Payable & Accrued Liabilities',
        'Debt, Notes Payable & Interest',
        'Other Liabilities',
        'Deferred Revenue & Revenue Recognition (ASC 606)',
        "Equity / Owner's Equity & Capital",
        'Revenue (Detailed Substantive Testing)',
        'Expenses & Cost of Revenue',
        'Payroll & Employee Benefits',
        'Financial Statement Preparation & Disclosures',
        'Engagement Wrap-Up & Quality Review',
        'Client Communication & Project Management',
        'Contingency & Out-of-Scope Reserve',
    ]

    rows = []
    for area in WORK_AREAS:
        budget_hrs = budget_map.get(area, 0.0)
        actual_hrs = actual_map.get(area, 0.0)
        rows.append({
            'work_area':  area,
            'budget_hrs': budget_hrs,
            'actual_hrs': round(actual_hrs, 2),
            'var_hrs':    round(budget_hrs - actual_hrs, 2),
            'budget_usd': round(budget_hrs * HOURLY_RATE, 2),
            'actual_usd': round(actual_hrs * HOURLY_RATE, 2),
            'var_usd':    round((budget_hrs - actual_hrs) * HOURLY_RATE, 2),
            'pct_complete': round((actual_hrs / budget_hrs * 100) if budget_hrs > 0 else 0, 1),
        })

    total_budget_hrs = sum(r['budget_hrs'] for r in rows)
    total_actual_hrs = sum(r['actual_hrs'] for r in rows)
    return jsonify({
        'client':          client,
        'rows':            rows,
        'total_budget_hrs': total_budget_hrs,
        'total_actual_hrs': round(total_actual_hrs, 2),
        'hrs_remaining':   round(total_budget_hrs - total_actual_hrs, 2),
        'budget_usd':      round(total_budget_hrs * HOURLY_RATE, 2),
        'actual_usd':      round(total_actual_hrs * HOURLY_RATE, 2),
        'var_usd':         round((total_budget_hrs - total_actual_hrs) * HOURLY_RATE, 2),
        'realization_rate': round((total_actual_hrs / total_budget_hrs * 100) if total_budget_hrs > 0 else 0, 1),
        'hourly_rate':     HOURLY_RATE,
        'cached_age':      round(time.time() - _wb_cache['fetched_at']),
    })


BUDGET_JSON_PATH = '/home/jlobel/lac_automation/budget_backups.json'

# Clockify Fixed Fee project task name → task_id (source of truth for work area names)
_CK_FF_PROJECT  = '6a244432c9387aecfb2d67ac'
_CK_TASKS = {
    'Engagement Planning & Administration':              '6a24446b7b8dfc6725d148ee',
    'Cash & Restricted Cash':                           '6a24447641257465f53d8e8b',
    'Accounts Receivable & Revenue Cutoff':             '6a24447ec9387aecfb2d6f49',
    'Fixed Assets & Depreciation':                      '6a2444873b1c9faa378ebe72',
    'Prepaid Expenses & Other Current Assets':          '6a24449341257465f53d907f',
    'Other Assets (Incl. Income Taxes & Deferred Taxes)': '6a2444d9177ca3150f886d72',
    'Accounts Payable & Accrued Liabilities':           '6a2444e241257465f53d9a00',
    'Debt, Notes Payable & Interest':                   '6a2444e8148dd6010d858e4c',
    'Other Liabilities':                                '6a2444edc9387aecfb2d787e',
    'Deferred Revenue & Revenue Recognition (ASC 606)': '6a2444fbc9387aecfb2d7c8a',
    "Equity / Owner's Equity & Capital":                '6a24452c7b8dfc6725d15693',
    'Revenue (Detailed Substantive Testing)':           '6a244536c9387aecfb2d80ef',
    'Expenses & Cost of Revenue':                       '6a24453b41257465f53da402',
    'Payroll & Employee Benefits':                      '6a244542148dd6010d85998a',
    'Financial Statement Preparation & Disclosures':    '6a24454e497f3ef707ef49bf',
    'Engagement Wrap-Up & Quality Review':              '6a2445ae6922488bf183a9fb',
    'Client Communication & Project Management':        '6a2445bc41257465f53daede',
    'Contingency & Out-of-Scope Reserve':               '6a2445c7497f3ef707ef5677',
}

def _load_budget_json() -> dict:
    try:
        with open(BUDGET_JSON_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _save_budget_json(data: dict):
    with open(BUDGET_JSON_PATH, 'w') as f:
        json.dump(data, f, indent=2)


@app.route('/budget/save', methods=['POST'])
@require_jwt
def budget_save():
    """Write a budget_hrs update to local JSON and immediately to workbook Budget Data tab."""
    body = request.get_json(force=True) or {}
    client     = str(body.get('client', '')).strip()
    work_area  = str(body.get('work_area', '')).strip()
    budget_hrs = body.get('budget_hrs')

    if not client or not work_area or budget_hrs is None:
        return jsonify({'error': 'client, work_area, budget_hrs required'}), 400
    try:
        budget_hrs = float(budget_hrs)
    except (TypeError, ValueError):
        return jsonify({'error': 'budget_hrs must be a number'}), 400

    today_str = date.today().strftime('%m/%d/%Y')
    key = f"{client}|{work_area}"

    # 1. Write to local JSON
    try:
        data = _load_budget_json()
        data[key] = {
            'client':     client,
            'work_area':  work_area,
            'budget_hrs': budget_hrs,
            'updated':    date.today().isoformat(),
        }
        _save_budget_json(data)
    except Exception as exc:
        app.logger.error('budget_save JSON error: %s', exc)
        return jsonify({'error': str(exc)}), 500

    # 2. Write to workbook Budget Data tab immediately
    wb_err = None
    try:
        wb, file_id, svc = _wb_download_fresh()
        if 'Budget Data' in wb.sheetnames:
            ws = wb['Budget Data']
            # Find existing row by Key column (col E = index 5, 1-based)
            target_row = None
            last_data_row = 1
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and any(v is not None for v in row[:4]):
                    last_data_row = i
                    if str(row[4] or '').strip() == key:
                        target_row = i
                        break
            if target_row:
                ws.cell(row=target_row, column=3).value = budget_hrs
                ws.cell(row=target_row, column=4).value = today_str
            else:
                new_row = last_data_row + 1
                ws.cell(row=new_row, column=1).value = client
                ws.cell(row=new_row, column=2).value = work_area
                ws.cell(row=new_row, column=3).value = budget_hrs
                ws.cell(row=new_row, column=4).value = today_str
                ws.cell(row=new_row, column=5).value = key
        _wb_upload_async(wb, file_id, svc)
    except Exception as exc:
        app.logger.error('budget_save workbook error: %s', exc)
        wb_err = str(exc)

    # 3. Push estimate to Clockify Fixed Fee task (if work_area matches a known task)
    ck_err = None
    task_id = _CK_TASKS.get(work_area)
    if task_id:
        try:
            def _to_dur(hrs):
                total_m = round(hrs * 60)
                h, m = total_m // 60, total_m % 60
                return f"PT{h}H{m}M" if m else f"PT{h}H"
            CK_API_KEY   = os.getenv('CLOCKIFY_API_KEY')
            CK_WORKSPACE = os.getenv('CLOCKIFY_WORKSPACE_ID')
            ck_headers = {'X-Api-Key': CK_API_KEY, 'Content-Type': 'application/json'}
            url = (f"https://api.clockify.me/api/v1/workspaces/{CK_WORKSPACE}"
                   f"/projects/{_CK_FF_PROJECT}/tasks/{task_id}")
            resp = requests.put(url, headers=ck_headers, json={
                'name': work_area, 'estimate': _to_dur(budget_hrs),
                'status': 'ACTIVE', 'billable': True,
            })
            if resp.status_code != 200:
                ck_err = f"Clockify {resp.status_code}: {resp.text[:200]}"
                app.logger.error('budget_save clockify error: %s', ck_err)
        except Exception as exc:
            ck_err = str(exc)
            app.logger.error('budget_save clockify exception: %s', exc)

    return jsonify({'success': True, 'client': client,
                    'work_area': work_area, 'budget_hrs': budget_hrs,
                    'wb_sync': 'error' if wb_err else 'ok',
                    'ck_sync': 'skipped' if not task_id else ('error' if ck_err else 'ok'),
                    'wb_error': wb_err, 'ck_error': ck_err})


@app.route('/clockify/tasks', methods=['GET'])
@require_jwt
def clockify_tasks():
    """Return all tasks with current estimates for both Clockify projects."""
    import re as _re
    CK_API_KEY    = os.getenv('CLOCKIFY_API_KEY')
    CK_WORKSPACE  = os.getenv('CLOCKIFY_WORKSPACE_ID')
    FF_PROJECT    = os.getenv('CLOCKIFY_FIXED_FEE_PROJECT_ID')
    HR_PROJECT    = os.getenv('CLOCKIFY_HOURLY_PROJECT_ID')
    ck_headers    = {'X-Api-Key': CK_API_KEY}

    def _parse_dur(s):
        if not s or s == 'PT0S':
            return 0.0
        m = _re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+(?:\.\d+)?)S)?', s or '')
        if not m:
            return 0.0
        return round(int(m.group(1) or 0) + int(m.group(2) or 0) / 60, 4)

    def _get_tasks(project_id, billing_type):
        url  = f"https://api.clockify.me/api/v1/workspaces/{CK_WORKSPACE}/projects/{project_id}/tasks"
        resp = requests.get(url, headers=ck_headers, params={'limit': 100, 'status': 'ACTIVE'})
        return [
            {'task_id': t['id'], 'name': t['name'],
             'budget_hrs': _parse_dur(t.get('estimate')), 'billing_type': billing_type,
             'project_id': project_id}
            for t in resp.json()
        ]

    tasks = _get_tasks(FF_PROJECT, 'Fixed Fee') + _get_tasks(HR_PROJECT, 'Hourly')
    return jsonify({'tasks': tasks})


@app.route('/clockify/set-estimate', methods=['POST'])
@require_jwt
def clockify_set_estimate():
    """Set a task's hour estimate in Clockify. Body: {task_id, task_name, budget_hrs, project_id}"""
    import re as _re
    body       = request.get_json(force=True) or {}
    task_id    = str(body.get('task_id', '')).strip()
    task_name  = str(body.get('task_name', '')).strip()
    project_id = str(body.get('project_id', '')).strip()
    budget_hrs = body.get('budget_hrs')

    if not task_id or not task_name or not project_id or budget_hrs is None:
        return jsonify({'error': 'task_id, task_name, project_id, budget_hrs required'}), 400
    try:
        budget_hrs = float(budget_hrs)
    except (TypeError, ValueError):
        return jsonify({'error': 'budget_hrs must be a number'}), 400

    def _to_dur(hrs):
        total_m = round(hrs * 60)
        h, m = total_m // 60, total_m % 60
        return f"PT{h}H{m}M" if m else f"PT{h}H"

    CK_API_KEY   = os.getenv('CLOCKIFY_API_KEY')
    CK_WORKSPACE = os.getenv('CLOCKIFY_WORKSPACE_ID')
    ck_headers   = {'X-Api-Key': CK_API_KEY, 'Content-Type': 'application/json'}
    url = f"https://api.clockify.me/api/v1/workspaces/{CK_WORKSPACE}/projects/{project_id}/tasks/{task_id}"
    resp = requests.put(url, headers=ck_headers, json={
        'name': task_name, 'estimate': _to_dur(budget_hrs),
        'status': 'ACTIVE', 'billable': True,
    })
    if resp.status_code != 200:
        return jsonify({'error': f'Clockify API error {resp.status_code}', 'detail': resp.text}), 502

    return jsonify({'success': True, 'task_id': task_id, 'task_name': task_name,
                    'budget_hrs': budget_hrs, 'estimate': _to_dur(budget_hrs)})


# ---------------------------------------------------------------------------
# Document Pipeline  (/docs/*)
# ---------------------------------------------------------------------------

import json as _json
import sys as _sys

_DOC_STATE_PATH = '/home/jlobel/lac_automation/phase8/doc_pipeline_state.json'


def _doc_state_path():
    if has_request_context() and request.headers.get('X-Env') == 'test':
        try:
            if request.client_name == '__staff__':
                return '/home/jlobel/lac_automation/phase8/doc_pipeline_state_test.json'
        except Exception:
            pass
    return _DOC_STATE_PATH


def _load_doc_state():
    path = _doc_state_path()
    if os.path.exists(path):
        with open(path) as f:
            return _json.load(f)
    return {'deal_stages': {}, 'triggered': {}, 'pending': []}


def _save_doc_state(state):
    path = _doc_state_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w') as f:
        _json.dump(state, f, indent=2)


DOC_TYPE_LABELS = {
    'proposal':          'Proposal',
    'engagement_letter': 'Engagement Letter',
    'rep_letter':        'Rep Letter',
    'invoice_1':         'Invoice 1 (50% Deposit)',
    'invoice_2':         'Invoice 2 (Final)',
}


_DOC_TEMPLATE_LINKS = {
    'proposal':          'https://drive.google.com/file/d/1K0-xYkFguFWhjbRYzTqc7EOKV06tspO0/view',
    'engagement_letter': 'https://drive.google.com/file/d/1X2Dz7yrW1CePdIBxn3XzGC9wi01klGd-/view',
    'rep_letter':        'https://drive.google.com/file/d/1vuTfYZK4W-cw8vgdM5YTP9PpXljr3ifc/view',
}


@app.route('/docs/pending', methods=['GET'])
@require_jwt
def docs_pending():
    state = _load_doc_state()
    items = []
    for item in state.get('pending', []):
        data     = item.get('data', {})
        doc_type = item['doc_type']

        # Proposals and engagement letters show the full fee; invoices show their invoice amount
        if doc_type in ('proposal', 'engagement_letter', 'rep_letter'):
            try:
                amount = float(data.get('_total_fee') or 0) or None
            except (TypeError, ValueError):
                amount = None
        else:
            try:
                amount = float(data.get('INVOICE_AMOUNT') or 0) or None
            except (TypeError, ValueError):
                amount = None

        items.append({
            'id':            item['id'],
            'deal_id':       item.get('deal_id', ''),
            'client':        item['client'],
            'doc_type':      doc_type,
            'label':         DOC_TYPE_LABELS.get(doc_type, doc_type),
            'trigger':       item.get('trigger', ''),
            'created':       item.get('created', ''),
            'amount':        amount,
            'template_link': _DOC_TEMPLATE_LINKS.get(doc_type, ''),
        })
    return jsonify({'pending': items, 'count': len(items)})


@app.route('/docs/preview', methods=['POST'])
@require_jwt
def docs_preview():
    """Pre-generate draft, upload to Google Docs/Slides. Returns pdf_url + edit_url."""
    body = request.get_json(silent=True) or {}
    item_id = body.get('id')
    if not item_id:
        return jsonify({'error': 'id required'}), 400

    state = _load_doc_state()
    item  = next((p for p in state['pending'] if p['id'] == item_id), None)
    if not item:
        return jsonify({'error': 'item not found'}), 404

    # Return cached result if preview already exists
    if item.get('draft_pdf') and os.path.exists(item['draft_pdf']) and item.get('edit_url') is not None:
        return jsonify({
            'pdf_url':  f'/docs/pdf/{item_id}',
            'edit_url': item.get('edit_url', ''),
        })

    phase8_path = '/home/jlobel/lac_automation/phase8'
    if phase8_path not in _sys.path:
        _sys.path.insert(0, phase8_path)

    try:
        import doc_generator
        pdf_path, gid, edit_url = doc_generator.run_preview(item)
    except Exception as exc:
        app.logger.error('docs_preview error: %s', exc)
        return jsonify({'error': str(exc)}), 500

    item['draft_pdf']      = pdf_path
    item['google_file_id'] = gid or ''
    item['edit_url']       = edit_url or ''
    _save_doc_state(state)

    return jsonify({
        'pdf_url':  f'/docs/pdf/{item_id}',
        'edit_url': edit_url or '',
    })


@app.route('/docs/pdf/<item_id>', methods=['GET'])
def docs_pdf(item_id):
    """Serve pre-generated PDF for a pending item. JWT via ?token= or Authorization header."""
    from flask import send_file as _send_file
    token = request.args.get('token') or (request.headers.get('Authorization', '').replace('Bearer ', '').strip() or None)
    if not token:
        return jsonify({'error': 'Missing authorization header'}), 401
    try:
        jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except Exception:
        return jsonify({'error': 'Invalid token'}), 401
    state = _load_doc_state()
    item  = next((p for p in state['pending'] if p['id'] == item_id), None)
    if not item:
        item = state.get('recently_approved', {}).get(item_id)
    if not item:
        return jsonify({'error': 'item not found'}), 404
    pdf_path = item.get('draft_pdf')
    if not pdf_path or not os.path.exists(pdf_path):
        return jsonify({'error': 'preview not generated yet'}), 404

    # Persist document to client document store
    try:
        _store_client_doc(
            item['client'], item_id, item['doc_type'],
            DOC_TYPE_LABELS.get(item['doc_type'], item['doc_type']),
            pdf_path,
        )
    except Exception as exc:
        app.logger.warning('_store_client_doc failed: %s', exc)

    return _send_file(pdf_path, mimetype='application/pdf')


@app.route('/docs/approve', methods=['POST'])
@require_jwt
def docs_approve():
    body = request.get_json(silent=True) or {}
    item_id = body.get('id')
    if not item_id:
        return jsonify({'error': 'id required'}), 400

    state = _load_doc_state()
    item = next((p for p in state['pending'] if p['id'] == item_id), None)
    if not item:
        return jsonify({'error': 'item not found'}), 404

    phase8_path = '/home/jlobel/lac_automation/phase8'
    if phase8_path not in _sys.path:
        _sys.path.insert(0, phase8_path)

    try:
        import doc_generator

        gid = item.get('google_file_id')
        if gid and item['doc_type'] in ('proposal', 'engagement_letter', 'rep_letter'):
            # Export the current state of the Google Doc/Slides (captures any edits)
            os.makedirs('/tmp/lac_drafts', exist_ok=True)
            pdf_bytes = doc_generator.export_google_to_pdf(gid)
            edited_pdf = os.path.join('/tmp/lac_drafts', f'approved_{item_id}.pdf')
            with open(edited_pdf, 'wb') as f:
                f.write(pdf_bytes)
            doc_generator.send_approved(item, edited_pdf)
            pdf_path = edited_pdf
        else:
            pdf_path = doc_generator.run_generate(item)
    except Exception as exc:
        app.logger.error('docs_approve error: %s', exc)
        return jsonify({'error': str(exc)}), 500

    # Preserve generated PDF so /docs/pdf/ can serve it after the item leaves pending
    item['draft_pdf'] = pdf_path
    if 'recently_approved' not in state:
        state['recently_approved'] = {}
    state['recently_approved'][item_id] = item

    # Record sent date (used by rep_letter for ENGAGEMENT_LETTER_DATE)
    key = f"{item['deal_id']}|{item['doc_type']}"
    if key in state['triggered']:
        state['triggered'][key]['sent'] = datetime.now().date().isoformat()
        state['triggered'][key]['sent_date'] = datetime.now().strftime('%B %d, %Y')

    # If engagement_letter just sent, store date for future rep_letter
    if item['doc_type'] == 'engagement_letter':
        rep_key = f"{item['deal_id']}|rep_letter"
        if rep_key in state['triggered']:
            state['triggered'][rep_key]['eng_letter_date'] = datetime.now().strftime('%B %d, %Y')
        # Also pre-populate any queued rep_letter data
        for pending_item in state['pending']:
            if (pending_item['deal_id'] == item['deal_id'] and
                    pending_item['doc_type'] == 'rep_letter'):
                pending_item['data']['ENGAGEMENT_LETTER_DATE'] = (
                    datetime.now().strftime('%B %d, %Y'))

    # Remove from pending
    state['pending'] = [p for p in state['pending'] if p['id'] != item_id]
    _save_doc_state(state)

    _log_activity(item['client'], 'admin', 'Document Approved',
                  DOC_TYPE_LABELS.get(item['doc_type'], item['doc_type']))
    return jsonify({
        'success': True,
        'client': item['client'],
        'doc_type': item['doc_type'],
        'label': DOC_TYPE_LABELS.get(item['doc_type'], item['doc_type']),
        'pdf_url': f'/docs/pdf/{item_id}',
    })


@app.route('/docs/reject', methods=['POST'])
@require_jwt
def docs_reject():
    body = request.get_json(silent=True) or {}
    item_id = body.get('id')
    if not item_id:
        return jsonify({'error': 'id required'}), 400

    state = _load_doc_state()
    item = next((p for p in state['pending'] if p['id'] == item_id), None)
    if not item:
        return jsonify({'error': 'item not found'}), 404

    # Remove from pending
    state['pending'] = [p for p in state['pending'] if p['id'] != item_id]

    # Remove from triggered so the poller can re-queue if stage re-triggers
    tkey = f"{item['deal_id']}|{item['doc_type']}"
    state['triggered'].pop(tkey, None)

    _save_doc_state(state)
    _log_activity(item['client'], 'admin', 'Document Dismissed',
                  DOC_TYPE_LABELS.get(item['doc_type'], item['doc_type']))
    return jsonify({'success': True})


@app.route('/docs/send', methods=['POST'])
@require_jwt
def docs_send():
    """Send document for e-signature via DocuSeal."""
    body = request.get_json(silent=True) or {}
    item_id      = body.get('id')
    signer_email = body.get('signer_email', '')
    if not item_id:
        return jsonify({'error': 'id required'}), 400
    if not DOCUSEAL_API_KEY:
        return jsonify({'error': 'DocuSeal not configured — set DOCUSEAL_API_KEY in .env'}), 503
    state = _load_doc_state()
    item = next((p for p in state['pending'] if p['id'] == item_id), None)
    if not item:
        return jsonify({'error': 'item not found'}), 404
    return jsonify({
        'error': 'DocuSeal setup required — visit https://sign.lobelaccountancy.com to complete configuration'
    }), 503


# ---------------------------------------------------------------------------
# Engagement Pipeline — list + DocuSeal engagement letter
# ---------------------------------------------------------------------------

def _parse_pipeline_full(wb):
    """Return all engagement pipeline rows with full field detail."""
    if 'Engagement Pipeline' not in wb.sheetnames:
        return []
    ws = wb['Engagement Pipeline']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        client = row[0] if len(row) > 0 else None
        if not client:
            continue
        def _s(v):
            return str(v).strip() if v is not None else ''
        def _f(v):
            try: return float(v)
            except: return 0.0
        rows.append({
            'client':       _s(row[0]),
            'client_num':   _s(row[1]),
            'stage':        _s(row[2]),
            'eng_type':     _s(row[3]),
            'billing_type': _s(row[4]),
            'entity_type':  _s(row[5]),
            'fee':          _f(row[6]),
            'status':       _s(row[7]),
            'start_date':   _s(row[8]),
            'issue_date':   _s(row[9]),
            'fye':          _s(row[10]),
            'client_title': _s(row[11]),
            'email':        _s(row[12]),
            'address':      _s(row[13]),
        })
    return rows


@app.route('/engagement/list', methods=['GET'])
@require_jwt
def engagement_list():
    try:
        wb = _workbook()
        return jsonify(_parse_pipeline_full(wb))
    except Exception as e:
        return jsonify({'error': str(e)}), 503


@app.route('/engagement/templates', methods=['GET'])
@require_jwt
def engagement_templates():
    """List DocuSeal templates."""
    if not DOCUSEAL_API_KEY:
        return jsonify({'error': 'DOCUSEAL_API_KEY not set'}), 503
    try:
        import urllib.request as _ur
        req = _ur.Request(
            f'{DOCUSEAL_URL}/api/templates',
            headers={'X-Auth-Token': DOCUSEAL_API_KEY}
        )
        with _ur.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read())
        return jsonify(data.get('data', []))
    except Exception as e:
        return jsonify({'error': str(e)}), 503


@app.route('/engagement/send-letter', methods=['POST'])
@require_jwt
def engagement_send_letter():
    """Create a DocuSeal submission from an engagement letter template,
    pre-filled with data from the Engagement Pipeline, and email the client."""
    if not DOCUSEAL_API_KEY:
        return jsonify({'error': 'DOCUSEAL_API_KEY not set'}), 503

    body        = request.get_json(silent=True) or {}
    client_name = body.get('client', '').strip()
    template_id = body.get('template_id')

    if not client_name:
        return jsonify({'error': 'client required'}), 400
    if not template_id:
        return jsonify({'error': 'template_id required'}), 400

    # Look up client in Engagement Pipeline
    try:
        wb = _workbook()
    except Exception as e:
        return jsonify({'error': f'Could not load workbook: {e}'}), 503

    rows = _parse_pipeline_full(wb)
    eng = next((r for r in rows if r['client'].lower() == client_name.lower()), None)
    if not eng:
        return jsonify({'error': f'Client "{client_name}" not found in Engagement Pipeline'}), 404

    if not eng['email']:
        return jsonify({'error': f'No email address for {client_name} in Engagement Pipeline'}), 400

    # Format fee as dollar string
    fee_str = f"${eng['fee']:,.2f}" if eng['fee'] else ''

    payload = {
        'template_id': int(template_id),
        'send_email':  True,
        'submitters': [{
            'role':  'Client',
            'email': eng['email'],
            'name':  eng['client_title'] or eng['client'],
            'fields': [
                {'name': 'Client Name',      'default_value': eng['client'],       'readonly': True},
                {'name': 'Client Title',     'default_value': eng['client_title'], 'readonly': True},
                {'name': 'Client Address',   'default_value': eng['address'],      'readonly': True},
                {'name': 'Engagement Type',  'default_value': eng['eng_type'],     'readonly': True},
                {'name': 'Entity Type',      'default_value': eng['entity_type'],  'readonly': True},
                {'name': 'Fee Amount',       'default_value': fee_str,             'readonly': True},
                {'name': 'Fiscal Year End',  'default_value': eng['fye'],          'readonly': True},
                {'name': 'Issuance Date',    'default_value': eng['issue_date'],   'readonly': True},
            ]
        }]
    }

    try:
        import urllib.request as _ur
        data = _json.dumps(payload).encode()
        req = _ur.Request(
            f'{DOCUSEAL_URL}/api/submissions',
            data=data,
            headers={
                'X-Auth-Token':  DOCUSEAL_API_KEY,
                'Content-Type':  'application/json',
            },
            method='POST'
        )
        with _ur.urlopen(req, timeout=15) as resp:
            result = _json.loads(resp.read())
        return jsonify({'ok': True, 'submission': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 503


@app.route('/engagement/create-template', methods=['POST'])
@require_jwt
def engagement_create_template():
    """Accept .docx/.pdf, convert if needed, create DocuSeal template via Rails runner."""
    import subprocess as _sp, tempfile as _tf, uuid as _uuid

    if not DOCUSEAL_API_KEY:
        return jsonify({'error': 'DOCUSEAL_API_KEY not set'}), 503
    if 'file' not in request.files:
        return jsonify({'error': 'file required'}), 400

    f = request.files['file']
    name = (request.form.get('name') or f.filename.rsplit('.', 1)[0] or 'Engagement Letter').strip()
    ext  = os.path.splitext(f.filename)[1].lower()

    with _tf.TemporaryDirectory() as tmp:
        src = os.path.join(tmp, 'input' + ext)
        f.save(src)

        if ext in ('.docx', '.doc', '.odt', '.rtf', '.pptx', '.ppt', '.odp'):
            res = _sp.run(
                ['libreoffice', '--headless', '--convert-to', 'pdf', src, '--outdir', tmp],
                capture_output=True, text=True, timeout=90
            )
            pdf_path = os.path.join(tmp, 'input.pdf')
            if not os.path.exists(pdf_path):
                return jsonify({'error': f'PDF conversion failed: {res.stderr[:300]}'}), 500
        elif ext == '.pdf':
            pdf_path = src
        else:
            return jsonify({'error': f'Unsupported file type: {ext}'}), 400

        # Copy PDF into DocuSeal container
        container_pdf = f'/tmp/tpl_{_uuid.uuid4().hex}.pdf'
        cp = _sp.run(['docker', 'cp', pdf_path, f'docuseal:{container_pdf}'],
                     capture_output=True, text=True)
        if cp.returncode != 0:
            return jsonify({'error': f'docker cp failed: {cp.stderr}'}), 500

        # Create template + attach document via Rails runner
        escaped_name = name.replace("'", "\\'")
        rails = f"""
begin
  user = User.first
  tpl  = Template.new(name: '{escaped_name}', author: user, account: user.account)
  tpl.submitters  = [{{'name' => 'Client', 'uuid' => SecureRandom.uuid}}]
  tpl.schema      = []
  tpl.fields      = []
  tpl.save!
  blob = ActiveStorage::Blob.create_and_upload!(
    io: File.open('{container_pdf}', 'rb'),
    filename: '{escaped_name}.pdf',
    content_type: 'application/pdf',
    metadata: {{identified: true, analyzed: true, pdf: {{annotations: []}}}}
  )
  doc = tpl.documents.create!(blob: blob)
  tpl.update!(schema: [{{'attachment_uuid' => doc.uuid, 'name' => '{escaped_name}'}}])
  Templates::ProcessDocument.call(doc, File.read('{container_pdf}', mode: 'rb'), extract_fields: true)
  File.delete('{container_pdf}') rescue nil
  puts tpl.id
rescue => e
  STDERR.puts e.message
  exit 1
end
"""
        run = _sp.run(
            ['docker', 'exec', '-w', '/app', 'docuseal', 'bundle', 'exec', 'rails', 'runner', rails],
            capture_output=True, text=True, timeout=120
        )
        if run.returncode != 0:
            return jsonify({'error': f'Template creation failed: {run.stderr[:400]}'}), 500

        template_id = run.stdout.strip().split('\n')[-1]
        return jsonify({
            'ok':          True,
            'template_id': int(template_id),
            'edit_url':    f'{DOCUSEAL_URL}/templates/{template_id}/edit',
            'name':        name,
        })


# ---------------------------------------------------------------------------
# AI Workbook API — natural-language reads/writes for Open WebUI tools
# All endpoints accept either a user JWT or the service JWT (role=service).
# ---------------------------------------------------------------------------

def _next_client_num(ws):
    """Return the next available 4-digit client number from Engagement Pipeline."""
    used = set()
    for row in ws.iter_rows(min_row=_EP_DATA_START, values_only=True):
        try:
            used.add(int(row[_EP_COL['client_num']]))
        except (TypeError, ValueError):
            pass
    n = 1000
    while n in used:
        n += 1
    return n

def _find_ep_row(ws, client_name):
    """Return the (row_index, row_values) for a client in Engagement Pipeline, or None."""
    for i, row in enumerate(ws.iter_rows(min_row=_EP_DATA_START, values_only=False), start=_EP_DATA_START):
        val = row[_EP_COL['client']].value
        if val and str(val).strip().lower() == client_name.strip().lower():
            return i, row
    return None, None

def _find_ar_row(ws, client_name=None, invoice_num=None):
    """Return (row_index, row_cells) for an AR row by client name or invoice number."""
    for i, row in enumerate(ws.iter_rows(min_row=_AR_DATA_START, values_only=False), start=_AR_DATA_START):
        if invoice_num:
            val = row[_AR_COL['invoice']].value
            if val and str(val).strip().lower() == invoice_num.strip().lower():
                return i, row
        elif client_name:
            val = row[_AR_COL['client']].value
            if val and str(val).strip().lower() == client_name.strip().lower():
                return i, row
    return None, None


@app.route('/wb/clients', methods=['GET'])
@require_jwt
def wb_clients():
    """List all clients in Engagement Pipeline and AR Aging."""
    wb = _workbook()
    pipeline = []
    if 'Engagement Pipeline' in wb.sheetnames:
        for row in wb['Engagement Pipeline'].iter_rows(min_row=_EP_DATA_START, values_only=True):
            c = row[_EP_COL['client']]
            if c:
                pipeline.append({
                    'client':   str(c),
                    'stage':    str(row[_EP_COL['stage']] or ''),
                    'status':   str(row[_EP_COL['status']] or ''),
                    'eng_type': str(row[_EP_COL['eng_type']] or ''),
                    'email':    str(row[_EP_COL['email']] or ''),
                })
    ar = []
    if 'AR Aging' in wb.sheetnames:
        for row in wb['AR Aging'].iter_rows(min_row=_AR_DATA_START, values_only=True):
            c = row[_AR_COL['client']]
            if c:
                ar.append({
                    'client':  str(c),
                    'invoice': str(row[_AR_COL['invoice']] or ''),
                    'amount':  row[_AR_COL['amount']],
                    'status':  str(row[_AR_COL['status']] or ''),
                })
    return jsonify({'pipeline': pipeline, 'ar': ar})


@app.route('/wb/prospect/add', methods=['POST'])
@require_jwt
def wb_prospect_add():
    """Add a new row to Engagement Pipeline (prospect or engaged)."""
    body = request.get_json(silent=True) or {}
    client = (body.get('client') or '').strip()
    if not client:
        return jsonify({'error': 'client name required'}), 400

    wb, fid, svc = _wb_download_writable()
    ws = wb['Engagement Pipeline']

    # Check for duplicate
    row_i, _ = _find_ep_row(ws, client)
    if row_i:
        return jsonify({'error': f'"{client}" already exists in Engagement Pipeline'}), 409

    client_num = _next_client_num(ws)
    today_str  = date.today().strftime('%m/%d/%Y')

    new_row = [''] * 14
    new_row[_EP_COL['client']]       = client
    new_row[_EP_COL['client_num']]   = client_num
    new_row[_EP_COL['stage']]        = body.get('stage', 'Prospect')
    new_row[_EP_COL['eng_type']]     = body.get('eng_type', '')
    new_row[_EP_COL['billing_type']] = body.get('billing_type', 'Fixed Fee')
    new_row[_EP_COL['entity_type']]  = body.get('entity_type', '')
    new_row[_EP_COL['fee']]          = float(body.get('fee', 0) or 0)
    new_row[_EP_COL['status']]       = body.get('status', 'Active')
    new_row[_EP_COL['start_date']]   = body.get('start_date', today_str)
    new_row[_EP_COL['issue_date']]   = body.get('issue_date', '')
    new_row[_EP_COL['fye']]          = body.get('fye', '')
    new_row[_EP_COL['client_title']] = body.get('client_title', '')
    new_row[_EP_COL['email']]        = body.get('email', '')
    new_row[_EP_COL['address']]      = body.get('address', '')

    # Find first empty row after last client data (avoid appending to row 1000+)
    last_data_row = _EP_DATA_START - 1
    for i, row in enumerate(ws.iter_rows(min_row=_EP_DATA_START, values_only=True), start=_EP_DATA_START):
        if row and row[_EP_COL['client']]:
            last_data_row = i
    target_row = last_data_row + 1
    for col_idx, value in enumerate(new_row, start=1):
        ws.cell(row=target_row, column=col_idx).value = value

    _wb_save_and_upload(wb, fid, svc)
    return jsonify({'ok': True, 'message': f'Added "{client}" (#{client_num}) as {new_row[_EP_COL["stage"]]} to Engagement Pipeline.'})


@app.route('/wb/engagement/update', methods=['POST'])
@require_jwt
def wb_engagement_update():
    """Update fields on an existing Engagement Pipeline row."""
    body   = request.get_json(silent=True) or {}
    client = (body.get('client') or '').strip()
    updates = body.get('updates') or {}
    if not client:
        return jsonify({'error': 'client required'}), 400
    if not updates:
        return jsonify({'error': 'updates dict required'}), 400

    wb, fid, svc = _wb_download_writable()
    ws  = wb['Engagement Pipeline']
    row_i, row = _find_ep_row(ws, client)
    if row_i is None:
        return jsonify({'error': f'"{client}" not found in Engagement Pipeline'}), 404

    changed = []
    for field, col in _EP_COL.items():
        if field in updates:
            old = row[col].value
            row[col].value = updates[field]
            changed.append(f'{field}: {old!r} → {updates[field]!r}')

    _wb_save_and_upload(wb, fid, svc)
    summary = ', '.join(changed) if changed else 'no changes'
    return jsonify({'ok': True, 'message': f'Updated "{client}": {summary}.'})


@app.route('/wb/ar/add', methods=['POST'])
@require_jwt
def wb_ar_add():
    """Add a new row to AR Aging."""
    body = request.get_json(silent=True) or {}
    client = (body.get('client') or '').strip()
    invoice = (body.get('invoice') or '').strip()
    amount  = float(body.get('amount', 0) or 0)
    if not client or not invoice or not amount:
        return jsonify({'error': 'client, invoice, and amount required'}), 400

    wb, fid, svc = _wb_download_writable()
    ws_ar = wb['AR Aging']

    # Look up client number from Engagement Pipeline
    client_num = ''
    if 'Engagement Pipeline' in wb.sheetnames:
        _, ep_row = _find_ep_row(wb['Engagement Pipeline'], client)
        if ep_row:
            client_num = ep_row[_EP_COL['client_num']].value or ''

    today_str  = date.today().strftime('%m/%d/%Y')
    service    = body.get('service', 'Professional Services')
    paid       = float(body.get('paid', 0) or 0)
    outstanding = amount - paid
    status     = 'Paid' if outstanding <= 0 else body.get('status', 'Unpaid')

    new_row = [''] * 14
    new_row[_AR_COL['client']]      = client
    new_row[_AR_COL['client_num']]  = client_num
    new_row[_AR_COL['invoice']]     = invoice
    new_row[_AR_COL['inv_date']]    = body.get('inv_date', today_str)
    new_row[_AR_COL['due_date']]    = body.get('due_date', today_str)
    new_row[_AR_COL['service']]     = service
    new_row[_AR_COL['amount']]      = amount
    new_row[_AR_COL['paid']]        = paid
    new_row[_AR_COL['outstanding']] = outstanding
    new_row[_AR_COL['days']]        = 0
    new_row[_AR_COL['status']]      = status
    new_row[_AR_COL['email']]       = body.get('email', '')
    new_row[_AR_COL['address']]     = body.get('address', '')
    new_row[_AR_COL['reminder']]    = ''

    ws_ar.append(new_row)
    _wb_save_and_upload(wb, fid, svc)
    return jsonify({'ok': True, 'message': f'Added AR entry {invoice} for "{client}" — ${amount:,.2f} {status}.'})


@app.route('/wb/ar/update', methods=['POST'])
@require_jwt
def wb_ar_update():
    """Update an AR Aging row (e.g. mark paid, change status)."""
    body        = request.get_json(silent=True) or {}
    client      = (body.get('client') or '').strip()
    invoice_num = (body.get('invoice') or '').strip()
    updates     = body.get('updates') or {}
    if not (client or invoice_num):
        return jsonify({'error': 'client or invoice required'}), 400

    wb, fid, svc = _wb_download_writable()
    ws = wb['AR Aging']
    row_i, row = _find_ar_row(ws, client_name=client or None, invoice_num=invoice_num or None)
    if row_i is None:
        return jsonify({'error': f'AR row not found for {client or invoice_num}'}), 404

    # Convenience: if paid is set, auto-compute outstanding and status
    if 'paid' in updates:
        paid   = float(updates['paid'])
        amount = float(row[_AR_COL['amount']].value or 0)
        outstanding = max(0, amount - paid)
        updates.setdefault('outstanding', outstanding)
        updates.setdefault('status', 'Paid' if outstanding == 0 else 'Partial')

    changed = []
    for field, col in _AR_COL.items():
        if field in updates:
            old = row[col].value
            row[col].value = updates[field]
            changed.append(f'{field}: {old!r} → {updates[field]!r}')

    _wb_save_and_upload(wb, fid, svc)
    summary = ', '.join(changed) if changed else 'no changes'
    invoice_label = invoice_num or str(row[_AR_COL['invoice']].value or '')
    return jsonify({'ok': True, 'message': f'Updated AR {invoice_label} for "{client or invoice_num}": {summary}.'})


# ---------------------------------------------------------------------------
# Email — Zoho IMAP
# ---------------------------------------------------------------------------

MAIL_ALIASES = {
    'jlobel':  'jlobel@lobelaccountancy.com',
    'info':    'info@lobelaccountancy.com',
    'billing': 'billing@lobelaccountancy.com',
}

_EMAIL_CACHE = {'data': None, 'ts': 0, 'error': None}
_EMAIL_TTL   = 120   # seconds


def _hdr(raw):
    """Decode a potentially RFC-2047-encoded mail header to plain text."""
    if not raw:
        return ''
    parts = _decode_mime_header(raw)
    out = []
    for part, enc in parts:
        if isinstance(part, bytes):
            out.append(part.decode(enc or 'utf-8', errors='replace'))
        else:
            out.append(str(part))
    return ' '.join(out).strip()


def _body_snippet(msg, max_len=400):
    """Extract plain-text snippet from an email.Message; strip excess whitespace."""
    candidates = [msg] if not msg.is_multipart() else list(msg.walk())
    for part in candidates:
        if part.get_content_type() != 'text/plain':
            continue
        if 'attachment' in str(part.get('Content-Disposition', '')):
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue
        charset = part.get_content_charset() or 'utf-8'
        text = payload.decode(charset, errors='replace')
        text = re.sub(r'\s+', ' ', text).strip()
        return text[:max_len]
    return ''


def _fetch_zoho_emails():
    """Return (emails_list, error_str). Cached for _EMAIL_TTL seconds."""
    now = time.time()
    if _EMAIL_CACHE['data'] is not None and now - _EMAIL_CACHE['ts'] < _EMAIL_TTL:
        return _EMAIL_CACHE['data'], _EMAIL_CACHE['error']

    import socket
    old_timeout = socket.getdefaulttimeout()
    socket.setdefaulttimeout(25)
    try:
        imap = imaplib.IMAP4_SSL('imap.zoho.com', 993)
        imap.login(ZOHO_EMAIL, ZOHO_PASSWORD)
        imap.select('INBOX', readonly=True)

        _, data = imap.search(None, 'UNSEEN')
        all_ids = data[0].split()

        # Fetch the 60 most recent unread, newest first — single batch round-trip
        recent = list(reversed(all_ids[-60:]))

        emails = []
        if recent:
            uid_str = b','.join(recent)
            _, raw_list = imap.fetch(uid_str, '(BODY.PEEK[])')
            for item in (raw_list or []):
                if not isinstance(item, tuple):
                    continue
                try:
                    msg = _email_pkg.message_from_bytes(item[1])

                    from_raw = _hdr(msg.get('From', ''))
                    to_raw   = _hdr(msg.get('To', ''))
                    subject  = _hdr(msg.get('Subject', '')) or '(no subject)'
                    date_raw = msg.get('Date', '')

                    from_name, from_addr = _email_utils.parseaddr(from_raw)
                    from_display = from_name.strip() or from_addr

                    # Format date: today → time, this week → weekday, older → Mon D
                    try:
                        tup = _email_utils.parsedate_tz(date_raw)
                        ts  = _email_utils.mktime_tz(tup) if tup else None
                        if ts:
                            dt    = datetime.fromtimestamp(ts)
                            today = datetime.now().date()
                            delta = (today - dt.date()).days
                            if delta == 0:
                                date_str = dt.strftime('%-I:%M %p')
                            elif delta < 7:
                                date_str = dt.strftime('%a')
                            else:
                                date_str = dt.strftime('%b %-d')
                        else:
                            date_str = date_raw[:10]
                    except Exception:
                        date_str = date_raw[:10]

                    message_id   = (msg.get('Message-ID') or '').strip()
                    reply_to_raw = _hdr(msg.get('Reply-To', '')) or from_raw
                    _, reply_to_addr = _email_utils.parseaddr(reply_to_raw)

                    emails.append({
                        'from_name':  from_display,
                        'from_email': from_addr,
                        'reply_to':   reply_to_addr or from_addr,
                        'to':         to_raw,
                        'subject':    subject,
                        'date':       date_str,
                        'snippet':    _body_snippet(msg),
                        'body':       _body_snippet(msg, max_len=1500),
                        'message_id': message_id,
                    })
                except Exception:
                    continue

        imap.logout()
        _EMAIL_CACHE.update({'data': emails, 'ts': now, 'error': None})
        return emails, None

    except Exception as exc:
        err = str(exc)
        _EMAIL_CACHE['error'] = err
        return _EMAIL_CACHE.get('data') or [], err
    finally:
        socket.setdefaulttimeout(old_timeout)


@app.route('/data/emails')
@require_jwt
def get_emails():
    alias_key = request.args.get('alias', 'jlobel')
    alias_addr = MAIL_ALIASES.get(alias_key, MAIL_ALIASES['jlobel'])

    emails, error = _fetch_zoho_emails()

    # Per-alias unread counts
    counts = {}
    for key, addr in MAIL_ALIASES.items():
        counts[key] = sum(1 for e in emails if addr.lower() in e['to'].lower())

    # Filter to the requested alias
    filtered = [e for e in emails if alias_addr.lower() in e['to'].lower()]

    return jsonify({
        'alias':        alias_addr,
        'emails':       filtered,
        'counts':       counts,
        'total_unread': len(emails),
        'error':        error,
    })


# ---------------------------------------------------------------------------
# Data: Google Drive browser + upload
# ---------------------------------------------------------------------------

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024   # 50 MB upload cap


_drive_browse_cache: dict = {}
_DRIVE_BROWSE_TTL = 60  # seconds


@app.route('/data/drive', methods=['GET'])
@require_jwt
def drive_list():
    """List Drive files/folders accessible to the service account.

    Params:
      folder_id / folder: folder to list (default 'root' = top-level)
      page_token:         pagination token
      q:                  search string (name contains)
    """
    folder_id  = (request.args.get('folder_id') or request.args.get('folder') or 'root').strip()
    page_token = (request.args.get('page_token') or '').strip()
    search_q   = (request.args.get('q') or '').strip()

    cache_key = f'{folder_id}|{page_token}|{search_q}'
    cached = _drive_browse_cache.get(cache_key)
    if cached and time.time() - cached['ts'] < _DRIVE_BROWSE_TTL:
        return jsonify(cached['data'])

    try:
        svc = _drive_service()

        if search_q:
            safe_q = search_q.replace("'", "\\'")
            q = f"name contains '{safe_q}' and trashed=false"
        elif folder_id == 'root':
            q = 'trashed=false'
        else:
            q = f"'{folder_id}' in parents and trashed=false"

        params = dict(
            q=q,
            fields='nextPageToken,files(id,name,mimeType,modifiedTime,size,webViewLink,parents)',
            orderBy='folder,name',
            pageSize=100,
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            corpora='allDrives',
        )
        if page_token:
            params['pageToken'] = page_token

        result = svc.files().list(**params).execute()
        files  = result.get('files', [])

        # For root, keep only top-level items (parents not in the result set)
        if folder_id == 'root' and not search_q and not page_token:
            ids   = {f['id'] for f in files}
            files = [f for f in files if not any(p in ids for p in f.get('parents', []))]

        folder_name = 'My Drive'
        if folder_id != 'root':
            try:
                info = svc.files().get(
                    fileId=folder_id,
                    fields='id,name',
                    supportsAllDrives=True,
                ).execute()
                folder_name = info.get('name', folder_id)
            except Exception:
                folder_name = folder_id

        data = {
            'files':           files,
            'folder_id':       folder_id,
            'folder_name':     folder_name,
            'next_page_token': result.get('nextPageToken'),
            'total':           len(files),
        }
        _drive_browse_cache[cache_key] = {'data': data, 'ts': time.time()}
        return jsonify(data)
    except Exception as exc:
        app.logger.error('drive_list error: %s', exc)
        return jsonify({'error': str(exc)}), 503


@app.route('/data/drive/upload', methods=['POST'])
@require_jwt
def drive_upload():
    import io as _io
    from googleapiclient.http import MediaIoBaseUpload as _MU

    folder_id = request.form.get('folder_id', 'root')
    f         = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'No file provided'}), 400

    try:
        svc       = _drive_rw_service()
        buf       = _io.BytesIO(f.read())
        mime      = f.content_type or 'application/octet-stream'
        media     = _MU(buf, mimetype=mime, resumable=False)
        metadata  = {'name': f.filename, 'parents': [folder_id]}

        created = svc.files().create(
            body=metadata,
            media_body=media,
            fields='id,name,webViewLink,mimeType',
        ).execute()

        return jsonify({
            'success':  True,
            'id':       created['id'],
            'name':     created['name'],
            'web_link': created.get('webViewLink', ''),
        })
    except Exception as exc:
        app.logger.error('drive_upload error: %s', exc)
        return jsonify({'error': str(exc)}), 500


_EMAIL_SIGNATURE = (
    "\n\n--\n"
    "Jeffrey Lobel, CPA\n"
    "Lobel Accountancy Corporation\n"
    "(949) 345-1925\n"
    "jlobel@lobelaccountancy.com"
)


@app.route('/data/email/reply', methods=['POST'])
@require_jwt
def email_reply():
    body        = request.get_json(silent=True) or {}
    to_addr     = (body.get('to')         or '').strip()
    subject     = (body.get('subject')    or '').strip()
    reply_body  = (body.get('body')       or '').strip()
    in_reply_to = (body.get('in_reply_to') or '').strip()
    alias_key   = (body.get('alias')      or 'jlobel')

    if not to_addr or not reply_body:
        return jsonify({'error': 'to and body are required'}), 400

    from_addr = MAIL_ALIASES.get(alias_key, MAIL_ALIASES['jlobel'])
    subj      = subject if subject.lower().startswith('re:') else f'Re: {subject}'

    msg = MIMEText(reply_body, 'plain')
    msg['Subject'] = subj
    msg['From']    = from_addr
    msg['To']      = to_addr
    if in_reply_to:
        msg['In-Reply-To'] = in_reply_to
        msg['References']  = in_reply_to

    try:
        with smtplib.SMTP_SSL('smtp.zoho.com', 465) as server:
            server.login(ZOHO_EMAIL, ZOHO_PASSWORD)
            server.sendmail(from_addr, to_addr, msg.as_string())
    except Exception as exc:
        app.logger.error('email_reply failed: %s', exc)
        return jsonify({'error': str(exc)}), 500

    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# Financials — P&L and Balance Sheet from LAC Workbook
# ---------------------------------------------------------------------------

MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

COMPANY_START_MONTH = 5  # May — company inception

def _available_months():
    """Return month indices from May through the current month.
    Past months require close+5d; the current month is always included."""
    today = date.today()
    available = []
    for m in range(COMPANY_START_MONTH, today.month + 1):
        if m == today.month:
            available.append(m)  # always show current month
            continue
        month_end = date(today.year, m + 1, 1) - timedelta(days=1)
        if today >= month_end + timedelta(days=5):
            available.append(m)
    return available

_EXCEL_EPOCH = date(1899, 12, 30)

def _build_txn_data(wb):
    """Read Transactions sheet and return {(acct_int, month_int): {'debit': float, 'credit': float}}.
    Google Sheets exports don't cache formula results, so we compute the month from the raw date."""
    if 'Transactions' not in wb.sheetnames:
        return {}
    ws = wb['Transactions']
    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        date_val, _, _, acct, _, debit, credit = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if acct is None:
            continue
        # Resolve date to month
        if isinstance(date_val, (datetime, date)):
            month = date_val.month
        elif isinstance(date_val, (int, float)):
            # Excel serial date: days since 1899-12-30
            try:
                d = _EXCEL_EPOCH + timedelta(days=int(date_val))
                month = d.month
            except Exception:
                continue
        elif isinstance(date_val, str):
            d = _to_date(date_val)
            if d is None:
                continue
            month = d.month
        else:
            continue
        try:
            acct_int = int(float(acct))
        except (ValueError, TypeError):
            continue
        key = (acct_int, month)
        entry = data.setdefault(key, {'debit': 0.0, 'credit': 0.0})
        if debit:
            entry['debit'] += float(debit)
        if credit:
            entry['credit'] += float(credit)
    return data


def _txn_value(txn_data, acct_int, month):
    """Return the signed value for an account in a given month.
    Revenue (4xxx): net credit.  Expense (5xxx-8xxx): net debit.
    Asset (1xxx): net debit.  Liability/Equity (2xxx-3xxx): net credit."""
    entry = txn_data.get((acct_int, month), {})
    debit  = entry.get('debit',  0.0)
    credit = entry.get('credit', 0.0)
    if 4000 <= acct_int <= 4999:
        return credit - debit
    elif 1000 <= acct_int <= 1999:
        return debit - credit
    elif 2000 <= acct_int <= 3999:
        return credit - debit
    else:
        return debit - credit


def _txn_cumulative(txn_data, acct_int, through_month):
    """Running balance from inception (month 1) through through_month.
    Used for balance sheet accounts so each period shows the full balance,
    not just that month's activity."""
    total_debit  = sum(txn_data.get((acct_int, m), {}).get('debit',  0.0) for m in range(1, through_month + 1))
    total_credit = sum(txn_data.get((acct_int, m), {}).get('credit', 0.0) for m in range(1, through_month + 1))
    if 4000 <= acct_int <= 4999:
        return total_credit - total_debit
    elif 1000 <= acct_int <= 1999:
        return total_debit  - total_credit
    elif 2000 <= acct_int <= 3999:
        return total_credit - total_debit
    else:
        return total_debit  - total_credit


def _parse_financials_sheet(ws, available_months, txn_data=None, bs_mode=False):
    """
    Parse a P&L or Balance Sheet worksheet into structured sections.
    Returns list of {label, is_section, is_total, months: {idx: value}}.
    When formula cached values are absent (Google Sheets export), falls back
    to computing from txn_data built from the Transactions sheet.
    """
    rows = []
    section_accounts = []  # track account rows for current section total accumulation

    for row in ws.iter_rows(min_row=4, values_only=True):
        label = row[0]
        if label is None:
            continue
        label = str(label).strip()
        if not label:
            continue

        is_section = label.isupper() and all(
            (row[i] is None or row[i] == 0) for i in range(1, 13)
            if i <= len(row) - 1
        ) and not label.startswith('TOTAL') and 'NET ' not in label
        is_total = (label.upper().startswith('TOTAL') or
                    label.startswith('Net Income') or
                    label.startswith('NET INCOME') or
                    label == 'Balance Check [should be $0]:')

        # Try to extract account number from label prefix
        acct_match = re.match(r'^(\d{4})', label)
        acct_int = int(acct_match.group(1)) if acct_match else None

        month_vals = {}
        for mi in available_months:
            # First try cached formula value from sheet
            raw = row[mi] if mi < len(row) else None
            if raw is not None and raw != 0:
                v = round(float(raw), 2)
            elif txn_data is not None and acct_int:
                # Balance sheet needs the running balance (all months through mi),
                # P&L needs only the current month's activity.
                if bs_mode:
                    v = round(_txn_cumulative(txn_data, acct_int, mi), 2)
                else:
                    v = round(_txn_value(txn_data, acct_int, mi), 2)
            elif txn_data is not None and is_total:
                # For total rows, sum the section accounts accumulated so far
                if bs_mode:
                    v = round(sum(_txn_cumulative(txn_data, a, mi) for a in section_accounts), 2)
                else:
                    v = round(sum(_txn_value(txn_data, a, mi) for a in section_accounts), 2)
            else:
                v = 0.0
            month_vals[MONTHS[mi - 1]] = v

        if is_section:
            section_accounts = []  # start fresh accumulator for this section
        elif acct_int and not is_total:
            section_accounts.append(acct_int)
        elif is_total:
            section_accounts = []  # reset after total row

        rows.append({
            'label':      label,
            'is_section': is_section,
            'is_total':   is_total,
            'months':     month_vals,
        })
    return rows


def _group_txn_by_month(txn_data):
    """Single-pass grouping: returns (rev_by_month, exp_by_month) dicts keyed by month int."""
    rev_by_month = defaultdict(float)
    exp_by_month = defaultdict(float)
    for (a, m), v in txn_data.items():
        if 4000 <= a <= 4999:
            rev_by_month[m] += v['credit'] - v['debit']
        elif 5000 <= a <= 9999:
            exp_by_month[m] += v['debit'] - v['credit']
    return rev_by_month, exp_by_month


def _monthly_net_income(txn_data, available_months):
    """Net income per month: revenue (4xxx credits) minus expenses (5xxx-9xxx debits)."""
    rev, exp = _group_txn_by_month(txn_data)
    return {MONTHS[mi - 1]: round(rev[mi] - exp[mi], 2) for mi in available_months}


def _ytd_net_income(txn_data, available_months):
    """Cumulative YTD net income through each available month."""
    rev, exp = _group_txn_by_month(txn_data)
    cumulative = 0.0
    result = {}
    for mi in range(1, max(available_months) + 1):
        cumulative += rev[mi] - exp[mi]
        if mi in available_months:
            result[MONTHS[mi - 1]] = round(cumulative, 2)
    return result


def _patch_pl_totals(pl_rows, txn_data, available_months):
    """Fix TOTAL EXPENSES and NET INCOME rows whose multi-level sums compute to 0."""
    monthly_ni = _monthly_net_income(txn_data, available_months)
    for row in pl_rows:
        lbl_up = row['label'].upper()
        if lbl_up.startswith('TOTAL EXPENSE'):
            for mi in available_months:
                mo = MONTHS[mi - 1]
                row['months'][mo] = round(
                    sum(v['debit'] - v['credit']
                        for (a, m), v in txn_data.items()
                        if m == mi and 5000 <= a <= 9999), 2)
        elif lbl_up.startswith('NET INCOME') or lbl_up.startswith('NET LOSS'):
            for mo, v in monthly_ni.items():
                row['months'][mo] = v


def _patch_bs_equity(bs_rows, txn_data, available_months):
    """Inject YTD net income into equity section so balance sheet foots to zero."""
    ytd_ni = _ytd_net_income(txn_data, available_months)
    in_equity = False
    equity_items = []
    totals = {}

    for row in bs_rows:
        lbl = row['label']
        lbl_up = lbl.upper()

        if lbl_up == 'EQUITY' and row['is_section']:
            in_equity = True
            equity_items = []

        if in_equity:
            if lbl.startswith('Net Income') or lbl.startswith('NET INCOME'):
                for mo, v in ytd_ni.items():
                    row['months'][mo] = v
                row['is_total'] = False
                equity_items.append(row)
            elif lbl_up.startswith('TOTAL EQUITY'):
                for mo in ytd_ni:
                    row['months'][mo] = round(
                        sum(r['months'].get(mo, 0) for r in equity_items), 2)
                in_equity = False
                totals['total_equity'] = row
            elif not row['is_section'] and not row['is_total']:
                equity_items.append(row)

        if lbl_up == 'TOTAL LIABILITIES':
            totals['total_liabilities'] = row
        elif 'TOTAL LIABILITIES &' in lbl_up:
            totals['total_liabilities_equity'] = row
        elif lbl_up == 'TOTAL ASSETS':
            totals['total_assets'] = row
        elif 'BALANCE CHECK' in lbl_up:
            totals['balance_check'] = row

    if all(k in totals for k in ('total_liabilities_equity', 'total_liabilities', 'total_equity')):
        tl = totals['total_liabilities']
        te = totals['total_equity']
        for mo in ytd_ni:
            totals['total_liabilities_equity']['months'][mo] = round(
                tl['months'].get(mo, 0) + te['months'].get(mo, 0), 2)

    if all(k in totals for k in ('balance_check', 'total_assets', 'total_liabilities_equity')):
        ta = totals['total_assets']
        tl_eq = totals['total_liabilities_equity']
        for mo in ytd_ni:
            totals['balance_check']['months'][mo] = round(
                ta['months'].get(mo, 0) - tl_eq['months'].get(mo, 0), 2)


def _add_ytd(rows, available_months, cumulative=True):
    """Add a ytd dict to each row.
    For P&L (cumulative=True): ytd[month] = sum from first month through that month.
    For BS (cumulative=False): ytd[month] = same as months[month] (balance is already point-in-time).
    """
    for row in rows:
        ytd = {}
        running = 0.0
        for mi in available_months:
            mo = MONTHS[mi - 1]
            v  = row['months'].get(mo, 0.0) or 0.0
            if cumulative:
                running = round(running + v, 2)
                ytd[mo] = running
            else:
                ytd[mo] = v
        row['ytd'] = ytd


@app.route('/data/financials', methods=['GET'])
@require_jwt
def financials():
    available = _available_months()
    if not available:
        return jsonify({'months': [], 'pl': [], 'bs': []})
    try:
        wb = _workbook()
        txn_data = _build_txn_data(wb)
        pl_rows = _parse_financials_sheet(wb['Income Statement'], available, txn_data, bs_mode=False)
        bs_rows = _parse_financials_sheet(wb['Balance Sheet'],    available, txn_data, bs_mode=True)
        _patch_pl_totals(pl_rows, txn_data, available)
        _patch_bs_equity(bs_rows, txn_data, available)
        _add_ytd(pl_rows, available, cumulative=True)
        _add_ytd(bs_rows, available, cumulative=False)
        month_labels = [MONTHS[m-1] for m in available]
        return jsonify({'months': month_labels, 'pl': pl_rows, 'bs': bs_rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 503


@app.route('/data/transactions', methods=['GET'])
@require_jwt
def get_transactions():
    """Return individual transaction rows for a GL account + month.
    Query params: account=<int>, month=<1-12>
    """
    try:
        account   = int(request.args.get('account', 0))
        month_num = int(request.args.get('month', 0))
    except (TypeError, ValueError):
        return jsonify({'error': 'account and month must be integers'}), 400
    if not account or not month_num:
        return jsonify({'error': 'account and month are required'}), 400
    try:
        wb = _workbook()
    except Exception as exc:
        return jsonify({'error': str(exc)}), 503
    if 'Transactions' not in wb.sheetnames:
        return jsonify({'account': account, 'month': MONTHS[month_num - 1], 'transactions': []})
    ws = wb['Transactions']
    results = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        date_val = row[0]
        je_num   = str(row[1] or '').strip()
        desc     = str(row[2] or '').strip()
        acct     = row[3]
        notes    = str(row[7] or '').strip() if len(row) > 7 else ''
        debit    = row[5]
        credit   = row[6]
        try:
            acct_int = int(float(acct))
        except (TypeError, ValueError):
            continue
        if acct_int != account:
            continue
        if isinstance(date_val, (datetime, date)):
            row_month = date_val.month
            date_str  = date_val.strftime('%b %d, %Y')
        elif isinstance(date_val, (int, float)):
            try:
                d = _EXCEL_EPOCH + timedelta(days=int(date_val))
                row_month = d.month
                date_str  = d.strftime('%b %d, %Y')
            except Exception:
                continue
        elif isinstance(date_val, str):
            d = _to_date(date_val)
            if d is None:
                continue
            row_month = d.month
            date_str  = d.strftime('%b %d, %Y')
        else:
            continue
        if row_month != month_num:
            continue
        results.append({
            'date':   date_str,
            'je':     je_num,
            'desc':   desc,
            'debit':  round(float(debit), 2)  if debit  else None,
            'credit': round(float(credit), 2) if credit else None,
            'notes':  notes,
        })
    return jsonify({'account': account, 'month': MONTHS[month_num - 1], 'transactions': results})


# ---------------------------------------------------------------------------
# PBC (Prepared by Client) Request System
# ---------------------------------------------------------------------------

import csv as _csv
import io as _io
import uuid as _uuid
import subprocess as _subprocess

_PBC_DATA_PATH  = '/home/jlobel/lac_automation/pbc_data.json'
_PBC_DRIVE_ROOT = 'Clients'

# ---------------------------------------------------------------------------
# Client document store
# ---------------------------------------------------------------------------

_CLIENT_DOCS_PATH    = '/home/jlobel/lac_automation/client_docs.json'
_CLIENT_DOCS_PDF_DIR = '/home/jlobel/lac_automation/client_docs_pdf'


def _load_client_docs() -> dict:
    try:
        with open(_CLIENT_DOCS_PATH) as f:
            return _json.load(f)
    except (FileNotFoundError, _json.JSONDecodeError):
        return {}


def _save_client_docs(data: dict):
    with open(_CLIENT_DOCS_PATH, 'w') as f:
        _json.dump(data, f, indent=2)


def _store_client_doc(client_name: str, item_id: str, doc_type: str,
                      label: str, pdf_path: str):
    """Persist a copy of a document PDF for the client's document history."""
    import shutil as _shutil
    client_pdf_dir = f'{_CLIENT_DOCS_PDF_DIR}/{client_name}'
    os.makedirs(client_pdf_dir, exist_ok=True)

    safe_label = label.replace(' ', '_')
    dest = f'{client_pdf_dir}/{item_id}_{safe_label}.pdf'

    try:
        if pdf_path != dest and os.path.exists(pdf_path):
            _shutil.copy2(pdf_path, dest)
    except Exception as exc:
        app.logger.warning('_store_client_doc copy failed: %s', exc)
        return

    docs = _load_client_docs()
    client_list = docs.setdefault(client_name, [])

    # Deduplicate by item_id
    if any(d.get('item_id') == item_id for d in client_list):
        return

    import uuid as _uuid2
    client_list.append({
        'id':        str(_uuid2.uuid4())[:8],
        'item_id':   item_id,
        'doc_type':  doc_type,
        'label':     label,
        'date':      date.today().isoformat(),
        'pdf_path':  dest,
        'filename':  os.path.basename(dest),
    })
    _save_client_docs(docs)


# ---------------------------------------------------------------------------

SECTION_CODES = {
    'Cash': 'C',
    'Cash & Bank': 'CB',
    'Accounts Receivable': 'AR',
    'Fixed Assets': 'FA',
    'Inventory': 'INV',
    'Prepaid': 'PP',
    'Prepaid & Other Current Assets': 'PP',
    'Other Assets': 'OA',
    'Accounts Payable': 'AP',
    'Accrued Liabilities': 'AL',
    'Other Liabilities': 'OL',
    'Debt': 'D',
    'Debt & Financing': 'DF',
    'Equity': 'EQ',
    'Revenue': 'REV',
    'Expenses': 'EXP',
    'Payroll': 'PAY',
    'Payroll & Benefits': 'PAY',
    'Income Tax': 'TAX',
    'Financial Statements': 'FS',
    'Corporate & Legal': 'CL',
    'Related Parties': 'RP',
    'Planning': 'PL',
    'Internal Controls': 'IC',
    'General': 'G',
    'Income & Revenue': 'REV',
    'Tax - Business': 'TAX',
    'Estimated Taxes & Payments': 'TAX',
}

_PBC_TEMPLATES_PATH = '/home/jlobel/lac_automation/pbc_templates.json'

def _load_pbc_templates():
    if os.path.exists(_PBC_TEMPLATES_PATH):
        with open(_PBC_TEMPLATES_PATH) as f:
            return _json.load(f)
    return {}

def _pbc_data_path():
    if has_request_context() and request.headers.get('X-Env') == 'test':
        try:
            if request.client_name == '__staff__':
                return '/home/jlobel/lac_automation/pbc_data_test.json'
        except Exception:
            pass
    return _PBC_DATA_PATH

def _load_pbc():
    path = _pbc_data_path()
    if os.path.exists(path):
        with open(path) as f:
            return _json.load(f)
    return {}

def _save_pbc(data):
    path = _pbc_data_path()
    with open(path, 'w') as f:
        _json.dump(data, f, indent=2)

def _client_folder_name(client_name: str, client_number: str) -> str:
    num = str(client_number).replace(',', '').strip() if client_number else ''
    if num:
        return f"{num} - {client_name}"
    return client_name

def _rclone(*args):
    r = _subprocess.run(['rclone', *args], capture_output=True, text=True, timeout=60)
    return r.returncode == 0

def _ensure_drive_folder(path: str):
    """Create folder in Drive via rclone if it doesn't exist."""
    _rclone('mkdir', f'gdrive_user:{path}')

def _pbc_drive_path(client_name: str, client_number: str, section: str = None) -> str:
    folder = _client_folder_name(client_name, client_number)
    base = f"{_PBC_DRIVE_ROOT}/{folder}"
    if section:
        return f"{base}/{section}"
    return base

def _get_client_number(client_name: str) -> str:
    """Look up client number from Engagement Pipeline col B."""
    try:
        wb = _workbook()
        if 'Engagement Pipeline' not in wb.sheetnames:
            return ''
        for row in wb['Engagement Pipeline'].iter_rows(min_row=3, values_only=True):
            if row and row[0] and str(row[0]).strip().lower() == client_name.lower():
                num = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                return num.replace(',', '').strip()
    except Exception:
        pass
    return ''

def _next_request_number(requests: list, section_code: str) -> str:
    existing = [r['request_number'] for r in requests
                if r['request_number'].startswith(f"{section_code}-")]
    if not existing:
        return f"{section_code}-01"
    nums = []
    for n in existing:
        try:
            nums.append(int(n.split('-')[1]))
        except (IndexError, ValueError):
            pass
    return f"{section_code}-{(max(nums) + 1):02d}" if nums else f"{section_code}-01"


# ---------------------------------------------------------------------------
# Activity log
# ---------------------------------------------------------------------------

_ACTIVITY_LOG_PATH = '/home/jlobel/lac_automation/activity_log.json'


def _log_activity(client: str, actor_type: str, action: str, details: str = ''):
    """Append an event to the activity log. actor_type: 'admin' | 'client'"""
    try:
        log = []
        if os.path.exists(_ACTIVITY_LOG_PATH):
            with open(_ACTIVITY_LOG_PATH) as _f:
                log = _json.load(_f)
        log.append({
            'id':         str(_uuid.uuid4()),
            'timestamp':  datetime.utcnow().isoformat(),
            'client':     client,
            'actor_type': actor_type,
            'action':     action,
            'details':    details,
        })
        if len(log) > 2000:
            log = log[-2000:]
        with open(_ACTIVITY_LOG_PATH, 'w') as _f:
            _json.dump(log, _f)
    except Exception as _e:
        app.logger.warning('activity log error: %s', _e)


def _get_client_email(client_name: str) -> str | None:
    """Return email address for a client from AR Aging tab."""
    try:
        wb = _workbook()
        if 'AR Aging' not in wb.sheetnames:
            return None
        cl = client_name.lower()
        for row in wb['AR Aging'].iter_rows(min_row=AR_DATA_START, values_only=True):
            if row and row[AR_C_CLIENT] and str(row[AR_C_CLIENT]).strip().lower() == cl:
                email = row[AR_C_EMAIL] if len(row) > AR_C_EMAIL else None
                if email:
                    return str(email).strip()
    except Exception:
        pass
    return None


@app.route('/activity/log', methods=['GET'])
@require_jwt
def activity_log_endpoint():
    """Return recent activity events. Admin-only."""
    if request.client_name != '__staff__':
        return jsonify({'error': 'Staff only'}), 403
    log = []
    if os.path.exists(_ACTIVITY_LOG_PATH):
        with open(_ACTIVITY_LOG_PATH) as _f:
            log = _json.load(_f)
    client_filter = request.args.get('client', '').strip()
    limit = min(int(request.args.get('limit', 300)), 1000)
    if client_filter:
        log = [e for e in log if e.get('client') == client_filter]
    # Unique client list for filter dropdown
    clients = sorted({e['client'] for e in log if e.get('client')})
    return jsonify({'events': list(reversed(log[-limit:])), 'clients': clients})


@app.route('/pbc/import', methods=['POST'])
@require_jwt
def pbc_import():
    """Admin uploads CSV of PBC requests for a client.
    CSV columns: client, section, description, due_date (optional), notes (optional)
    Merges with existing requests — won't duplicate.
    """
    client_name = (request.form.get('client') or '').strip()
    if not client_name:
        return jsonify({'error': 'client required'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'file required'}), 400

    f = request.files['file']
    text = f.read().decode('utf-8-sig')
    reader = _csv.DictReader(_io.StringIO(text))

    data = _load_pbc()
    if client_name not in data:
        cn = _get_client_number(client_name)
        data[client_name] = {
            'client_number': cn,
            'requests': [],
            'archived': False,
            'archive_date': None,
        }
        # Create Drive folder
        path = _pbc_drive_path(client_name, cn)
        _ensure_drive_folder(path)

    client_data = data[client_name]
    requests_list = client_data['requests']
    added = 0

    for row in reader:
        section = (row.get('section') or row.get('Section') or '').strip()
        desc    = (row.get('description') or row.get('Description') or '').strip()
        if not section or not desc:
            continue
        # Deduplicate by section + description
        if any(r['section'] == section and r['description'] == desc
               for r in requests_list):
            continue
        code = SECTION_CODES.get(section, section[:2].upper())
        req_num = _next_request_number(requests_list, code)
        due = (row.get('due_date') or row.get('Due Date') or '').strip()
        note = (row.get('notes') or row.get('Notes') or '').strip()
        requests_list.append({
            'id':             str(_uuid.uuid4()),
            'request_number': req_num,
            'section':        section,
            'section_code':   code,
            'description':    desc,
            'upload_date':    date.today().isoformat(),
            'due_date':       due or None,
            'status':         'Not Provided',
            'notes':          [{'author': 'admin', 'text': note,
                                'date': date.today().isoformat()}] if note else [],
            'file_name':      None,
            'drive_path':     None,
            'provided_date':  None,
            'completed_date': None,
        })
        added += 1

    _save_pbc(data)
    if added:
        _log_activity(client_name, 'admin', 'PBC Import',
                      f'{added} request{"s" if added != 1 else ""} added')
    return jsonify({'success': True, 'added': added,
                    'total': len(requests_list)})


@app.route('/pbc/add', methods=['POST'])
@require_jwt
def pbc_add():
    """Admin adds a single PBC request (or creates a new client workspace)."""
    body        = request.get_json(silent=True) or {}
    client_name = body.get('client', '').strip()
    section     = body.get('section', '').strip()
    description = body.get('description', '').strip()
    due_date    = body.get('due_date', '').strip() or None
    note_text   = body.get('note', '').strip()
    engagement  = body.get('engagement', '').strip()

    if not client_name:
        return jsonify({'error': 'client required'}), 400
    if request.client_name != '__staff__':
        return jsonify({'error': 'Staff only'}), 403

    data = _load_pbc()
    if client_name not in data:
        cn = _get_client_number(client_name)
        data[client_name] = {
            'client_number': cn,
            'engagement': engagement,
            'requests': [],
            'archived': False,
            'archive_date': None,
        }
        path = _pbc_drive_path(client_name, cn)
        _ensure_drive_folder(path)
    elif engagement:
        data[client_name]['engagement'] = engagement

    if section and description:
        requests_list = data[client_name]['requests']
        code    = SECTION_CODES.get(section, section[:2].upper())
        req_num = _next_request_number(requests_list, code)
        requests_list.append({
            'id':             str(_uuid.uuid4()),
            'request_number': req_num,
            'section':        section,
            'section_code':   code,
            'description':    description,
            'upload_date':    date.today().isoformat(),
            'due_date':       due_date,
            'status':         'Not Provided',
            'notes':          [{'author': 'admin', 'text': note_text,
                                'date': date.today().isoformat()}] if note_text else [],
            'file_name':      None,
            'drive_path':     None,
            'provided_date':  None,
            'completed_date': None,
        })
        _log_activity(client_name, 'admin', 'Request Added', f'{req_num}: {description[:60]}')

    _save_pbc(data)
    return jsonify({'success': True, 'total': len(data[client_name]['requests'])})


@app.route('/pbc/delete', methods=['POST'])
@require_jwt
def pbc_delete():
    """Admin deletes a single PBC request."""
    body        = request.get_json(silent=True) or {}
    client_name = body.get('client', '').strip()
    req_id      = body.get('id', '').strip()

    if request.client_name != '__staff__':
        return jsonify({'error': 'Staff only'}), 403

    data = _load_pbc()
    client_data = data.get(client_name)
    if not client_data:
        return jsonify({'error': 'Client not found'}), 404

    before = len(client_data['requests'])
    client_data['requests'] = [r for r in client_data['requests'] if r['id'] != req_id]
    if len(client_data['requests']) == before:
        return jsonify({'error': 'Request not found'}), 404

    _save_pbc(data)
    return jsonify({'success': True})


@app.route('/pbc/templates', methods=['GET'])
@require_jwt
def pbc_templates():
    """Return list of available template names."""
    return jsonify({'templates': list(_load_pbc_templates().keys())})


@app.route('/pbc/apply-template', methods=['POST'])
@require_jwt
def pbc_apply_template():
    """Apply a standard template to a client, skipping any already-existing requests."""
    if request.client_name != '__staff__':
        return jsonify({'error': 'Staff only'}), 403

    body          = request.get_json(silent=True) or {}
    client_name   = body.get('client', '').strip()
    template_name = body.get('template', '').strip()
    due_offset    = int(body.get('due_days', 30))  # days from today for initial due dates

    if not client_name or not template_name:
        return jsonify({'error': 'client and template required'}), 400

    templates = _load_pbc_templates()
    if template_name not in templates:
        return jsonify({'error': f'Template "{template_name}" not found'}), 404

    data = _load_pbc()
    if client_name not in data:
        cn = _get_client_number(client_name)
        data[client_name] = {
            'client_number': cn,
            'engagement': '',
            'requests': [],
            'archived': False,
            'archive_date': None,
        }
        _ensure_drive_folder(_pbc_drive_path(client_name, cn))

    requests_list = data[client_name]['requests']
    existing_descs = {(r['section'], r['description']) for r in requests_list}

    added = 0
    base_due = (date.today() + __import__('datetime').timedelta(days=due_offset)).isoformat()

    for section_block in templates[template_name]:
        section = section_block['section']
        code    = SECTION_CODES.get(section, section[:2].upper())
        for desc in section_block['items']:
            if (section, desc) in existing_descs:
                continue
            req_num = _next_request_number(requests_list, code)
            requests_list.append({
                'id':             str(_uuid.uuid4()),
                'request_number': req_num,
                'section':        section,
                'section_code':   code,
                'description':    desc,
                'upload_date':    date.today().isoformat(),
                'due_date':       base_due,
                'status':         'Not Provided',
                'notes':          [],
                'file_name':      None,
                'drive_path':     None,
                'provided_date':  None,
                'completed_date': None,
            })
            existing_descs.add((section, desc))
            added += 1

    _save_pbc(data)
    _log_activity(client_name, 'admin', 'Template Applied',
                  f'{template_name} — {added} requests added')
    return jsonify({'success': True, 'added': added, 'total': len(requests_list)})


@app.route('/pbc/list', methods=['GET'])
@require_jwt
def pbc_list():
    """Return PBC requests. Staff sees all clients or filters by client.
    Portal client sees only their own."""
    data = _load_pbc()
    is_staff = (request.client_name == '__staff__')

    if is_staff:
        client_filter = request.args.get('client', '').strip()
        if client_filter:
            clients = {client_filter: data.get(client_filter, {})}
        else:
            clients = data
    else:
        client_name = request.client_name
        clients = {client_name: data.get(client_name, {})} if client_name else {}

    result = {}
    for cname, cdata in clients.items():
        if not cdata:
            continue
        reqs = cdata.get('requests', [])
        # Compute section progress
        sections = {}
        for r in reqs:
            s = r['section']
            if s not in sections:
                sections[s] = {'total': 0, 'provided': 0, 'complete': 0}
            sections[s]['total'] += 1
            if r['status'] == 'Provided':
                sections[s]['provided'] += 1
            elif r['status'] == 'Complete':
                sections[s]['complete'] += 1
        result[cname] = {
            'client_number': cdata.get('client_number', ''),
            'engagement': cdata.get('engagement', ''),
            'archived': cdata.get('archived', False),
            'requests': reqs,
            'sections': sections,
        }
    return jsonify(result)


@app.route('/pbc/status', methods=['POST'])
@require_jwt
def pbc_status():
    """Update request status. Staff can set any status; client can only set Provided."""
    body = request.get_json(silent=True) or {}
    client_name = body.get('client', '').strip()
    req_id      = body.get('id', '').strip()
    new_status  = body.get('status', '').strip()

    valid = {'Not Provided', 'Provided', 'Complete'}
    if new_status not in valid:
        return jsonify({'error': f'status must be one of {valid}'}), 400

    is_staff = (request.client_name == '__staff__')
    if not is_staff and new_status != 'Provided':
        return jsonify({'error': 'Clients may only set status to Provided'}), 403

    data = _load_pbc()
    client_data = data.get(client_name)
    if not client_data:
        return jsonify({'error': 'Client not found'}), 404

    for r in client_data['requests']:
        if r['id'] == req_id:
            r['status'] = new_status
            if new_status == 'Provided':
                r['provided_date'] = date.today().isoformat()
            elif new_status == 'Complete':
                r['completed_date'] = date.today().isoformat()
            elif new_status == 'Not Provided':
                r['provided_date'] = None
                r['completed_date'] = None
            _save_pbc(data)
            actor = 'admin' if is_staff else 'client'
            _log_activity(client_name, actor, 'Status Update',
                          f'{r["request_number"]}: → {new_status}')
            if new_status == 'Provided':
                _notify_admin_pbc_provided(client_name, r)
            return jsonify({'success': True})

    return jsonify({'error': 'Request not found'}), 404


@app.route('/pbc/update', methods=['POST'])
@require_jwt
def pbc_update():
    """Admin updates any field on a request (description, due_date, section)."""
    body = request.get_json(silent=True) or {}
    client_name = body.get('client', '').strip()
    req_id      = body.get('id', '').strip()

    data = _load_pbc()
    client_data = data.get(client_name)
    if not client_data:
        return jsonify({'error': 'Client not found'}), 404

    for r in client_data['requests']:
        if r['id'] == req_id:
            for field in ('description', 'due_date', 'section'):
                if field in body:
                    r[field] = body[field]
            _save_pbc(data)
            return jsonify({'success': True})
    return jsonify({'error': 'Request not found'}), 404


@app.route('/pbc/note', methods=['POST'])
@require_jwt
def pbc_note():
    """Add a note to a request."""
    body = request.get_json(silent=True) or {}
    client_name = body.get('client', '').strip()
    req_id      = body.get('id', '').strip()
    text        = (body.get('text') or '').strip()
    if not text:
        return jsonify({'error': 'text required'}), 400

    is_staff = (request.client_name == '__staff__')
    author   = 'admin' if is_staff else 'client'

    data = _load_pbc()
    client_data = data.get(client_name)
    if not client_data:
        return jsonify({'error': 'Client not found'}), 404

    for r in client_data['requests']:
        if r['id'] == req_id:
            r['notes'].append({
                'author': author,
                'text':   text,
                'date':   date.today().isoformat(),
            })
            _save_pbc(data)
            _log_activity(client_name, author, 'Note Added',
                          f'{r["request_number"]}: {text[:80]}{"…" if len(text) > 80 else ""}')
            return jsonify({'success': True})
    return jsonify({'error': 'Request not found'}), 404


@app.route('/pbc/upload', methods=['POST'])
@require_jwt
def pbc_upload():
    """Client uploads a file for a specific request. Saves to Drive."""
    from werkzeug.utils import secure_filename as _secure
    client_name = (request.form.get('client') or '').strip()
    req_id      = (request.form.get('id') or '').strip()
    if not client_name or not req_id:
        return jsonify({'error': 'client and id required'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'file required'}), 400

    data = _load_pbc()
    client_data = data.get(client_name)
    if not client_data:
        return jsonify({'error': 'Client not found'}), 404

    for r in client_data['requests']:
        if r['id'] == req_id:
            f = request.files['file']
            filename = _secure(f.filename)
            section  = r['section']
            cn       = client_data.get('client_number', '')

            # Save locally then upload to Drive
            tmp_path = f'/tmp/pbc_{req_id}_{filename}'
            f.save(tmp_path)

            drive_section_path = _pbc_drive_path(client_name, cn, section)
            _ensure_drive_folder(drive_section_path)
            _rclone('copyto', tmp_path, f'gdrive_user:{drive_section_path}/{filename}')
            os.unlink(tmp_path)

            r['file_name']  = filename
            r['drive_path'] = f'{drive_section_path}/{filename}'
            r['status']     = 'Provided'
            r['provided_date'] = date.today().isoformat()
            _save_pbc(data)
            _notify_admin_pbc_provided(client_name, r)
            _log_activity(client_name, 'client', 'File Uploaded',
                          f'{r["request_number"]}: {filename}')
            return jsonify({'success': True, 'file_name': filename})

    return jsonify({'error': 'Request not found'}), 404


@app.route('/pbc/export/csv', methods=['GET'])
def pbc_export_csv():
    # Allow JWT via query param for direct window.open() downloads
    token = request.args.get('token') or (request.headers.get('Authorization','')[7:])
    try:
        jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except Exception:
        return jsonify({'error': 'Unauthorized'}), 401

    """Export PBC requests as CSV. ?client=X for single client, else all."""
    from flask import Response as _Response
    import openpyxl as _openpyxl

    data = _load_pbc()
    client_filter = request.args.get('client', '').strip()

    if client_filter:
        clients = {client_filter: data.get(client_filter, {})}
        fname = f"PBC_{client_filter.replace(' ','_')}.csv"
    else:
        clients = data
        fname = 'PBC_All_Clients.csv'

    output = _io.StringIO()
    writer = _csv.writer(output)
    writer.writerow(['Client', 'Client Number', 'Request #', 'Section',
                     'Description', 'Upload Date', 'Due Date', 'Status',
                     'Provided Date', 'Completed Date', 'File'])

    for cname, cdata in sorted(clients.items()):
        for r in cdata.get('requests', []):
            writer.writerow([
                cname,
                cdata.get('client_number', ''),
                r.get('request_number', ''),
                r.get('section', ''),
                r.get('description', ''),
                r.get('upload_date', ''),
                r.get('due_date', ''),
                r.get('status', ''),
                r.get('provided_date', ''),
                r.get('completed_date', ''),
                r.get('file_name', ''),
            ])

    return _Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.route('/pbc/complete', methods=['POST'])
@require_jwt
def pbc_complete():
    """Mark engagement as complete and start 6-month archive timer."""
    body = request.get_json(silent=True) or {}
    client_name = body.get('client', '').strip()
    data = _load_pbc()
    if client_name not in data:
        return jsonify({'error': 'Client not found'}), 404
    data[client_name]['archived']     = True
    data[client_name]['archive_date'] = date.today().isoformat()
    _save_pbc(data)
    _log_activity(client_name, 'admin', 'Engagement Archived', '')
    return jsonify({'success': True,
                    'message': f'{client_name} marked complete. Drive files will be deleted after 6 months.'})


@app.route('/pbc/clients', methods=['GET'])
@require_jwt
def pbc_clients():
    """Return list of clients that have PBC data."""
    data = _load_pbc()
    result = [
        {'name': name, 'client_number': cd.get('client_number', ''),
         'archived': cd.get('archived', False),
         'total': len(cd.get('requests', [])),
         'complete': sum(1 for r in cd.get('requests', []) if r['status'] == 'Complete')}
        for name, cd in sorted(data.items())
    ]
    return jsonify({'clients': result})


def _notify_admin_pbc_provided(client_name: str, req: dict):
    """Email admin when a client provides a document."""
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'PBC Provided — {client_name}: {req["request_number"]}'
        msg['From']    = ZOHO_EMAIL
        msg['To']      = ZOHO_EMAIL
        body = f"""
        <p><strong>{client_name}</strong> has provided support for:</p>
        <p><strong>{req['request_number']} — {req['description']}</strong></p>
        <p>File: {req.get('file_name') or '(status only)'}</p>
        <p>Log in to the dashboard to review and mark Complete.</p>
        """
        msg.attach(MIMEText(body, 'html'))
        with smtplib.SMTP_SSL('smtp.zoho.com', 465) as s:
            s.login(ZOHO_EMAIL, ZOHO_PASSWORD)
            s.send_message(msg)
    except Exception as e:
        app.logger.warning('PBC notify failed: %s', e)


# ---------------------------------------------------------------------------
# Plaid integration
# ---------------------------------------------------------------------------

import plaid
from plaid.api import plaid_api
from plaid.model.link_token_create_request import LinkTokenCreateRequest
from plaid.model.link_token_create_request_user import LinkTokenCreateRequestUser
from plaid.model.products import Products
from plaid.model.country_code import CountryCode
from plaid.model.item_public_token_exchange_request import ItemPublicTokenExchangeRequest
from plaid.model.transactions_get_request import TransactionsGetRequest
from plaid.model.transactions_get_request_options import TransactionsGetRequestOptions
from plaid.model.accounts_balance_get_request import AccountsBalanceGetRequest

_PLAID_CLIENT_ID  = os.getenv('PLAID_CLIENT_ID', '')
_PLAID_SECRET     = os.getenv('PLAID_SECRET', '')
_PLAID_ENV        = os.getenv('PLAID_ENV', 'sandbox')
_PLAID_TOKEN_PATH = '/home/jlobel/lac_automation/plaid_access_token.json'

_PLAID_HOST_MAP = {
    'sandbox':     plaid.Environment.Sandbox,
    'development': 'https://development.plaid.com',
    'production':  plaid.Environment.Production,
}


def _plaid_client():
    cfg = plaid.Configuration(
        host=_PLAID_HOST_MAP.get(_PLAID_ENV, plaid.Environment.Sandbox),
        api_key={'clientId': _PLAID_CLIENT_ID, 'secret': _PLAID_SECRET},
    )
    return plaid_api.PlaidApi(plaid.ApiClient(cfg))


def _load_plaid_token() -> str:
    if os.path.exists(_PLAID_TOKEN_PATH):
        with open(_PLAID_TOKEN_PATH) as f:
            return _json.load(f).get('access_token', '')
    return os.getenv('PLAID_ACCESS_TOKEN', '')


def _save_plaid_token(access_token: str):
    with open(_PLAID_TOKEN_PATH, 'w') as f:
        _json.dump({'access_token': access_token}, f)


@app.route('/plaid/link-token', methods=['POST'])
@require_jwt
def plaid_link_token():
    """Create a Plaid Link token to initiate the account linking flow."""
    try:
        client = _plaid_client()
        req = LinkTokenCreateRequest(
            user=LinkTokenCreateRequestUser(client_user_id='lac-admin'),
            client_name='Lobel Accountancy',
            products=[Products('transactions')],
            country_codes=[CountryCode('US')],
            language='en',
            redirect_uri='https://auth.lobelaccountancy.com/plaid/oauth-return',
        )
        resp = client.link_token_create(req)
        return jsonify({'link_token': resp['link_token']})
    except Exception as e:
        app.logger.error('plaid link_token error: %s', e)
        return jsonify({'error': str(e)}), 500


@app.route('/plaid/link', methods=['GET'])
@app.route('/plaid/oauth-return', methods=['GET'])
def plaid_oauth_return():
    """Plaid Link page and OAuth return landing."""
    return send_from_directory(
        os.path.join(os.path.dirname(__file__), '../docs'),
        'plaid_link.html',
    )


@app.route('/plaid/exchange', methods=['POST'])
@require_jwt
def plaid_exchange():
    """Exchange Plaid public token for a permanent access token."""
    body = request.get_json(silent=True) or {}
    public_token = body.get('public_token')
    if not public_token:
        return jsonify({'error': 'public_token required'}), 400
    try:
        client = _plaid_client()
        resp = client.item_public_token_exchange(
            ItemPublicTokenExchangeRequest(public_token=public_token)
        )
        _save_plaid_token(resp['access_token'])
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error('plaid exchange error: %s', e)
        return jsonify({'error': str(e)}), 500


@app.route('/plaid/transactions', methods=['GET'])
@require_jwt
def plaid_transactions():
    """Return recent transactions for the linked account."""
    access_token = _load_plaid_token()
    if not access_token:
        return jsonify({'error': 'No account linked — visit /plaid/link-token first'}), 400
    days = int(request.args.get('days', 30))
    end   = date.today()
    start = end - timedelta(days=days)
    try:
        client = _plaid_client()
        resp = client.transactions_get(TransactionsGetRequest(
            access_token=access_token,
            start_date=start,
            end_date=end,
            options=TransactionsGetRequestOptions(count=500),
        ))
        txns = [
            {
                'date':        t['date'].isoformat() if hasattr(t['date'], 'isoformat') else str(t['date']),
                'name':        t['name'],
                'amount':      t['amount'],
                'category':    (t.get('category') or [''])[0],
                'account_id':  t['account_id'],
            }
            for t in resp['transactions']
        ]
        accounts = [
            {
                'id':       a['account_id'],
                'name':     a['name'],
                'balance':  a['balances']['current'],
                'type':     str(a['type']),
            }
            for a in resp['accounts']
        ]
        return jsonify({'transactions': txns, 'accounts': accounts,
                        'total': resp['total_transactions']})
    except Exception as e:
        app.logger.error('plaid transactions error: %s', e)
        return jsonify({'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Audit Tools — server-side document analysis via Ollama or Claude
# ---------------------------------------------------------------------------

OLLAMA_URL = 'http://localhost:11434/api/generate'
OLLAMA_MODEL = 'llama3.2:3b'

ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY', '')
CLAUDE_MODEL = 'claude-sonnet-4-6'

_AUDIT_SYSTEM = {
    'ask': (
        'You are an audit assistant. Answer the question based only on the provided documents. '
        'Be precise and cite the document name when referencing specific data.'
    ),
    'crossref': (
        'You are an audit assistant performing cross-referencing. Find every occurrence of the '
        'requested value, amount, or term across the documents. For each match, state the document '
        'name, page or section, and the surrounding context. Note any discrepancies if the same '
        'item appears with different values in different documents.'
    ),
    'footing': (
        'You are an audit assistant performing a footing and cross-footing check. '
        'Examine every numerical table in the documents. Verify that each column total and row '
        'total is arithmetically correct. Report the table name/location, the expected total, '
        'the stated total, and whether it agrees or disagrees. List all discrepancies clearly.'
    ),
    'extract': (
        'You are an audit assistant. Extract all financial figures, dates, party names, and '
        'key data points from the documents into a structured list. Group by document. '
        'Include amounts with their currency, dates in MM/DD/YYYY format, and labels.'
    ),
}


def _extract_file_text(file_obj):
    """Extract readable text from an uploaded file (PDF, xlsx, txt)."""
    name = (file_obj.filename or '').lower()
    data = file_obj.read()

    if name.endswith('.pdf'):
        pages = []
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ''
                tables = page.extract_tables() or []
                for tbl in tables:
                    text += '\n[TABLE]\n'
                    for row in tbl:
                        text += ' | '.join(str(c or '').strip() for c in row) + '\n'
                pages.append(f'[Page {i}]\n{text}')
        return '\n\n'.join(pages)

    if name.endswith(('.xlsx', '.xls', '.xlsm')):
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(io.BytesIO(data), data_only=True)
            parts = []
            for sh in wb.sheetnames:
                ws = wb[sh]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        rows.append(' | '.join(str(c or '').strip() for c in row))
                if rows:
                    parts.append(f'[Sheet: {sh}]\n' + '\n'.join(rows))
            return '\n\n'.join(parts)
        except Exception as exc:
            return f'[Could not read Excel file: {exc}]'

    # Plain text / CSV / anything else
    return data.decode('utf-8', errors='replace')


def _search_text(text, term):
    """Return list of (line_num, snippet) for lines containing term (case-insensitive)."""
    hits = []
    term_lower = term.lower()
    for i, line in enumerate(text.splitlines(), 1):
        if term_lower in line.lower():
            hits.append({'line': i, 'snippet': line.strip()})
    return hits


def _query_ollama(system_prompt, context, user_prompt):
    """Send to local Ollama, return response string."""
    full_prompt = f"DOCUMENTS:\n{context}\n\nTASK: {user_prompt}"
    payload = json.dumps({
        'model': OLLAMA_MODEL,
        'system': system_prompt,
        'prompt': full_prompt,
        'stream': False,
    }).encode()
    req = urllib.request.Request(
        OLLAMA_URL,
        data=payload,
        headers={'Content-Type': 'application/json'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            result = json.loads(resp.read())
            return result.get('response', '')
    except Exception as exc:
        return f'[Ollama error: {exc}]'


def _query_claude(system_prompt, context, user_prompt):
    """Send to Anthropic Claude API, return response string."""
    if not ANTHROPIC_API_KEY:
        return '[Claude error: ANTHROPIC_API_KEY not set]'
    full_prompt = f"DOCUMENTS:\n{context}\n\nTASK: {user_prompt}"
    try:
        resp = requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'x-api-key': ANTHROPIC_API_KEY,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json',
            },
            json={
                'model': CLAUDE_MODEL,
                'max_tokens': 4096,
                'system': system_prompt,
                'messages': [{'role': 'user', 'content': full_prompt}],
            },
            timeout=120,
        )
        data = resp.json()
        if resp.status_code != 200:
            return f'[Claude error: {data.get("error", {}).get("message", resp.status_code)}]'
        return data['content'][0]['text']
    except Exception as exc:
        return f'[Claude error: {exc}]'


@app.route('/audit-tools/extract', methods=['POST'])
@require_jwt
def audit_extract():
    """Extract text/tables from uploaded files without AI processing."""
    files = request.files.getlist('files')
    if not files or not files[0].filename:
        return jsonify({'error': 'No files uploaded'}), 400
    docs = []
    for f in files:
        text = _extract_file_text(f)
        docs.append({'name': f.filename, 'text': text, 'chars': len(text)})
    return jsonify({'docs': docs})


@app.route('/audit-tools/search', methods=['POST'])
@require_jwt
def audit_search():
    """Search for a value across multiple uploaded documents (no AI)."""
    files = request.files.getlist('files')
    term = (request.form.get('term') or '').strip()
    if not term:
        return jsonify({'error': 'Search term required'}), 400
    results = []
    for f in files:
        text = _extract_file_text(f)
        hits = _search_text(text, term)
        results.append({'name': f.filename, 'hits': hits, 'total': len(hits)})
    return jsonify({'term': term, 'results': results})


@app.route('/audit-tools/analyze', methods=['POST'])
@require_jwt
def audit_analyze():
    """Extract text from files and analyze with Ollama or Claude."""
    files = request.files.getlist('files')
    prompt = (request.form.get('prompt') or '').strip()
    mode = (request.form.get('mode') or 'ask').strip()
    engine = (request.form.get('engine') or 'ollama').strip()

    if not files or not files[0].filename:
        return jsonify({'error': 'No files uploaded'}), 400
    if not prompt:
        return jsonify({'error': 'Prompt required'}), 400
    if mode not in _AUDIT_SYSTEM:
        mode = 'ask'
    if engine not in ('ollama', 'claude'):
        engine = 'ollama'

    docs = []
    for f in files:
        text = _extract_file_text(f)
        docs.append({'name': f.filename, 'text': text})

    # Claude supports much larger context; Ollama cap stays at 8k per doc
    doc_cap = 40000 if engine == 'claude' else 8000
    context = '\n\n'.join(
        f'=== {d["name"]} ===\n{d["text"][:doc_cap]}'
        for d in docs
    )

    system_prompt = _AUDIT_SYSTEM[mode]
    if engine == 'claude':
        answer = _query_claude(system_prompt, context, prompt)
    else:
        answer = _query_ollama(system_prompt, context, prompt)

    return jsonify({
        'mode': mode,
        'engine': engine,
        'prompt': prompt,
        'docs': [{'name': d['name'], 'chars': len(d['text'])} for d in docs],
        'result': answer,
    })


@app.route('/audit-tools/ollama-status', methods=['GET'])
@require_jwt
def audit_ollama_status():
    """Check if local Ollama is running and whether Claude is configured."""
    claude_available = bool(ANTHROPIC_API_KEY and not ANTHROPIC_API_KEY.startswith('your_'))
    try:
        req = urllib.request.Request('http://localhost:11434/api/tags', method='GET')
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
            models = [m['name'] for m in data.get('models', [])]
            return jsonify({
                'running': True,
                'models': models,
                'active_model': OLLAMA_MODEL,
                'claude_available': claude_available,
                'claude_model': CLAUDE_MODEL,
            })
    except Exception:
        return jsonify({
            'running': False,
            'models': [],
            'active_model': OLLAMA_MODEL,
            'claude_available': claude_available,
            'claude_model': CLAUDE_MODEL,
        })


# ---------------------------------------------------------------------------
# AI Chat — local Ollama assistant
# ---------------------------------------------------------------------------

import datetime as _dt


@app.route('/email/summarize', methods=['POST'])
@require_jwt
def email_summarize():
    data = request.get_json(silent=True) or {}
    subject   = (data.get('subject')   or '').strip()
    from_name = (data.get('from_name') or '').strip()
    body      = (data.get('body')      or '').strip()
    if not body:
        return jsonify({'error': 'body required'}), 400
    system_prompt = (
        'You are an email assistant for Jeffrey Lobel, CPA. '
        'Summarize the key points of this email in 2-3 short bullet points (use • as the bullet). '
        'Focus on action items, requests, and important information. '
        'Each bullet should be one concise sentence. Return only the bullets, no preamble.'
    )
    result = _query_ollama(system_prompt, '', f'From: {from_name}\nSubject: {subject}\n\n{body}')
    return jsonify({'summary': result})


@app.route('/email/draft', methods=['POST'])
@require_jwt
def email_draft():
    data = request.get_json(silent=True) or {}
    subject   = (data.get('subject')   or '').strip()
    from_name = (data.get('from_name') or '').strip()
    body      = (data.get('body')      or '').strip()
    if not body:
        return jsonify({'error': 'body required'}), 400
    system_prompt = (
        'You are writing a professional email reply for Jeffrey Lobel, CPA at Lobel Accountancy Corporation. '
        'Write a concise, professional reply to this email. '
        'Do not include a subject line, greeting salutation opener like "Dear X", or signature — just the reply body paragraphs. '
        'Be warm but professional. Keep it brief unless a detailed response is clearly needed.'
    )
    result = _query_ollama(system_prompt, '', f'From: {from_name}\nSubject: {subject}\n\nOriginal email:\n{body}')
    return jsonify({'draft': result})


def _build_chat_context():
    lines = [
        f'Today: {_dt.date.today().isoformat()}',
        'Firm: Lobel Accountancy Corporation (LAC), CPA firm',
        '',
    ]
    try:
        with open(_PBC_DATA_PATH) as f:
            pbc = json.load(f)
        lines.append('PBC REQUEST STATUS:')
        for client, eng in pbc.items():
            if eng.get('archived'):
                continue
            reqs = eng.get('requests', [])
            counts = {}
            for r in reqs:
                s = r.get('status', 'Unknown')
                counts[s] = counts.get(s, 0) + 1
            lines.append(f'  {client} — {eng.get("engagement", "")}:')
            for status, n in counts.items():
                lines.append(f'    {status}: {n}')
            pending = [r for r in reqs if r.get('status') not in ('Complete', 'Provided')]
            if pending:
                lines.append(f'    Open items ({len(pending)}):')
                for r in pending[:15]:
                    lines.append(
                        f'      [{r.get("request_number","")}] {r.get("description","")} '
                        f'(due {r.get("due_date","")}, {r.get("status","")})'
                    )
    except Exception:
        pass
    return '\n'.join(lines)


@app.route('/chat', methods=['POST'])
@require_jwt
def chat():
    data = request.get_json(silent=True) or {}
    user_message = (data.get('message') or '').strip()
    if not user_message:
        return jsonify({'error': 'message required'}), 400
    if len(user_message) > 2000:
        return jsonify({'error': 'message too long'}), 400

    context = _build_chat_context()
    system_prompt = (
        'You are the internal dashboard assistant for LAC (Lobel Accountancy Corporation). '
        'You help Jeffrey Lobel, the firm owner and CPA, with questions about clients, '
        'PBC requests, deadlines, and firm operations. '
        'Be concise and accurate. Use the provided context. '
        'If you do not know something, say so.'
    )
    answer = _query_ollama(system_prompt, context, user_message)
    return jsonify({'answer': answer})


# ---------------------------------------------------------------------------
# Client portal — static file serving for clients.lobelaccountancy.com
# ---------------------------------------------------------------------------

@app.route('/', methods=['GET'])
@app.route('/<path:path>', methods=['GET'])
def portal_static(path='auth.html'):
    host = request.host.split(':')[0]
    if host != 'clients.lobelaccountancy.com':
        if request.path == '/':
            return redirect('https://dashboard.lobelaccountancy.com')
        return jsonify({'error': 'Not found'}), 404
    portal_dir = os.path.abspath(PORTAL_DIR)
    try:
        return send_from_directory(portal_dir, path)
    except Exception:
        return send_from_directory(portal_dir, 'auth.html')


# ---------------------------------------------------------------------------
# AI Tool endpoints — Clockify, Calendar write, DocuSeal status, Stirling PDF
# ---------------------------------------------------------------------------

@app.route('/clockify/entries', methods=['GET'])
@require_jwt
def clockify_entries():
    """Return time entries for today or this week. ?range=today|week (default today)"""
    import re as _re
    CK_API_KEY   = os.getenv('CLOCKIFY_API_KEY')
    CK_WORKSPACE = os.getenv('CLOCKIFY_WORKSPACE_ID')
    CK_USER_ID   = os.getenv('CLOCKIFY_USER_ID')
    rng = request.args.get('range', 'today')
    now_utc = datetime.now(timezone.utc)
    if rng == 'week':
        start = (now_utc - timedelta(days=now_utc.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        start = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
    end = now_utc

    def _dur_hrs(s):
        if not s or s == 'PT0S':
            return 0.0
        m = _re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?', s or '')
        if not m:
            return 0.0
        return round(int(m.group(1) or 0) + int(m.group(2) or 0) / 60, 2)

    try:
        url = f"https://api.clockify.me/api/v1/workspaces/{CK_WORKSPACE}/user/{CK_USER_ID}/time-entries"
        resp = requests.get(url, headers={'X-Api-Key': CK_API_KEY}, params={
            'start': start.strftime('%Y-%m-%dT%H:%M:%SZ'),
            'end':   end.strftime('%Y-%m-%dT%H:%M:%SZ'),
            'hydrated': True,
            'limit': 50,
        })
        entries = []
        for e in resp.json():
            interval = e.get('timeInterval', {})
            dur = _dur_hrs(interval.get('duration', ''))
            entries.append({
                'id':          e.get('id'),
                'description': e.get('description', ''),
                'project':     (e.get('project') or {}).get('name', ''),
                'task':        (e.get('task') or {}).get('name', ''),
                'start':       interval.get('start', ''),
                'end':         interval.get('end', ''),
                'hours':       dur,
            })
        total = round(sum(e['hours'] for e in entries), 2)
        return jsonify({'entries': entries, 'total_hours': total, 'range': rng})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/clockify/log-time', methods=['POST'])
@require_jwt
def clockify_log_time():
    """Log a time entry. Body: {description, project_id, task_id, hours, date (YYYY-MM-DD, default today)}"""
    import re as _re
    body       = request.get_json(force=True) or {}
    desc       = str(body.get('description', '')).strip()
    project_id = str(body.get('project_id', '')).strip()
    task_id    = str(body.get('task_id', '')).strip()
    hours      = body.get('hours')
    work_date  = str(body.get('date', '')).strip() or date.today().isoformat()

    if not desc or not project_id or hours is None:
        return jsonify({'error': 'description, project_id, hours required'}), 400
    try:
        hours = float(hours)
    except (TypeError, ValueError):
        return jsonify({'error': 'hours must be a number'}), 400

    CK_API_KEY   = os.getenv('CLOCKIFY_API_KEY')
    CK_WORKSPACE = os.getenv('CLOCKIFY_WORKSPACE_ID')

    # Build ISO 8601 interval from work_date + hours
    start_dt = datetime.strptime(work_date, '%Y-%m-%d').replace(hour=9, minute=0, tzinfo=timezone.utc)
    total_min = round(hours * 60)
    end_dt    = start_dt + timedelta(minutes=total_min)

    payload = {
        'start':       start_dt.strftime('%Y-%m-%dT%H:%M:%SZ'),
        'end':         end_dt.strftime('%Y-%m-%dT%H:%M:%SZ'),
        'description': desc,
        'projectId':   project_id,
        'billable':    True,
    }
    if task_id:
        payload['taskId'] = task_id

    try:
        url  = f"https://api.clockify.me/api/v1/workspaces/{CK_WORKSPACE}/time-entries"
        resp = requests.post(url, headers={'X-Api-Key': CK_API_KEY, 'Content-Type': 'application/json'}, json=payload)
        if resp.status_code not in (200, 201):
            return jsonify({'error': f'Clockify error {resp.status_code}', 'detail': resp.text}), 502
        entry = resp.json()
        return jsonify({'success': True, 'entry_id': entry.get('id'), 'hours': hours, 'description': desc})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/calendar/create', methods=['POST'])
@require_jwt
def calendar_create():
    """Create a Google Calendar event. Body: {title, start, end, description, location}
    start/end: ISO 8601 datetime or YYYY-MM-DD for all-day events."""
    body      = request.get_json(force=True) or {}
    title     = str(body.get('title', '')).strip()
    start_raw = str(body.get('start', '')).strip()
    end_raw   = str(body.get('end', '')).strip()
    desc      = str(body.get('description', '')).strip()
    location  = str(body.get('location', '')).strip()

    if not title or not start_raw:
        return jsonify({'error': 'title and start are required'}), 400

    all_day = 'T' not in start_raw and len(start_raw) == 10  # YYYY-MM-DD

    if all_day:
        start_obj = {'date': start_raw}
        end_obj   = {'date': end_raw or start_raw}
    else:
        # Ensure timezone suffix
        if not start_raw.endswith('Z') and '+' not in start_raw:
            start_raw += '-07:00'
        if end_raw and not end_raw.endswith('Z') and '+' not in end_raw:
            end_raw += '-07:00'
        if not end_raw:
            # Default to 1 hour
            from dateutil import parser as _dp
            s = _dp.parse(start_raw)
            end_raw = (s + timedelta(hours=1)).isoformat()
        start_obj = {'dateTime': start_raw}
        end_obj   = {'dateTime': end_raw}

    event_body = {'summary': title, 'start': start_obj, 'end': end_obj}
    if desc:     event_body['description'] = desc
    if location: event_body['location']    = location

    cal_id = os.getenv('GOOGLE_CALENDAR_ID', 'jlobel@lobelaccountancy.com')
    try:
        # Need write scope — build with full calendar scope credentials
        creds = service_account.Credentials.from_service_account_file(
            CREDENTIALS_PATH,
            scopes=['https://www.googleapis.com/auth/calendar'])
        svc   = build('calendar', 'v3', credentials=creds)
        ev    = svc.events().insert(calendarId=cal_id, body=event_body).execute()
        return jsonify({'success': True, 'event_id': ev.get('id'), 'title': title,
                        'link': ev.get('htmlLink', '')})
    except Exception as exc:
        app.logger.error('Calendar create failed: %s', exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/docuseal/submissions', methods=['GET'])
@require_jwt
def docuseal_submissions():
    """List recent DocuSeal submissions (documents sent for signature)."""
    if not DOCUSEAL_API_KEY:
        return jsonify({'error': 'DOCUSEAL_API_KEY not set'}), 503
    try:
        resp = requests.get(
            f'{DOCUSEAL_URL}/api/submissions',
            headers={'X-Auth-Token': DOCUSEAL_API_KEY},
            params={'limit': request.args.get('limit', 20)},
            timeout=10,
        )
        return jsonify(resp.json())
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/pdf/process', methods=['POST'])
@require_jwt
def pdf_process():
    """Proxy a file operation through Stirling PDF.
    Body: {operation, drive_file_id} OR upload multipart file.
    operation: merge, compress, ocr, rotate, split
    Returns: {ok, result_url} or {ok, text} for ocr."""
    import subprocess as _sp2, tempfile as _tf2
    body      = request.get_json(force=True) or {}
    operation = str(body.get('operation', 'ocr')).strip()
    drive_id  = str(body.get('drive_file_id', '')).strip()

    STIRLING = 'http://localhost:8081'
    OP_MAP = {
        'ocr':      ('/api/v1/misc/ocr-pdf',         {'languages': 'eng', 'ocrType': '2', 'ocrRenderType': 'hocr', 'sidecar': False}),
        'compress': ('/api/v1/misc/compress-pdf',     {'optimizeLevel': 2}),
        'merge':    ('/api/v1/general/merge-pdfs',   {}),
        'rotate':   ('/api/v1/general/rotate-pdf',   {'angle': body.get('angle', 90)}),
        'split':    ('/api/v1/general/split-pdf-by-sections', {'splitAfterPage': body.get('page', 1)}),
        'extract':  ('/api/v1/misc/extract-images',  {'format': 'pdf'}),
    }
    if operation not in OP_MAP:
        return jsonify({'error': f'Unknown operation. Choose: {", ".join(OP_MAP)}'}), 400

    if not drive_id:
        return jsonify({'error': 'drive_file_id required'}), 400

    # Download from Drive
    try:
        svc  = _drive_rw_service()
        req  = svc.files().get_media(fileId=drive_id)
        buf  = io.BytesIO()
        dl   = googleapiclient.http.MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        pdf_bytes = buf.getvalue()
    except Exception as exc:
        return jsonify({'error': f'Drive download failed: {exc}'}), 502

    path, extra_params = OP_MAP[operation]
    files = {'fileInput': ('input.pdf', pdf_bytes, 'application/pdf')}

    try:
        resp = requests.post(f'{STIRLING}{path}', files=files, data=extra_params, timeout=60)
    except Exception as exc:
        return jsonify({'error': f'Stirling PDF error: {exc}'}), 502

    if resp.status_code != 200:
        return jsonify({'error': f'Stirling PDF returned {resp.status_code}'}), 502

    ct = resp.headers.get('Content-Type', '')
    if 'application/pdf' in ct:
        # Save result to /tmp and return a download URL
        out = f'/tmp/stirling_{operation}_{drive_id[:8]}.pdf'
        with open(out, 'wb') as f:
            f.write(resp.content)
        return jsonify({'ok': True, 'operation': operation, 'bytes': len(resp.content),
                        'note': f'Result saved to {out}. Use /pdf/download to retrieve it.',
                        'result_path': out})
    else:
        return jsonify({'ok': True, 'operation': operation, 'text': resp.text[:2000]})


# ---------------------------------------------------------------------------
# OpenRouter proxy — filters to free-only models so paid options never appear
# Open WebUI connects here instead of openrouter.ai directly
# ---------------------------------------------------------------------------

_OPENROUTER_KEY = os.getenv('OPENROUTER_API_KEY', '')
_OPENROUTER_BASE = 'https://openrouter.ai/api/v1'


@app.route('/openrouter/v1/models', methods=['GET'])
def openrouter_models():
    """Return only free OpenRouter models."""
    key = _OPENROUTER_KEY
    if not key:
        return jsonify({'error': 'OPENROUTER_API_KEY not set'}), 503
    try:
        resp = requests.get(f'{_OPENROUTER_BASE}/models',
                            headers={'Authorization': f'Bearer {key}'}, timeout=15)
        data = resp.json().get('data', [])
        free = [m for m in data
                if float(m.get('pricing', {}).get('prompt', 1)) == 0
                and m['id'].endswith(':free')]
        return jsonify({'object': 'list', 'data': free})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 502


@app.route('/openrouter/v1/chat/completions', methods=['POST'])
def openrouter_chat():
    """Proxy chat completions to OpenRouter, enforcing free-only models."""
    key = _OPENROUTER_KEY
    if not key:
        return jsonify({'error': 'OPENROUTER_API_KEY not set'}), 503

    body = request.get_json(force=True) or {}
    model = body.get('model', '')

    # Refuse if the requested model is not a free model
    if not model.endswith(':free'):
        return jsonify({'error': f'Model "{model}" is not a free model. Only :free models are allowed.'}), 400

    stream = body.get('stream', False)
    headers = {
        'Authorization': f'Bearer {key}',
        'Content-Type': 'application/json',
        'HTTP-Referer': 'https://ai.lobelaccountancy.com',
        'X-Title': 'LAC AI',
    }
    try:
        r = requests.post(f'{_OPENROUTER_BASE}/chat/completions',
                          json=body, headers=headers,
                          stream=stream, timeout=120)
        if stream:
            def generate():
                for chunk in r.iter_content(chunk_size=None):
                    yield chunk
            return Response(generate(), status=r.status_code,
                            content_type=r.headers.get('Content-Type', 'text/event-stream'))
        return Response(r.content, status=r.status_code,
                        content_type=r.headers.get('Content-Type', 'application/json'))
    except Exception as exc:
        return jsonify({'error': str(exc)}), 502


# ---------------------------------------------------------------------------
# Open WebUI Tool Server — OpenAPI spec + tool call endpoints
# Configure in Open WebUI: Admin Panel → Settings → Tools
#   URL: https://auth.lobelaccountancy.com
#   Path: /ai/openapi.json
#   Auth: Bearer  Key: <service JWT>
# ---------------------------------------------------------------------------

_AI_TOOLS_SPEC = {
    "openapi": "3.0.0",
    "info": {"title": "LAC Tools", "version": "1.0.0",
             "description": "LAC Workbook, Clockify, Google Calendar, DocuSeal, Stirling PDF"},
    "paths": {
        "/ai/call/list_clients": {"post": {
            "operationId": "list_clients",
            "summary": "List all clients in Engagement Pipeline and AR Aging",
            "requestBody": {"required": False, "content": {"application/json": {"schema": {"type": "object", "properties": {}}}}},
            "responses": {"200": {"description": "Client list"}}}},
        "/ai/call/get_pipeline_summary": {"post": {
            "operationId": "get_pipeline_summary",
            "summary": "Get Engagement Pipeline grouped by stage (Prospect, Engaged, Complete, Lost)",
            "requestBody": {"required": False, "content": {"application/json": {"schema": {"type": "object", "properties": {}}}}},
            "responses": {"200": {"description": "Pipeline summary"}}}},
        "/ai/call/get_ar_summary": {"post": {
            "operationId": "get_ar_summary",
            "summary": "Get all outstanding (unpaid/partial) Accounts Receivable",
            "requestBody": {"required": False, "content": {"application/json": {"schema": {"type": "object", "properties": {}}}}},
            "responses": {"200": {"description": "AR summary"}}}},
        "/ai/call/add_prospect": {"post": {
            "operationId": "add_prospect",
            "summary": "Add a new client to the Engagement Pipeline spreadsheet",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "required": ["client_name", "engagement_type", "email"],
                "properties": {
                    "client_name":      {"type": "string", "description": "Full client/company name"},
                    "engagement_type":  {"type": "string", "description": "e.g. Tax Return, Audit, Bookkeeping"},
                    "email":            {"type": "string", "description": "Client email address"},
                    "fee":              {"type": "number", "description": "Fee amount in dollars"},
                    "billing_type":     {"type": "string", "description": "Fixed Fee or Hourly"},
                    "entity_type":      {"type": "string", "description": "e.g. Individual, LLC, S-Corp"},
                    "fye":              {"type": "string", "description": "Fiscal year end MM/DD/YYYY"},
                    "address":          {"type": "string", "description": "Client mailing address"},
                    "client_title":     {"type": "string", "description": "Contact name/title"},
                    "stage":            {"type": "string", "description": "Prospect, Engaged, Complete, or Lost"},
                }}}}},
            "responses": {"200": {"description": "Confirmation"}}}},
        "/ai/call/update_engagement": {"post": {
            "operationId": "update_engagement",
            "summary": "Update a client record in the Engagement Pipeline (stage, fee, email, etc.)",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "required": ["client_name"],
                "properties": {
                    "client_name":     {"type": "string"},
                    "stage":           {"type": "string", "description": "Prospect, Engaged, Complete, Lost"},
                    "status":          {"type": "string", "description": "Active or Inactive"},
                    "fee":             {"type": "number"},
                    "email":           {"type": "string"},
                    "issue_date":      {"type": "string", "description": "MM/DD/YYYY"},
                    "fye":             {"type": "string"},
                    "engagement_type": {"type": "string"},
                    "billing_type":    {"type": "string"},
                    "entity_type":     {"type": "string"},
                    "address":         {"type": "string"},
                    "client_title":    {"type": "string"},
                }}}}},
            "responses": {"200": {"description": "Confirmation"}}}},
        "/ai/call/add_ar_entry": {"post": {
            "operationId": "add_ar_entry",
            "summary": "Add a new invoice row to the AR Aging spreadsheet",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "required": ["client_name", "invoice_number", "amount", "service_description"],
                "properties": {
                    "client_name":          {"type": "string"},
                    "invoice_number":       {"type": "string", "description": "e.g. INV-202606-003"},
                    "amount":              {"type": "number"},
                    "service_description": {"type": "string"},
                    "email":               {"type": "string"},
                    "invoice_date":        {"type": "string", "description": "MM/DD/YYYY"},
                    "due_date":            {"type": "string", "description": "MM/DD/YYYY"},
                    "paid":                {"type": "number", "description": "Amount already paid"},
                }}}}},
            "responses": {"200": {"description": "Confirmation"}}}},
        "/ai/call/update_ar_entry": {"post": {
            "operationId": "update_ar_entry",
            "summary": "Update an AR Aging row — mark paid, change status, etc.",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "properties": {
                    "client_name":    {"type": "string"},
                    "invoice_number": {"type": "string"},
                    "paid":           {"type": "number", "description": "Amount paid"},
                    "status":         {"type": "string", "description": "Unpaid, Partial, Paid"},
                }}}}},
            "responses": {"200": {"description": "Confirmation"}}}},
        "/ai/call/get_calendar_events": {"post": {
            "operationId": "get_calendar_events",
            "summary": "Get upcoming Google Calendar events for the next 14 days",
            "requestBody": {"required": False, "content": {"application/json": {"schema": {"type": "object", "properties": {}}}}},
            "responses": {"200": {"description": "Event list"}}}},
        "/ai/call/create_calendar_event": {"post": {
            "operationId": "create_calendar_event",
            "summary": "Create a Google Calendar event",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "required": ["title", "start"],
                "properties": {
                    "title":       {"type": "string"},
                    "start":       {"type": "string", "description": "ISO 8601 datetime or YYYY-MM-DD for all-day"},
                    "end":         {"type": "string", "description": "ISO 8601 datetime (default 1hr after start)"},
                    "description": {"type": "string"},
                    "location":    {"type": "string"},
                }}}}},
            "responses": {"200": {"description": "Confirmation"}}}},
        "/ai/call/get_time_entries": {"post": {
            "operationId": "get_time_entries",
            "summary": "Get Clockify time entries logged today or this week",
            "requestBody": {"required": False, "content": {"application/json": {"schema": {"type": "object",
                "properties": {
                    "range": {"type": "string", "description": "today or week (default today)"},
                }}}}},
            "responses": {"200": {"description": "Time entries"}}}},
        "/ai/call/get_clockify_tasks": {"post": {
            "operationId": "get_clockify_tasks",
            "summary": "List all Clockify tasks with IDs. Call this before log_time to get project_id and task_id.",
            "requestBody": {"required": False, "content": {"application/json": {"schema": {"type": "object", "properties": {}}}}},
            "responses": {"200": {"description": "Task list"}}}},
        "/ai/call/log_time": {"post": {
            "operationId": "log_time",
            "summary": "Log time to Clockify. Call get_clockify_tasks first to find project_id.",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "required": ["description", "project_id", "hours"],
                "properties": {
                    "description": {"type": "string"},
                    "project_id":  {"type": "string"},
                    "task_id":     {"type": "string"},
                    "hours":       {"type": "number"},
                    "work_date":   {"type": "string", "description": "YYYY-MM-DD (default today)"},
                }}}}},
            "responses": {"200": {"description": "Confirmation"}}}},
        "/ai/call/list_docuseal_templates": {"post": {
            "operationId": "list_docuseal_templates",
            "summary": "List DocuSeal e-signature templates",
            "requestBody": {"required": False, "content": {"application/json": {"schema": {"type": "object", "properties": {}}}}},
            "responses": {"200": {"description": "Template list"}}}},
        "/ai/call/send_engagement_letter": {"post": {
            "operationId": "send_engagement_letter",
            "summary": "Send an engagement letter via DocuSeal to a client",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "required": ["client_name", "template_id"],
                "properties": {
                    "client_name": {"type": "string"},
                    "template_id": {"type": "integer"},
                }}}},},
            "responses": {"200": {"description": "Confirmation"}}}},
        "/ai/call/list_pending_signatures": {"post": {
            "operationId": "list_pending_signatures",
            "summary": "List documents pending e-signature in DocuSeal",
            "requestBody": {"required": False, "content": {"application/json": {"schema": {"type": "object", "properties": {}}}}},
            "responses": {"200": {"description": "Signature request list"}}}},
        "/ai/call/process_pdf": {"post": {
            "operationId": "process_pdf",
            "summary": "Run a PDF operation (compress, OCR, rotate, split) on a Google Drive file via Stirling PDF",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "required": ["drive_file_id", "operation"],
                "properties": {
                    "drive_file_id": {"type": "string", "description": "Google Drive file ID (from the file URL)"},
                    "operation":     {"type": "string", "description": "compress, ocr, rotate, split, or extract"},
                    "angle":         {"type": "integer", "description": "Rotation angle for rotate operation: 90, 180, or 270"},
                    "page":          {"type": "integer", "description": "Page number to split after for split operation"},
                }}}}},
            "responses": {"200": {"description": "Result or extracted text"}}}},
        "/ai/call/set_clockify_estimate": {"post": {
            "operationId": "set_clockify_estimate",
            "summary": "Set the hour budget/estimate on a Clockify task. Call get_clockify_tasks first to find task_id and project_id.",
            "requestBody": {"required": True, "content": {"application/json": {"schema": {"type": "object",
                "required": ["task_id", "task_name", "project_id", "budget_hrs"],
                "properties": {
                    "task_id":    {"type": "string", "description": "Clockify task ID"},
                    "task_name":  {"type": "string", "description": "Clockify task name"},
                    "project_id": {"type": "string", "description": "Clockify project ID"},
                    "budget_hrs": {"type": "number", "description": "Hour estimate to set"},
                }}}}},
            "responses": {"200": {"description": "Confirmation"}}}},
    }
}


@app.route('/ai/openapi.json', methods=['GET'])
def ai_openapi():
    return jsonify(_AI_TOOLS_SPEC)


@app.route('/ai/call/<tool_name>', methods=['POST'])
@require_jwt
def ai_tool_call(tool_name):
    """Dispatch tool calls from Open WebUI."""
    body = request.get_json(force=True) or {}

    # ── LAC Workbook ──────────────────────────────────────────────────────
    if tool_name == 'list_clients':
        data = _get_wb_clients_data()
        lines = ['=== Engagement Pipeline ===']
        for c in data.get('pipeline', []):
            lines.append(f"  {c['client']} | {c['stage']} | {c['eng_type']} | {c.get('email','')}")
        lines.append('\n=== AR Aging ===')
        for r in data.get('ar', []):
            lines.append(f"  {r['client']} | {r['invoice']} | ${r['amount']} | {r['status']}")
        return jsonify({'result': '\n'.join(lines)})

    if tool_name == 'get_pipeline_summary':
        data = _get_wb_clients_data()
        by_stage = defaultdict(list)
        for r in data.get('pipeline', []):
            by_stage[r['stage']].append(r['client'])
        text = '\n'.join(f"{s}: {', '.join(cs)}" for s, cs in sorted(by_stage.items())) or 'Pipeline is empty.'
        return jsonify({'result': text})

    if tool_name == 'get_ar_summary':
        data = _get_wb_clients_data()
        rows = [r for r in data.get('ar', []) if r.get('status', '').lower() != 'paid']
        if not rows:
            return jsonify({'result': 'All AR is current — no outstanding balances.'})
        lines = ['Outstanding AR:']
        for r in rows:
            lines.append(f"  {r['client']} | {r['invoice']} | ${r['amount']} | paid ${r.get('paid',0)} | {r['status']}")
        return jsonify({'result': '\n'.join(lines)})

    if tool_name == 'add_prospect':
        with app.test_request_context(json=body):
            # Re-use wb/prospect/add logic via internal call
            pass
        result = _wb_add_prospect_logic(body)
        return jsonify({'result': result})

    if tool_name == 'update_engagement':
        result = _wb_update_engagement_logic(body)
        return jsonify({'result': result})

    if tool_name == 'add_ar_entry':
        result = _wb_add_ar_logic(body)
        return jsonify({'result': result})

    if tool_name == 'update_ar_entry':
        result = _wb_update_ar_logic(body)
        return jsonify({'result': result})

    # ── Google Calendar ───────────────────────────────────────────────────
    if tool_name == 'get_calendar_events':
        cal_id = os.getenv('GOOGLE_CALENDAR_ID', 'jlobel@lobelaccountancy.com')
        try:
            svc    = _calendar_service()
            now_u  = datetime.now(timezone.utc).isoformat()
            end_u  = (datetime.now(timezone.utc) + timedelta(days=14)).isoformat()
            result = svc.events().list(calendarId=cal_id, timeMin=now_u, timeMax=end_u,
                                       maxResults=20, singleEvents=True, orderBy='startTime').execute()
            lines = []
            for ev in result.get('items', []):
                start = ev['start'].get('dateTime', ev['start'].get('date', ''))[:16].replace('T', ' ')
                lines.append(f"  {start}  {ev.get('summary', '(no title)')}  {ev.get('location', '')}".rstrip())
            return jsonify({'result': '\n'.join(lines) or 'No upcoming events.'})
        except Exception as exc:
            return jsonify({'result': f'Error: {exc}'})

    if tool_name == 'create_calendar_event':
        with app.test_request_context(json=body):
            resp = calendar_create()
        return jsonify({'result': str(resp.get_json())})

    # ── Clockify ──────────────────────────────────────────────────────────
    if tool_name == 'get_time_entries':
        rng = body.get('range', 'today')
        with app.test_request_context(f'/clockify/entries?range={rng}'):
            resp = clockify_entries()
        d = resp.get_json()
        if 'error' in d:
            return jsonify({'result': f"Error: {d['error']}"})
        entries = d.get('entries', [])
        total   = d.get('total_hours', 0)
        if not entries:
            return jsonify({'result': f'No time entries for {rng}.'})
        lines = [f"Time entries ({rng}) — {total}h total:"]
        for e in entries:
            label = ' / '.join(x for x in [e.get('project',''), e.get('task',''), e.get('description','')] if x)
            lines.append(f"  {e.get('hours',0)}h — {label}")
        return jsonify({'result': '\n'.join(lines)})

    if tool_name == 'get_clockify_tasks':
        with app.test_request_context('/clockify/tasks'):
            resp = clockify_tasks()
        d = resp.get_json()
        tasks = d.get('tasks', [])
        if not tasks:
            return jsonify({'result': 'No tasks found.'})
        lines = ['Clockify tasks:']
        for t in tasks:
            lines.append(f"  [{t['billing_type']}] {t['name']} | {t['budget_hrs']}h | task_id:{t['task_id']} | project_id:{t['project_id']}")
        return jsonify({'result': '\n'.join(lines)})

    if tool_name == 'log_time':
        with app.test_request_context(json=body):
            resp = clockify_log_time()
        return jsonify({'result': str(resp.get_json())})

    # ── DocuSeal ──────────────────────────────────────────────────────────
    if tool_name == 'list_docuseal_templates':
        try:
            resp = requests.get(f'{DOCUSEAL_URL}/api/templates',
                                headers={'X-Auth-Token': DOCUSEAL_API_KEY}, timeout=10)
            templates = resp.json()
            if not templates:
                return jsonify({'result': 'No templates found.'})
            lines = ['DocuSeal templates:']
            for t in (templates if isinstance(templates, list) else templates.get('data', [])):
                lines.append(f"  ID {t['id']}: {t['name']}")
            return jsonify({'result': '\n'.join(lines)})
        except Exception as exc:
            return jsonify({'result': f'Error: {exc}'})

    if tool_name == 'send_engagement_letter':
        with app.test_request_context(json=body):
            resp = engagement_send_letter()
        return jsonify({'result': str(resp.get_json())})

    if tool_name == 'list_pending_signatures':
        try:
            resp = requests.get(f'{DOCUSEAL_URL}/api/submissions',
                                headers={'X-Auth-Token': DOCUSEAL_API_KEY}, timeout=10)
            items = resp.json().get('data', [])
            if not items:
                return jsonify({'result': 'No signature requests found.'})
            lines = ['Signature requests:']
            for s in items:
                tmpl = s.get('template', {}).get('name', '')
                for sub in s.get('submitters', []):
                    signed = '✓ signed' if sub.get('completed_at') else '⏳ pending'
                    lines.append(f"  {tmpl} → {sub.get('name','')} <{sub.get('email','')}> [{signed}]")
            return jsonify({'result': '\n'.join(lines)})
        except Exception as exc:
            return jsonify({'result': f'Error: {exc}'})

    if tool_name == 'process_pdf':
        with app.test_request_context(json=body):
            resp = pdf_process()
        return jsonify({'result': str(resp.get_json())})

    if tool_name == 'set_clockify_estimate':
        with app.test_request_context(json=body):
            resp = clockify_set_estimate()
        d = resp.get_json()
        if d.get('success'):
            return jsonify({'result': f"Set {d['task_name']} estimate to {d['budget_hrs']}h"})
        return jsonify({'result': d.get('error', str(d))})

    return jsonify({'error': f'Unknown tool: {tool_name}'}), 404


def _get_wb_clients_data():
    """Shared helper for tool server — read pipeline and AR without HTTP round-trip."""
    wb = _workbook()
    ep_ws = wb['Engagement Pipeline']
    ar_ws = wb['AR Aging']
    pipeline, ar = [], []
    for row in ep_ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        pipeline.append({'client': row[0] or '', 'client_num': row[1] or '', 'stage': row[2] or '',
                         'eng_type': row[3] or '', 'fee': row[6] or 0, 'email': row[12] or ''})
    for row in ar_ws.iter_rows(min_row=13, values_only=True):
        if not row[0]:
            continue
        ar.append({'client': row[0] or '', 'invoice': row[2] or '', 'amount': row[6] or 0,
                   'paid': row[7] or 0, 'status': row[10] or ''})
    return {'pipeline': pipeline, 'ar': ar}


def _wb_add_prospect_logic(body):
    try:
        wb, fid, svc = _wb_download_writable()
        ws = wb['Engagement Pipeline']
        next_row = ws.max_row + 1
        num = _next_client_num(ws)
        today = date.today().strftime('%m/%d/%Y')
        ws.cell(next_row, 1,  body.get('client_name', ''))
        ws.cell(next_row, 2,  num)
        ws.cell(next_row, 3,  body.get('stage', 'Prospect'))
        ws.cell(next_row, 4,  body.get('engagement_type', ''))
        ws.cell(next_row, 5,  body.get('billing_type', 'Fixed Fee'))
        ws.cell(next_row, 6,  body.get('entity_type', ''))
        ws.cell(next_row, 7,  float(body.get('fee', 0)))
        ws.cell(next_row, 8,  body.get('status', 'Active'))
        ws.cell(next_row, 9,  today)
        ws.cell(next_row, 10, body.get('issue_date', ''))
        ws.cell(next_row, 11, body.get('fye', ''))
        ws.cell(next_row, 12, body.get('client_title', ''))
        ws.cell(next_row, 13, body.get('email', ''))
        ws.cell(next_row, 14, body.get('address', ''))
        _wb_save_and_upload(wb, fid, svc)
        return f"Added {body.get('client_name')} as {body.get('stage','Prospect')} (client #{num})"
    except Exception as exc:
        return f'Error: {exc}'


def _wb_update_engagement_logic(body):
    try:
        client = body.get('client_name', '')
        updates = {k: v for k, v in body.items() if k != 'client_name' and v not in (None, '')}
        wb, fid, svc = _wb_download_writable()
        ws = wb['Engagement Pipeline']
        row_idx = _find_ep_row(ws, client)
        if row_idx is None:
            return f'Client "{client}" not found in Engagement Pipeline'
        field_map = {'stage':5,'status':8,'fee':7,'email':13,'issue_date':10,
                     'fye':11,'eng_type':4,'billing_type':5,'entity_type':6,'address':14,'client_title':12}
        # Correct map: col index (1-based)
        col_map = {'stage':3,'eng_type':4,'billing_type':5,'entity_type':6,'fee':7,
                   'status':8,'issue_date':10,'fye':11,'client_title':12,'email':13,'address':14}
        changed = []
        for field, val in updates.items():
            col = col_map.get(field)
            if col:
                ws.cell(row_idx, col, val)
                changed.append(field)
        _wb_save_and_upload(wb, fid, svc)
        return f"Updated {client}: {', '.join(changed)}"
    except Exception as exc:
        return f'Error: {exc}'


def _wb_add_ar_logic(body):
    try:
        wb, fid, svc = _wb_download_writable()
        ep_ws = wb['Engagement Pipeline']
        ar_ws = wb['AR Aging']
        client = body.get('client_name', '')
        # Look up client number
        client_num = ''
        for row in ep_ws.iter_rows(min_row=3, values_only=True):
            if row[0] and row[0].strip().lower() == client.strip().lower():
                client_num = str(row[1] or '')
                break
        today = date.today().strftime('%m/%d/%Y')
        amount = float(body.get('amount', 0))
        paid   = float(body.get('paid', 0))
        outstanding = round(amount - paid, 2)
        status = 'Paid' if outstanding <= 0 else ('Partial' if paid > 0 else 'Unpaid')
        inv_date = body.get('invoice_date', today)
        due_date = body.get('due_date', today)
        next_row = ar_ws.max_row + 1
        ar_ws.cell(next_row, 1,  client)
        ar_ws.cell(next_row, 2,  client_num)
        ar_ws.cell(next_row, 3,  body.get('invoice_number', ''))
        ar_ws.cell(next_row, 4,  inv_date)
        ar_ws.cell(next_row, 5,  due_date)
        ar_ws.cell(next_row, 6,  body.get('service_description', ''))
        ar_ws.cell(next_row, 7,  amount)
        ar_ws.cell(next_row, 8,  paid)
        ar_ws.cell(next_row, 9,  outstanding)
        ar_ws.cell(next_row, 11, status)
        ar_ws.cell(next_row, 12, body.get('email', ''))
        _wb_save_and_upload(wb, fid, svc)
        return f"Added AR entry {body.get('invoice_number','')} for {client}: ${amount} ({status})"
    except Exception as exc:
        return f'Error: {exc}'


def _wb_update_ar_logic(body):
    try:
        wb, fid, svc = _wb_download_writable()
        ar_ws = wb['AR Aging']
        client  = body.get('client_name', '')
        inv_num = body.get('invoice_number', '')
        row_idx = _find_ar_row(ar_ws, client, inv_num)
        if row_idx is None:
            return f'AR row not found for client="{client}" invoice="{inv_num}"'
        if 'paid' in body:
            paid   = float(body['paid'])
            amount = float(ar_ws.cell(row_idx, 7).value or 0)
            outstanding = round(amount - paid, 2)
            status = 'Paid' if outstanding <= 0 else ('Partial' if paid > 0 else 'Unpaid')
            ar_ws.cell(row_idx, 8,  paid)
            ar_ws.cell(row_idx, 9,  outstanding)
            ar_ws.cell(row_idx, 11, body.get('status', status))
        elif 'status' in body:
            ar_ws.cell(row_idx, 11, body['status'])
        _wb_save_and_upload(wb, fid, svc)
        return f"Updated AR row for {client or inv_num}"
    except Exception as exc:
        return f'Error: {exc}'


# ---------------------------------------------------------------------------
# Bookmarks  (/bookmarks)
# ---------------------------------------------------------------------------

_BOOKMARKS_PATH = '/home/jlobel/lac_automation/bookmarks.json'

def _load_bookmarks():
    try:
        with open(_BOOKMARKS_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def _save_bookmarks(data):
    with open(_BOOKMARKS_PATH, 'w') as f:
        json.dump(data, f, indent=2)


@app.route('/bookmarks', methods=['GET'])
@require_jwt
def bookmarks_get():
    return jsonify({'bookmarks': _load_bookmarks()})


@app.route('/bookmarks', methods=['POST'])
@require_jwt
def bookmarks_save():
    """Replace the full bookmark list. Body: { bookmarks: [...] }"""
    body = request.get_json(force=True) or {}
    items = body.get('bookmarks')
    if not isinstance(items, list):
        return jsonify({'error': 'bookmarks must be an array'}), 400
    # Validate each entry
    for item in items:
        if not isinstance(item, dict) or not item.get('name') or not item.get('url'):
            return jsonify({'error': 'each bookmark needs name and url'}), 400
        if not item.get('id'):
            item['id'] = item['name'].lower().replace(' ', '_')[:24]
    try:
        _save_bookmarks(items)
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500
    return jsonify({'success': True, 'count': len(items)})


# ---------------------------------------------------------------------------
# Apollo.io Prospecting  (/apollo/*)
# ---------------------------------------------------------------------------

_APOLLO_EMAIL_LOG = '/home/jlobel/lac_automation/apollo_email_log.json'
_APOLLO_BASE      = 'https://api.apollo.io/v1'
DAILY_EMAIL_CAP   = 25


def _apollo_headers():
    key = os.getenv('APOLLO_API_KEY', '')
    return {'Content-Type': 'application/json', 'X-Api-Key': key}


def _load_email_log():
    try:
        with open(_APOLLO_EMAIL_LOG) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_email_log(data):
    with open(_APOLLO_EMAIL_LOG, 'w') as f:
        json.dump(data, f, indent=2)


def _today_email_count():
    log = _load_email_log()
    today = date.today().isoformat()
    return len(log.get(today, []))


@app.route('/apollo/search', methods=['POST'])
@require_jwt
def apollo_search():
    """Search Apollo for prospects.
    Body: { titles, keywords, locations, industries, employee_ranges, page }
    """
    if not os.getenv('APOLLO_API_KEY'):
        return jsonify({'error': 'APOLLO_API_KEY not configured'}), 503

    body = request.get_json(force=True) or {}
    payload = {
        'page':     body.get('page', 1),
        'per_page': 25,
    }
    if body.get('titles'):
        payload['person_titles'] = body['titles']
    if body.get('keywords'):
        payload['q_keywords'] = body['keywords']
    if body.get('locations'):
        payload['person_locations'] = body['locations']
    if body.get('industries'):
        payload['organization_industries'] = body['industries']
    if body.get('employee_ranges'):
        payload['organization_num_employees_ranges'] = body['employee_ranges']
    if body.get('revenue_range'):
        payload['organization_annual_revenue_ranges'] = [body['revenue_range']]
    if body.get('departments'):
        payload['departments'] = body['departments']
    if body.get('founded_year_min'):
        payload['organization_founded_year_min'] = int(body['founded_year_min'])
    if body.get('founded_year_max'):
        payload['organization_founded_year_max'] = int(body['founded_year_max'])
    if body.get('verified_only'):
        payload['contact_email_status_cd'] = ['verified']
    if body.get('recent_job_change'):
        payload['recently_changed_jobs'] = True

    verified_only = bool(body.get('verified_only'))

    try:
        resp = requests.post(f"{_APOLLO_BASE}/mixed_people/search",
                             headers=_apollo_headers(), json=payload, timeout=15)
        if resp.status_code != 200:
            return jsonify({'error': f'Apollo {resp.status_code}', 'detail': resp.text[:300]}), 502
        data = resp.json()
        people = []
        for p in data.get('people', []):
            org = p.get('organization') or {}
            email_status = p.get('email_status', '')
            if verified_only and email_status != 'verified':
                continue
            people.append({
                'id':            p.get('id'),
                'name':          p.get('name', ''),
                'title':         p.get('title', ''),
                'company':       org.get('name', ''),
                'industry':      org.get('industry', ''),
                'employees':     org.get('estimated_num_employees'),
                'founded_year':  org.get('founded_year'),
                'annual_revenue':org.get('annual_revenue'),
                'city':          p.get('city', ''),
                'state':         p.get('state', ''),
                'email':         p.get('email', ''),
                'email_status':  email_status,
                'linkedin_url':  p.get('linkedin_url', ''),
                'website':       org.get('website_url', ''),
            })
        return jsonify({
            'people':     people,
            'total':      data.get('pagination', {}).get('total_entries', 0),
            'page':       data.get('pagination', {}).get('page', 1),
            'total_pages':data.get('pagination', {}).get('total_pages', 1),
        })
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/apollo/email-status', methods=['GET'])
@require_jwt
def apollo_email_status():
    """Return today's email send count and cap."""
    today = date.today().isoformat()
    log   = _load_email_log()
    sent  = log.get(today, [])
    return jsonify({
        'date':      today,
        'sent':      len(sent),
        'cap':       DAILY_EMAIL_CAP,
        'remaining': max(0, DAILY_EMAIL_CAP - len(sent)),
        'emails':    sent,
    })


@app.route('/apollo/send-email', methods=['POST'])
@require_jwt
def apollo_send_email():
    """Send a prospecting email via Apollo.
    Body: { prospect_id, prospect_name, prospect_email, subject, body }
    Enforces 25/day cap. Requires APOLLO_API_KEY and email account connected.
    """
    if not os.getenv('APOLLO_API_KEY'):
        return jsonify({'error': 'APOLLO_API_KEY not configured'}), 503

    body = request.get_json(force=True) or {}
    prospect_id    = str(body.get('prospect_id', '')).strip()
    prospect_email = str(body.get('prospect_email', '')).strip()
    prospect_name  = str(body.get('prospect_name', '')).strip()
    subject        = str(body.get('subject', '')).strip()
    email_body     = str(body.get('body', '')).strip()

    if not prospect_email or not subject or not email_body:
        return jsonify({'error': 'prospect_email, subject, body required'}), 400

    today = date.today().isoformat()
    log   = _load_email_log()
    sent_today = log.get(today, [])

    if len(sent_today) >= DAILY_EMAIL_CAP:
        return jsonify({
            'error': f'Daily cap of {DAILY_EMAIL_CAP} emails reached.',
            'sent':  len(sent_today),
            'cap':   DAILY_EMAIL_CAP,
        }), 429

    # Apollo: create a one-off email send via emailer_messages
    try:
        resp = requests.post(f"{_APOLLO_BASE}/emailer_messages",
            headers=_apollo_headers(),
            json={
                'contact_id':    prospect_id,
                'to':            prospect_email,
                'subject':       subject,
                'body_html':     email_body,
                'body_text':     email_body,
                'send_at':       'now',
            },
            timeout=15,
        )
        if resp.status_code not in (200, 201):
            return jsonify({'error': f'Apollo {resp.status_code}', 'detail': resp.text[:300]}), 502

        entry = {
            'prospect_id':    prospect_id,
            'prospect_name':  prospect_name,
            'prospect_email': prospect_email,
            'subject':        subject,
            'sent_at':        date.today().isoformat(),
        }
        sent_today.append(entry)
        log[today] = sent_today
        _save_email_log(log)

        return jsonify({
            'success':   True,
            'sent':      len(sent_today),
            'remaining': max(0, DAILY_EMAIL_CAP - len(sent_today)),
        })
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/apollo/save-prospect', methods=['POST'])
@require_jwt
def apollo_save_prospect():
    """Save an Apollo prospect to the Engagement Pipeline as a Prospect stage entry."""
    body = request.get_json(force=True) or {}
    name    = str(body.get('name', '')).strip()
    company = str(body.get('company', '')).strip()
    email   = str(body.get('email', '')).strip()
    title   = str(body.get('title', '')).strip()
    website = str(body.get('website', '')).strip()

    if not name or not company:
        return jsonify({'error': 'name and company required'}), 400

    try:
        wb, file_id, svc = _wb_download_fresh()
        if 'Engagement Pipeline' not in wb.sheetnames:
            return jsonify({'error': 'Engagement Pipeline tab not found'}), 500
        ws = wb['Engagement Pipeline']

        # Find last data row
        last_data_row = _EP_DATA_START - 1
        for i, row in enumerate(ws.iter_rows(min_row=_EP_DATA_START, values_only=True),
                                 start=_EP_DATA_START):
            if row and row[_EP_COL['client']]:
                last_data_row = i
        target_row = last_data_row + 1

        ws.cell(row=target_row, column=_EP_COL['client']       + 1).value = company
        ws.cell(row=target_row, column=_EP_COL['stage']        + 1).value = 'Prospect'
        ws.cell(row=target_row, column=_EP_COL['status']       + 1).value = 'Pending'
        ws.cell(row=target_row, column=_EP_COL['client_title'] + 1).value = f"{name} — {title}" if title else name
        ws.cell(row=target_row, column=_EP_COL['email']        + 1).value = email

        _wb_upload(wb, file_id, svc)
        return jsonify({'success': True, 'company': company, 'row': target_row})
    except Exception as exc:
        app.logger.error('apollo_save_prospect error: %s', exc)
        return jsonify({'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# Cron Monitor
# ---------------------------------------------------------------------------

import subprocess as _subprocess

_CRON_JOBS = [
    {'name': 'Activity Digest',       'script': 'activity_digest.py',       'log': '/home/jlobel/lac_automation/logs/activity_digest.log',      'schedule': 'Daily 8am'},
    {'name': 'Approval Queue Digest', 'script': 'approval_queue_digest.py', 'log': '/home/jlobel/lac_automation/logs/approval_queue_digest.log', 'schedule': 'Daily 8am'},
    {'name': 'Shareholder Sync',      'script': 'shareholder_sync.py',      'log': '/home/jlobel/lac_automation/logs/shareholder_sync.log',      'schedule': 'Weekly Mon'},
    {'name': 'Backup',                'script': 'backup.py',                'log': '/home/jlobel/lac_automation/backup.log',                     'schedule': 'Daily 2am'},
    {'name': 'Drive Organizer',       'script': 'drive_organizer.py',       'log': '/home/jlobel/lac_automation/drive_organizer.log',            'schedule': 'Daily 3am'},
    {'name': 'Realization Alert',     'script': 'realization_alert.py',     'log': '/home/jlobel/lac_automation/realization.log',                'schedule': 'Weekly Mon'},
    {'name': 'Weekly KPI Digest',     'script': 'weekly_kpi_digest.py',     'log': '/home/jlobel/lac_automation/phase5/digest.log',              'schedule': 'Weekly Mon'},
    {'name': 'Daily Briefing',        'script': 'daily_briefing.py',        'log': '/home/jlobel/lac_automation/phase6/daily_briefing.log',      'schedule': 'Daily 7am'},
    {'name': 'Deadline Tracker',      'script': 'deadline_tracker.py',      'log': '/home/jlobel/lac_automation/phase6/deadline_tracker.log',    'schedule': 'Daily 8am'},
    {'name': 'Regulatory Scraper',    'script': 'regulatory_scraper.py',    'log': '/home/jlobel/lac_automation/phase6/regulatory_scraper.log',  'schedule': 'Weekly Sun'},
    {'name': 'Invoice Automation',    'script': 'invoice_automation.py',    'log': '/home/jlobel/lac_automation/invoice.log',                    'schedule': 'Monthly 1st'},
    {'name': 'Clockify',              'script': 'clockify.py',              'log': '/home/jlobel/lac_automation/clockify.log',                   'schedule': 'Weekly Fri'},
    {'name': 'HubSpot Sync',          'script': 'hubspot_sync.py',          'log': '/home/jlobel/lac_automation/hubspot.log',                    'schedule': 'Daily'},
    {'name': 'Paperless Sync',        'script': 'paperless_sync.py',        'log': '/home/jlobel/lac_automation/paperless.log',                  'schedule': 'Hourly'},
    {'name': 'Cash Recon Alert',      'script': 'cash_recon_alert.py',      'log': '/home/jlobel/lac_automation/logs/cash_recon.log',              'schedule': 'Daily 2:15am'},
]

def _parse_log_tail(log_path, n_lines=8):
    """Read last n lines of a log file; return (last_modified_iso, lines_text, status)."""
    import os as _os
    if not _os.path.exists(log_path):
        return None, None, 'unknown'
    try:
        mtime = _os.path.getmtime(log_path)
        last_run_iso = date.fromtimestamp(mtime).isoformat() + 'T' + \
                       time.strftime('%H:%M:%S', time.localtime(mtime))
        with open(log_path, 'r', errors='replace') as f:
            lines = f.readlines()
        tail = ''.join(lines[-n_lines:]).strip()
        low = tail.lower()
        error_signals = [
            'error', 'exception', 'traceback', 'failed',
            'no such file', 'cannot find', 'not found',
            'permission denied', 'command not found',
            'errno', 'ioerror', 'oserror', 'filenotfounderror',
        ]
        status = 'error' if any(w in low for w in error_signals) else 'ok'
        return last_run_iso, tail or '(empty)', status
    except Exception:
        return None, None, 'unknown'


@app.route('/cron/status', methods=['GET'])
@require_jwt
def cron_status():
    try:
        # Read actual crontab
        try:
            ct = _subprocess.run(['crontab', '-l'], capture_output=True, text=True, timeout=5)
            crontab_lines = ct.stdout.splitlines() if ct.returncode == 0 else []
        except Exception:
            crontab_lines = []

        jobs = []
        for job in _CRON_JOBS:
            last_run, last_lines, status = _parse_log_tail(job['log'])
            # Try to get exit code from last log line pattern (not always available)
            schedule = job['schedule']
            # Match crontab line to get real schedule
            for line in crontab_lines:
                if job['script'] in line and not line.startswith('#'):
                    parts = line.split()
                    if len(parts) >= 5:
                        schedule = ' '.join(parts[:5])
                    break
            jobs.append({
                'name':       job['name'],
                'script':     job['script'],
                'log_file':   job['log'],
                'schedule':   schedule,
                'last_run':   last_run,
                'last_lines': last_lines,
                'exit_code':  None,
                'status':     status,
            })

        return jsonify({'jobs': jobs})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# CPE Log — Becker CSV import
# ---------------------------------------------------------------------------

_CPE_HEADERS = ['Course Name', 'Credits', 'Completed On', 'Delivery Method',
                'Field of Study', 'Provider', 'CBA Period', 'CBA Category']
_CPE_DATA_ROW = 4   # first data row in CPE Log (1-based)



@app.route('/cpe/preview', methods=['POST'])
@require_jwt
def cpe_preview():
    """Parse uploaded Becker CSV and return rows not already in the log."""
    try:
        import csv as _csv, io as _io
        f = request.files.get('file')
        if not f:
            return jsonify({'error': 'No file uploaded'}), 400

        content = f.read().decode('utf-8-sig', errors='replace')
        reader  = _csv.DictReader(_io.StringIO(content))
        raw_headers = reader.fieldnames or []

        # Normalise header → index mapping (case-insensitive, strip whitespace)
        def _norm(s): return s.strip().lower() if s else ''
        nh = {_norm(h): h for h in raw_headers}

        # Map Becker column names → CPE Log columns
        col_map = {
            'Course Name':      next((nh[k] for k in nh if 'course' in k and 'name' in k or k == 'title'), None),
            'Credits':          next((nh[k] for k in nh if 'credit' in k), None),
            'Completed On':     next((nh[k] for k in nh if 'complet' in k and ('date' in k or 'on' in k)), None),
            'Delivery Method':  next((nh[k] for k in nh if 'delivery' in k or 'method' in k), None),
            'Field of Study':   next((nh[k] for k in nh if 'field' in k), None),
            'Provider':         next((nh[k] for k in nh if 'provider' in k), None),
            'CBA Period':       next((nh[k] for k in nh if 'period' in k or 'cba period' in k), None),
            'CBA Category':     next((nh[k] for k in nh if 'category' in k or 'cba cat' in k), None),
        }

        def _get(row, col): return str(row.get(col_map.get(col) or '', '') or '').strip()

        # Load existing CPE Log entries to detect duplicates
        wb_read = _get_workbook()
        existing = set()
        if 'CPE Log' in wb_read.sheetnames:
            ws_log = wb_read['CPE Log']
            for row in ws_log.iter_rows(min_row=_CPE_DATA_ROW, values_only=True):
                name = str(row[0] or '').strip().lower()
                completed = str(row[2] or '').strip()
                if name:
                    existing.add((name, completed))

        rows = []
        for r in reader:
            name       = _get(r, 'Course Name')
            completed  = _get(r, 'Completed On')
            if not name:
                continue
            if (name.lower(), completed) in existing:
                continue
            credits    = _get(r, 'Credits')
            delivery   = _get(r, 'Delivery Method')
            field      = _get(r, 'Field of Study')
            provider   = _get(r, 'Provider')
            cba_period = _get(r, 'CBA Period')
            cba_cat    = _get(r, 'CBA Category')
            rows.append({
                'course_name':     name,
                'credits':         credits,
                'completed_on':    completed,
                'delivery_method': delivery,
                'field_of_study':  field,
                'provider':        provider,
                'cba_period':      cba_period,
                'cba_category':    cba_cat,
            })

        return jsonify({'rows': rows, 'existing_count': len(existing)})
    except Exception as exc:
        app.logger.error('cpe_preview error: %s', exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/cpe/import', methods=['POST'])
@require_jwt
def cpe_import():
    """Write confirmed rows to CPE Log tab and push to Drive."""
    try:
        data = request.get_json(silent=True) or {}
        rows = data.get('rows', [])
        if not rows:
            return jsonify({'error': 'No rows to import'}), 400

        wb, fid, svc = _wb_download_writable()
        if 'CPE Log' not in wb.sheetnames:
            return jsonify({'error': 'CPE Log sheet not found'}), 404

        ws = wb['CPE Log']

        # Find first empty row at or after _CPE_DATA_ROW
        next_row = _CPE_DATA_ROW
        while ws.cell(row=next_row, column=1).value:
            next_row += 1

        added = 0
        for r in rows:
            rn = next_row
            ws.cell(row=rn, column=1).value  = r.get('course_name', '')
            ws.cell(row=rn, column=2).value  = _safe_float(r.get('credits'))
            ws.cell(row=rn, column=3).value  = r.get('completed_on', '')
            ws.cell(row=rn, column=4).value  = r.get('delivery_method', '')
            ws.cell(row=rn, column=5).value  = r.get('field_of_study', '')
            ws.cell(row=rn, column=6).value  = r.get('provider', '')
            ws.cell(row=rn, column=7).value  = r.get('cba_period', '')
            ws.cell(row=rn, column=8).value  = r.get('cba_category', '')
            # ACFE Eligible: Yes whenever a course name is present
            ws.cell(row=rn, column=9).value  = f'=IF(A{rn}<>"","Yes","")'
            # Fraud Topic: Yes if "fraud" appears in course name or field of study
            ws.cell(row=rn, column=10).value = (
                f'=IF(OR(ISNUMBER(SEARCH("fraud",A{rn})),'
                f'ISNUMBER(SEARCH("fraud",E{rn}))),"Yes","No")'
            )
            next_row += 1
            added += 1

        _wb_upload(wb, fid, svc)
        _wb_cache['wb'] = None
        _wb_cache['fetched_at'] = 0
        return jsonify({'added': added})
    except Exception as exc:
        app.logger.error('cpe_import error: %s', exc)
        return jsonify({'error': str(exc)}), 500


def _safe_float(v):
    try: return float(v)
    except (TypeError, ValueError): return v


# ---------------------------------------------------------------------------
# CPE Compliance Dashboard
# ---------------------------------------------------------------------------

@app.route('/data/cpe-compliance', methods=['GET'])
@require_jwt
def cpe_compliance():
    try:
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(os.getenv('LAC_WORKBOOK', '/home/jlobel/lac_automation/LAC_Workbook.xlsx'), data_only=True)
        if 'Compliance Dashboard' not in wb.sheetnames:
            return jsonify({'error': 'Compliance Dashboard sheet not found'}), 404
        ws = wb['Compliance Dashboard']

        sections = []
        current_section = None

        for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
            a = row[0]
            if a is None:
                continue
            a_str = str(a).strip()
            if not a_str:
                continue

            # Section header rows (no numeric columns)
            b = row[1]
            if b is None and a_str not in ('Requirement',) and not a_str.startswith('Current Period') and not a_str.startswith('Calendar Year') and not a_str.startswith('INSTRUCTIONS'):
                current_section = {'title': a_str, 'period': None, 'rows': []}
                sections.append(current_section)
                continue

            # Period/subheader row
            if a_str.startswith('Current Period') or a_str.startswith('Calendar Year'):
                if current_section:
                    current_section['period'] = a_str
                continue

            # Skip header row and instructions
            if a_str in ('Requirement',) or a_str.startswith('INSTRUCTIONS'):
                continue

            # Data row
            if current_section is not None and b is not None:
                try:
                    required  = float(b) if b is not None else 0
                    earned    = float(row[2]) if row[2] is not None else 0
                    remaining = float(row[3]) if row[3] is not None else 0
                    pct       = float(row[4]) if row[4] is not None else 0
                    status    = str(row[5]).strip() if row[5] is not None else ''
                    current_section['rows'].append({
                        'requirement': a_str,
                        'required':    required,
                        'earned':      earned,
                        'remaining':   remaining,
                        'pct':         round(pct * 100, 1) if pct <= 1 else round(pct, 1),
                        'status':      status,
                    })
                except (ValueError, TypeError):
                    continue

        return jsonify({'sections': sections})
    except Exception as exc:
        app.logger.error('cpe_compliance error: %s', exc)
        return jsonify({'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# Journal Entries
# ---------------------------------------------------------------------------

@app.route('/data/accounts', methods=['GET'])
@require_jwt
def get_accounts():
    """Return chart of accounts from Trial Balance tab."""
    try:
        wb = _workbook()
        if 'Trial Balance' not in wb.sheetnames:
            return jsonify({'error': 'Trial Balance sheet not found'}), 404
        ws = wb['Trial Balance']
        accounts = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                break
            try:
                num = int(float(row[0]))
            except (TypeError, ValueError):
                continue
            name     = str(row[1] or '').strip()
            acct_type = str(row[2] or '').strip()
            if num and name:
                accounts.append({'num': num, 'name': name, 'type': acct_type})
        return jsonify({'accounts': accounts})
    except Exception as exc:
        app.logger.error('get_accounts error: %s', exc)
        return jsonify({'error': str(exc)}), 500


def _parse_je_rows(ws):
    """Read Transactions tab and group rows by JE number. Returns ordered list of JEs."""
    je_map   = {}
    je_order = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        je_num = str(row[1] or '').strip()
        if not je_num:
            continue
        date_val = row[0]
        if isinstance(date_val, (datetime, date)):
            date_str = (date_val if isinstance(date_val, datetime) else datetime(date_val.year, date_val.month, date_val.day)).strftime('%Y-%m-%d')
        elif isinstance(date_val, str):
            d = _to_date(date_val)
            date_str = d.strftime('%Y-%m-%d') if d else date_val
        elif isinstance(date_val, (int, float)):
            try:
                d = _EXCEL_EPOCH + timedelta(days=int(date_val))
                date_str = d.strftime('%Y-%m-%d')
            except Exception:
                date_str = ''
        else:
            date_str = ''
        line = {
            'acct_num':  int(float(row[3])) if row[3] is not None else None,
            'acct_name': str(row[4] or '').strip(),
            'debit':     float(row[5]) if row[5] is not None else None,
            'credit':    float(row[6]) if row[6] is not None else None,
            'notes':     str(row[7] or '').strip(),
        }
        if je_num not in je_map:
            je_map[je_num] = {
                'je_num':      je_num,
                'date':        date_str,
                'description': str(row[2] or '').strip(),
                'lines':       [],
            }
            je_order.append(je_num)
        je_map[je_num]['lines'].append(line)
    return [je_map[j] for j in je_order]


@app.route('/data/journal', methods=['GET'])
@require_jwt
def get_journal():
    """Return journal entries grouped by JE number, newest first."""
    try:
        limit = request.args.get('limit')
        wb = _workbook()
        if 'Transactions' not in wb.sheetnames:
            return jsonify({'entries': []})
        ws = wb['Transactions']
        entries = list(reversed(_parse_je_rows(ws)))
        if limit:
            try:
                entries = entries[:int(limit)]
            except ValueError:
                pass
        return jsonify({'entries': entries})
    except Exception as exc:
        app.logger.error('get_journal error: %s', exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/journal/add', methods=['POST'])
@require_jwt
def journal_add():
    """Add a new double-entry journal entry to the Transactions tab."""
    try:
        data    = request.get_json(silent=True) or {}
        je_date = (data.get('date') or '').strip()
        je_desc = (data.get('description') or '').strip()
        lines   = data.get('lines', [])

        if not je_date or not je_desc:
            return jsonify({'error': 'date and description are required'}), 400
        if len(lines) < 2:
            return jsonify({'error': 'At least 2 lines required for double-entry'}), 400

        total_debit  = sum(float(l.get('debit')  or 0) for l in lines)
        total_credit = sum(float(l.get('credit') or 0) for l in lines)
        if abs(total_debit - total_credit) > 0.005:
            return jsonify({'error': f'Entry not balanced: debits {total_debit:.2f} ≠ credits {total_credit:.2f}'}), 400

        wb, fid, svc = _wb_download_writable()
        if 'Transactions' not in wb.sheetnames:
            return jsonify({'error': 'Transactions sheet not found'}), 404
        ws = wb['Transactions']

        # Determine next JE number
        max_je = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or row[1] is None:
                continue
            m = re.match(r'JE-(\d+)', str(row[1]).strip(), re.I)
            if m:
                max_je = max(max_je, int(m.group(1)))
        next_je = f'JE-{max_je + 1:03d}'

        # Find first empty row
        next_row = 3
        while ws.cell(row=next_row, column=1).value is not None:
            next_row += 1

        try:
            d = datetime.strptime(je_date, '%Y-%m-%d')
        except ValueError:
            d = _to_date(je_date)
            if d is None:
                return jsonify({'error': 'Invalid date format'}), 400
            d = datetime(d.year, d.month, d.day)

        for line in lines:
            ws.cell(row=next_row, column=1).value = d
            ws.cell(row=next_row, column=2).value = next_je
            ws.cell(row=next_row, column=3).value = je_desc
            ws.cell(row=next_row, column=4).value = int(line['acct_num'])
            ws.cell(row=next_row, column=5).value = str(line.get('acct_name', ''))
            ws.cell(row=next_row, column=6).value = float(line['debit'])  if line.get('debit')  else None
            ws.cell(row=next_row, column=7).value = float(line['credit']) if line.get('credit') else None
            ws.cell(row=next_row, column=8).value = str(line.get('notes', ''))
            ws.cell(row=next_row, column=9).value = d.month
            next_row += 1

        _wb_upload(wb, fid, svc)
        _wb_cache['wb'] = None
        _wb_cache['fetched_at'] = 0

        return jsonify({'je_num': next_je, 'lines_added': len(lines)})
    except Exception as exc:
        app.logger.error('journal_add error: %s', exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/journal/update', methods=['PUT'])
@require_jwt
def journal_update():
    """Replace all lines for an existing JE number."""
    try:
        data    = request.get_json(silent=True) or {}
        je_num  = (data.get('je_num') or '').strip()
        je_date = (data.get('date') or '').strip()
        je_desc = (data.get('description') or '').strip()
        lines   = data.get('lines', [])

        if not je_num:
            return jsonify({'error': 'je_num required'}), 400
        if len(lines) < 2:
            return jsonify({'error': 'At least 2 lines required'}), 400

        total_debit  = sum(float(l.get('debit')  or 0) for l in lines)
        total_credit = sum(float(l.get('credit') or 0) for l in lines)
        if abs(total_debit - total_credit) > 0.005:
            return jsonify({'error': f'Entry not balanced: {total_debit:.2f} ≠ {total_credit:.2f}'}), 400

        wb, fid, svc = _wb_download_writable()
        if 'Transactions' not in wb.sheetnames:
            return jsonify({'error': 'Transactions sheet not found'}), 404
        ws = wb['Transactions']

        rows_to_delete = [row[0].row for row in ws.iter_rows(min_row=3) if str(row[1].value or '').strip() == je_num]
        if not rows_to_delete:
            return jsonify({'error': f'{je_num} not found'}), 404

        insert_at = min(rows_to_delete)
        for rn in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(rn)

        try:
            d = datetime.strptime(je_date, '%Y-%m-%d')
        except (ValueError, TypeError):
            d = _to_date(je_date)
            if d is None:
                return jsonify({'error': 'Invalid date'}), 400
            d = datetime(d.year, d.month, d.day)

        ws.insert_rows(insert_at, amount=len(lines))
        for i, line in enumerate(lines):
            rn = insert_at + i
            ws.cell(row=rn, column=1).value = d
            ws.cell(row=rn, column=2).value = je_num
            ws.cell(row=rn, column=3).value = je_desc
            ws.cell(row=rn, column=4).value = int(line['acct_num'])
            ws.cell(row=rn, column=5).value = str(line.get('acct_name', ''))
            ws.cell(row=rn, column=6).value = float(line['debit'])  if line.get('debit')  else None
            ws.cell(row=rn, column=7).value = float(line['credit']) if line.get('credit') else None
            ws.cell(row=rn, column=8).value = str(line.get('notes', ''))
            ws.cell(row=rn, column=9).value = d.month

        _wb_upload(wb, fid, svc)
        _wb_cache['wb'] = None
        _wb_cache['fetched_at'] = 0

        return jsonify({'updated': je_num, 'lines': len(lines)})
    except Exception as exc:
        app.logger.error('journal_update error: %s', exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/journal/delete', methods=['DELETE'])
@require_jwt
def journal_delete():
    """Delete all rows for a given JE number."""
    try:
        data   = request.get_json(silent=True) or {}
        je_num = (data.get('je_num') or '').strip()
        if not je_num:
            return jsonify({'error': 'je_num required'}), 400

        wb, fid, svc = _wb_download_writable()
        if 'Transactions' not in wb.sheetnames:
            return jsonify({'error': 'Transactions sheet not found'}), 404
        ws = wb['Transactions']

        rows_to_delete = [row[0].row for row in ws.iter_rows(min_row=3) if str(row[1].value or '').strip() == je_num]
        if not rows_to_delete:
            return jsonify({'error': f'{je_num} not found'}), 404

        for rn in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(rn)

        _wb_upload(wb, fid, svc)
        _wb_cache['wb'] = None
        _wb_cache['fetched_at'] = 0

        return jsonify({'deleted': je_num, 'rows': len(rows_to_delete)})
    except Exception as exc:
        app.logger.error('journal_delete error: %s', exc)
        return jsonify({'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# Accounting News Feed
# ---------------------------------------------------------------------------

@app.route('/data/news', methods=['GET'])
@require_jwt
def accounting_news():
    """Fetch and parse accounting news RSS feeds."""
    import xml.etree.ElementTree as ET
    import html as _html
    import urllib.request as _ur

    FEEDS = [
        {'source': 'Journal of Accountancy', 'url': 'https://www.journalofaccountancy.com/rss/all.rss'},
        {'source': 'Accounting Today',        'url': 'https://www.accountingtoday.com/feed'},
        {'source': 'CPA Practice Advisor',    'url': 'https://www.cpapracticeadvisor.com/rss'},
        {'source': 'Going Concern',            'url': 'https://goingconcern.com/feed/'},
    ]

    items = []
    for feed in FEEDS:
        try:
            req = _ur.Request(feed['url'], headers={'User-Agent': 'LAC-Dashboard/1.0 (+https://dashboard.lobelaccountancy.com)'})
            with _ur.urlopen(req, timeout=6) as resp:
                content = resp.read()
            root = ET.fromstring(content)

            for item in root.findall('.//item')[:6]:
                title = item.findtext('title', '') or ''
                link  = item.findtext('link',  '') or ''
                pub   = item.findtext('pubDate', '') or ''
                desc  = item.findtext('description', '') or ''
                desc_clean = re.sub(r'<[^>]+>', '', _html.unescape(desc))[:220].strip()
                if title and link:
                    items.append({
                        'source':  feed['source'],
                        'title':   _html.unescape(title.strip()),
                        'url':     link.strip(),
                        'date':    pub.strip(),
                        'summary': desc_clean,
                    })
        except Exception as e:
            app.logger.warning('News feed %s failed: %s', feed['source'], e)

    return jsonify({'items': items[:40]})


# ---------------------------------------------------------------------------
# Global Search
# ---------------------------------------------------------------------------

@app.route('/search/global', methods=['GET'])
@require_jwt
def global_search():
    q = (request.args.get('q') or '').strip().lower()
    if len(q) < 2:
        return jsonify({'results': []})
    try:
        results = []
        # Search clients / pipeline from workbook
        try:
            wb = _workbook()
            client_map = _build_client_map(wb)
            for name in client_map:
                if q in name.lower():
                    results.append({'type': 'client', 'label': name, 'href': 'clients.html', 'sub': 'Client'})
            # Search pipeline descriptions
            if 'Engagement Pipeline' in wb.sheetnames:
                ws = wb['Engagement Pipeline']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    desc = str(row[1] if len(row) > 1 else '')
                    client = str(row[0])
                    if q in desc.lower() or q in client.lower():
                        label = f"{client}: {desc}" if desc else client
                        if not any(r['label'] == label for r in results):
                            results.append({'type': 'matter', 'label': label, 'href': 'engagement.html', 'sub': 'Pipeline'})
        except Exception:
            pass

        # Search Paperless documents
        try:
            import requests as _req
            pl_url  = os.getenv('PAPERLESS_URL', 'http://localhost:8000')
            pl_key  = os.getenv('PAPERLESS_API_KEY', '')
            r = _req.get(f'{pl_url}/api/documents/', params={'query': q, 'page_size': 5},
                         headers={'Authorization': f'Token {pl_key}'}, timeout=4)
            if r.ok:
                for doc in r.json().get('results', []):
                    results.append({
                        'type': 'document',
                        'label': doc.get('title', 'Untitled'),
                        'href':  'docs.html',
                        'sub':   'Document',
                    })
        except Exception:
            pass

        return jsonify({'results': results[:20], 'query': q})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


if __name__ == '__main__':
    port = int(os.getenv('AUTH_PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=False)
